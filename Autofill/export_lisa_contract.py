"""
export_lisa_contract.py — freeze LISA's real input contract from the workbook.

LISA generates finished IDPs only if every FillIndex row obeys the workbook's own
dropdown logic:
  - Type      in named range  Type_<WireCt>
  - S Symbol  in named range  KEY(Type)_<WireCt>_L
  - D Symbol  in named range  KEY(Type)_<WireCt>_R
  where KEY(Type) = Type with dashes removed (Excel SUBSTITUTE(type,"-","")).
  - each symbol activates exactly BlockTags[symbol] Tag/Term slots (rest grey out).

This script resolves those named ranges + BlockTags out of the template workbook
and writes Handoff/lisa_symbols.json (the LISA contract the extractor loads).
Re-run whenever the workbook's block library changes.

    python export_lisa_contract.py [path-to-workbook.xlsm]
"""
from __future__ import annotations

import json
import os
import re
import sys

import openpyxl
from openpyxl.utils import range_boundaries

_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_WB = os.path.join(
    _HERE, "..", "Excel template", "IDP_Workbook_CurrentWIP_3 - Claude Copy.xlsm"
)
OUT = os.path.join(_HERE, "Handoff", "lisa_symbols.json")

SYM_RE = re.compile(r"^(.+)_(\d+)_([LR])$")   # KEY_<ct>_<L|R>
TYPE_RE = re.compile(r"^Type_(\d+)$")


def key_of(type_name: str) -> str:
    """Excel SUBSTITUTE(type,'-','') — the named-range key transform."""
    return str(type_name).replace("-", "")


def _resolve(wb, name):
    dn = wb.defined_names[name]
    out = []
    for title, coord in dn.destinations:
        ws = wb[title]
        c0, r0, c1, r1 = range_boundaries(coord.replace("$", ""))
        # full-column / full-row refs leave a bound as None — clamp to the sheet
        r0 = r0 or 1
        r1 = r1 or ws.max_row
        c0 = c0 or 1
        c1 = c1 or ws.max_column
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    out.append(str(v).strip())
    return out


def build(path):
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    names = list(wb.defined_names.keys())

    types = {}
    for n in names:
        m = TYPE_RE.match(n)
        if m:
            types[int(m.group(1))] = _resolve(wb, n)

    symbols = {}   # KEY -> ct -> {"L":[...], "R":[...]}
    for n in names:
        m = SYM_RE.match(n)
        if not m:
            continue
        k, ct, side = m.group(1), int(m.group(2)), m.group(3)
        symbols.setdefault(k, {}).setdefault(str(ct), {"L": [], "R": []})[side] = _resolve(wb, n)

    blocktags = {}
    if "BlockTags" in wb.sheetnames:
        ws = wb["BlockTags"]
        for r in range(1, ws.max_row + 1):
            nm = ws.cell(row=r, column=1).value
            ct = ws.cell(row=r, column=2).value
            if nm not in (None, ""):
                try:
                    blocktags[str(nm).strip()] = int(ct)
                except (TypeError, ValueError):
                    pass

    colors = {}
    for n in ("WireColor", "Color_1", "Color_2", "Color_3", "Color_4"):
        if n in names:
            colors[n] = _resolve(wb, n)

    contract = {
        "_comment": "LISA INPUT CONTRACT — the workbook's own dropdown universe. "
                    "Every FillIndex row must obey this or LISA cannot generate the IDP. "
                    "Regenerate with export_lisa_contract.py when the block library changes.",
        "source_workbook": os.path.basename(path),
        "key_transform": "KEY(type) = type.replace('-','')",
        "type_by_wire_ct": types,
        "symbols_by_key_ct_side": symbols,
        "blocktags_slot_count": blocktags,
        "colors": colors,
        "special_cell_value": "Hide from Generation",
    }
    return contract


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    path = os.path.abspath(path)
    contract = build(path)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(contract, fh, indent=2)
    n_keys = len(contract["symbols_by_key_ct_side"])
    n_syms = len(contract["blocktags_slot_count"])
    print(f"wrote {OUT}")
    print(f"  type buckets: {sorted(contract['type_by_wire_ct'].keys())}")
    print(f"  symbol keys : {sorted(contract['symbols_by_key_ct_side'].keys())}")
    print(f"  blocktags   : {n_syms} symbols with slot counts")


if __name__ == "__main__":
    main()
