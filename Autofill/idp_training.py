"""
idp_training.py — the gap-bridging training loop.

Compares, per conduit, what our workbook produced ("ours") against the finished
IDP drawings ("ground truth"), turns each mismatch into a concrete correction,
and folds confirmed-safe corrections back into the durable stores (Remembered
Logic + KB) AND into the Claude skills' reference files so both the exe and the
skill improve together. Whatever it is NOT confident about is collected for
escalation to Claude (see idp_escalate).

Inputs (any may be omitted):
  plans      : source docs the workbook was built from (context; optional)
  finished   : finished IDP .dwg files / folders  -> ground truth
  generated  : our filled workbook (.xlsm) OR the DWGs LISA generated from it

run_training(...) returns a report dict:
  {conduits_compared, gaps:[...], learned:[...], uncertain:[...], summary}
"""
from __future__ import annotations

import os

import idp_project
import idp_dwg_extract
try:
    import idp_project_symbols
except Exception:
    idp_project_symbols = None
try:
    import idp_excel
except Exception:
    idp_excel = None


# ── load each side into comparable records ──────────────────────────────────
def _ground_truth_records(finished_paths):
    """Finished IDP ground truth -> records. Handles BOTH finished-IDP PDFs (AIC's submittal
    packages, parsed by idp_idp_pdf — this is how finished IDPs are usually delivered) AND
    finished DWGs (scanned, with the template-palette blocks filtered out so fill counts
    reflect REAL per-conduit usage, not the legend)."""
    if not finished_paths:
        return []
    # 1) finished-IDP PDFs first — a picked PDF, or the IDP-package PDFs inside a picked folder
    try:
        import idp_idp_pdf
        pdfs = []
        for p in finished_paths:
            if os.path.isdir(p):
                for dp, _dn, fns in os.walk(p):
                    for fn in fns:
                        if fn.lower().endswith(".pdf"):
                            pdfs.append(os.path.join(dp, fn))
            elif str(p).lower().endswith(".pdf"):
                pdfs.append(p)
        pdf_recs = []
        idp_pdfs = []
        for pdf in pdfs:
            try:
                if idp_idp_pdf.is_idp_package(pdf):
                    idp_pdfs.append(pdf)
                    pdf_recs += idp_idp_pdf.read_source(pdf)
            except Exception:
                continue
        # GRAPHICAL finished IDP (e.g. AIC interconnection diagrams): the text layer barely
        # reads, so idp_idp_pdf recovers only a handful. OCR each sheet to recover the FULL
        # conduit set, so training compares against real ground truth instead of a fragment.
        # Only fires when the text read was sparse, and only adopted if it recovers MORE.
        if idp_pdfs and len(pdf_recs) < 12:
            try:
                import idp_idp_ocr
                ocr_recs = []
                for pdf in idp_pdfs:
                    ocr_recs += idp_idp_ocr.read_interconnection_idp(pdf)[0]
                if len(ocr_recs) > len(pdf_recs):
                    pdf_recs = ocr_recs
            except Exception:
                pass
        if pdf_recs:
            return pdf_recs
    except Exception:
        pass
    # 2) finished DWGs (fallback — only when no finished-IDP PDF ground truth was found)
    if idp_project_symbols is None:
        return []
    root = idp_project.project_root(finished_paths)
    dwgs = idp_project_symbols.find_project_dwgs(root) if root else []
    if not dwgs:
        return []
    scan = idp_project_symbols.load_or_scan(root, dwgs)
    if not scan:
        return []
    template_xy = idp_project_symbols._template_positions(scan)
    filtered = {}
    for fn, blocks in scan.items():
        keep = []
        for b in blocks:
            xy = (round(b.get("x", 0), 1), round(b.get("y", 0), 1))
            # keep the per-sheet Conduit data block always; drop palette devices
            if b.get("name") == "Conduit" or xy not in template_xy:
                keep.append(b)
        filtered[fn] = keep
    return idp_dwg_extract.extract_from_data(filtered)


def _our_records(generated_paths):
    """Our output -> records. Accepts a filled workbook (.xlsm/.xlsx) directly, a
    folder (walked RECURSIVELY for workbooks and DWGs at any nesting depth), or
    the DWGs LISA generated from it (scanned like the finished side)."""
    # expand any folder into its nested files (all depths) so a picked folder
    # pulls workbooks/DWGs from every subfolder, not just its top level
    expanded = []
    for p in generated_paths or []:
        if os.path.isdir(p):
            for dp, _dn, fns in os.walk(p):
                for fn in fns:
                    if fn.lower().endswith((".xlsm", ".xlsx", ".xls", ".dwg")):
                        expanded.append(os.path.join(dp, fn))
            expanded.append(p)          # keep the dir too (for the DWG-root path)
        else:
            expanded.append(p)

    recs = []
    dwg_roots = []
    for p in expanded:
        ext = os.path.splitext(p)[1].lower()
        if ext in (".xlsm", ".xlsx", ".xls") and idp_excel is not None:
            try:
                recs += idp_excel.read_source(p)
            except Exception:
                continue
        elif ext == ".dwg" or os.path.isdir(p):
            dwg_roots.append(p)
    if dwg_roots and not recs:
        recs = _ground_truth_records(dwg_roots)   # same DWG->records path
    return recs


# ── the diff ────────────────────────────────────────────────────────────────
import re

def _norm(s):
    return str(s or "").strip().upper()


# Finished sets use OLDER template names; fold them to the current dropdown so a
# naming-era difference (PULL ROPE vs PULL_ROPE) isn't mistaken for a real gap.
_TYPE_CANON = {
    "PULL ROPE": "PULL_ROPE", "PULLROPE": "PULL_ROPE",
    "ETHERNET": "CAT-6", "CAT6": "CAT-6",
    "MFR CABLE": "MFG_CABLE", "MFR_CABLE": "MFG_CABLE", "MFG CABLE": "MFG_CABLE",
    "GROUND": "GND",
}
# Values that mean "unresolved / placeholder", never a real, learnable value.
_PLACEHOLDER = {"", "XXX", "XXXX", "TBD", "PER SPEC", 'X"', 'X"C', "N/A"}
# Names/text the finished-DWG scan misreads as conduits (title blocks, legends,
# I/O tags, review-comment prose) — not real conduit runs.
_JUNK_RE = re.compile(r"(DWG|REVIEW|COMMENT|LEGACY|STATUS OPEN|\bAIC\b|�)", re.I)


def _canon_type(t):
    n = _norm(t)
    return _TYPE_CANON.get(n, n)


def _is_real_conduit(name):
    """A real conduit tag is short and not a placeholder/scan-artifact string.
    Real tags look like C002 / H008A / A051 / C019 (E) — a compact alnum core
    with an optional letter or ' (X)' suffix. Reject prose, I/O channel tags, and
    comma/space/'#'-laden fragments the finished-DWG scan misreads as conduits."""
    n = str(name or "").strip()
    if not n or len(n) > 24:
        return False
    u = n.upper()
    if u in _PLACEHOLDER or re.fullmatch(r'X+"?C?', u) or re.fullmatch(r"\d+", u):
        return False
    if "," in n or "#" in n or _JUNK_RE.search(n):
        return False
    # at most one space, and only as a ' (X)'-style suffix
    if n.count(" ") > 1 or (" " in n and not re.search(r"\s\([^)]+\)$", n)):
        return False
    # PLC I/O channel references (PLC-A1, PLC-D10, PLC-EC-DI#2-A1) are wiring
    # points, not conduit runs
    if re.match(r"(?i)^PLC-", n):
        return False
    return True


def _fill_signature(rec):
    """A comparable per-conduit summary of the fill: list of (type, wire_ct,
    s_symbol, d_symbol) tuples, order-independent. Type is canonicalized to the
    current template so old-naming finished sets don't produce phantom gaps."""
    sig = []
    for g in rec.get("fill", []) or []:
        sig.append((_canon_type(g.get("type")),
                    g.get("wire_ct") or g.get("count") or 1,
                    _norm(g.get("s_symbol")), _norm(g.get("d_symbol"))))
    return sig


def diff_records(ground, ours):
    """Conduit-by-conduit gaps between ground-truth and our records.
    Each gap: {conduit, field, ground, ours, kind}."""
    # only compare real conduit runs — drop scan artifacts (title blocks, legends,
    # I/O tags, review-comment prose) that appear on only one side as noise.
    g_by = {_norm(r.get("name")): r for r in ground or [] if _is_real_conduit(r.get("name"))}
    o_by = {_norm(r.get("name")): r for r in ours or [] if _is_real_conduit(r.get("name"))}
    gaps = []
    for name in sorted(set(g_by) | set(o_by)):
        g, o = g_by.get(name), o_by.get(name)
        disp = (g or o).get("name")
        if g and not o:
            gaps.append({"conduit": disp, "field": "conduit", "kind": "missing_in_ours",
                        "ground": "present", "ours": "absent"})
            continue
        if o and not g:
            gaps.append({"conduit": disp, "field": "conduit", "kind": "extra_in_ours",
                        "ground": "absent", "ours": "present"})
            continue
        # conduit-level — skip when either side is a placeholder (unresolved, not
        # a real disagreement) or, for type, an old-naming equivalent of ours.
        for fld, key, canon in (("conduit_type", "ctype", True),
                                ("conduit_size", "size", False)):
            gv = _canon_type(g.get(key)) if canon else _norm(g.get(key))
            ov = _canon_type(o.get(key)) if canon else _norm(o.get(key))
            if (gv and ov and gv != ov
                    and gv not in _PLACEHOLDER and ov not in _PLACEHOLDER):
                gaps.append({"conduit": disp, "field": fld, "kind": "value_mismatch",
                            "ground": g.get(key), "ours": o.get(key)})
        # fill-level (counts + per-row type/symbol)
        gsig, osig = _fill_signature(g), _fill_signature(o)
        # A real conduit carries a handful of fills; an absurd ground-side count
        # is the finished-DWG reference palette miscounted, not a real gap.
        if len(gsig) != len(osig) and max(len(gsig), len(osig)) <= 20:
            gaps.append({"conduit": disp, "field": "fill_count", "kind": "count_mismatch",
                        "ground": len(gsig), "ours": len(osig)})
        for i in range(min(len(gsig), len(osig))):
            gt, gct, gs, gd = gsig[i]
            ot, oct_, os_, od = osig[i]
            # ground-encoding equivalence: the current template has no GROUND
            # type — ground is carried as a POWER/Ct1 fill with the GND symbol
            # (ensure_ground). So finished 'GROUND' (canon GND) vs our 'POWER' is
            # the same conductor, not a real gap.
            ground_pair = {gt, ot} == {"GND", "POWER"}
            if gt and ot and gt != ot and not ground_pair:
                gaps.append({"conduit": disp, "field": f"fill[{i}].type", "kind": "value_mismatch",
                            "ground": gt, "ours": ot})
            if gs and os_ and gs != os_:
                gaps.append({"conduit": disp, "field": f"fill[{i}].s_symbol", "kind": "symbol_mismatch",
                            "ground": gsig[i][2], "ours": osig[i][2]})
            if gd and od and gd != od:
                gaps.append({"conduit": disp, "field": f"fill[{i}].d_symbol", "kind": "symbol_mismatch",
                            "ground": gsig[i][3], "ours": osig[i][3]})
    return gaps


# ── learn from confirmed-safe gaps ──────────────────────────────────────────
def _learn_from_gaps(ground, gaps):
    """Fold high-confidence corrections into Remembered Logic. A symbol mismatch
    where the ground-truth block is legal in the current template AND we can tie
    it to a device name is a durable device->symbol rule. Returns learned rules."""
    if idp_project_symbols is None:
        return []
    legal = idp_project_symbols._legal_symbols()
    g_by = {_norm(r.get("name")): r for r in ground or []}
    learned, seen = [], set()
    import logic_store
    data = logic_store.load()
    existing = {_norm(r.get("match")) for r in data.get("rules", [])}
    for gap in gaps:
        if gap["kind"] != "symbol_mismatch":
            continue
        rec = g_by.get(_norm(gap["conduit"]))
        if not rec:
            continue
        side = "L" if gap["field"].endswith("s_symbol") else "R"
        gt_sym = gap["ground"]
        if not gt_sym or (legal and gt_sym not in legal):
            continue   # ground-truth uses an older/illegal name -> don't teach globally
        base = idp_project_symbols._strip_side(gt_sym)
        dev = (rec.get("source") or [""])[0] if side == "L" else (rec.get("dest") or [""])[0]
        key = _norm(dev)
        if not dev or key in existing or key in seen:
            continue
        seen.add(key)
        rule = {"type": "symbol_keyword", "match": dev, "result": base, "context": "",
                "note": f"learned from finished-IDP diff ({rec.get('name')}, {side} side): "
                        f"our symbol was wrong, finished drawing uses {gt_sym}"}
        data.setdefault("rules", []).append(rule)
        learned.append(rule)
    if learned:
        logic_store.save(data)
    return learned


def _uncertain_from_gaps(gaps):
    """Gaps we can't safely auto-learn -> escalate to Claude. Type/size mismatches
    and count mismatches need judgment; symbol gaps we couldn't tie to a legal
    block also qualify."""
    out = []
    for gap in gaps:
        if gap["kind"] in ("value_mismatch", "count_mismatch", "missing_in_ours"):
            out.append(gap)
    return out


# ── skill feedback ──────────────────────────────────────────────────────────
import sys


def _find_skills_dirs():
    """Locate EVERY Claude skills folder holding idp-fillindex/ — there are two
    that matter and they must never diverge: (1) the project tree's `skills/`
    (the source/distribution copy), found by walking up from this module or the
    exe; (2) the INSTALLED `~/.claude/skills` that the Skill tool actually runs.
    Earlier this returned only #1, so exported conventions never reached the
    running skills. Returns a de-duped list of existing skills dirs."""
    found, seen = [], set()

    def _add(cand):
        cand = os.path.abspath(cand)
        if cand not in seen and os.path.isdir(os.path.join(cand, "idp-fillindex")):
            seen.add(cand)
            found.append(cand)

    starts = [os.path.dirname(os.path.abspath(__file__))]
    if getattr(sys, "frozen", False):
        starts.insert(0, os.path.dirname(os.path.abspath(sys.executable)))
    for start in starts:
        d = start
        for _ in range(6):   # walk up a few levels
            _add(os.path.join(d, "skills"))
            parent = os.path.dirname(d)
            if parent == d:
                break
            d = parent
    # the installed skills the Skill tool loads
    _add(os.path.join(os.path.expanduser("~"), ".claude", "skills"))
    if not found:   # last resort so a first-ever export still lands somewhere
        found.append(os.path.abspath(os.path.join(starts[-1], "..", "skills")))
    return found


# backward-compat shim: some callers/tests referenced the single-dir helper
def _find_skills_dir():
    dirs = _find_skills_dirs()
    return dirs[0] if dirs else ""


_SKILLS_DIRS = _find_skills_dirs()
_SKILLS_DIR = _SKILLS_DIRS[0] if _SKILLS_DIRS else ""

# Conventions confirmed by comparing our generated output to finished IDP
# drawings (the plans↔finished↔generated training loop). These are curated,
# durable rules the Claude skills should apply — captured in each skill so it
# improves in tandem with the exe. CRITICAL: each convention is filed under the
# skill whose SCHEMA / END GOAL it serves. FillIndex owns wire fill (Type, Wire
# Ct, Color, S/D Symbol); ConduitExtractor owns conduit metadata (count,
# Conduit Type, source/destination). A convention only ever goes to its domain —
# the FillIndex skill must not carry conduit rules, and vice versa.

# ── FillIndex-domain conventions (wire fill / symbols) ───────────────────────
FILL_CONVENTIONS = [
    ("Populate the SUPPORTING DOCUMENTS table (Ref Documents & Deviations)", "Every "
     "conduit on a finished IDP cites the source drawings it was derived from, in the "
     "SUPPORTING DOCUMENTS block (DRAWING NUMBER / DESCRIPTION / MANUFACTURER). These "
     "live once, deduped, on the **Ref Documents & Deviations** sheet (A=index#, "
     "B=DWG#, C=Description, D=Manufacturer); each ConduitIndex row lists the matching "
     "index numbers in col J. Populate them, don't leave the template's prior-project "
     "refs (clear stale rows first). The set per conduit: (1) ALWAYS the design plans' "
     "conduit schedule / single-line the fill came from — DWG# = that sheet (e.g. "
     "`E-02`/`E-3`), MANUFACTURER = the owner/engineer of record (e.g. `STANISLAUS "
     "COUNTY`) read from the plan title block; (2) the AIC EDC drawing for the "
     "conduit's devices, BY TYPE — power feeder → `THREE-LINE DIAGRAM`, control/analog "
     "→ `PLC I/O & CONTROL WIRING DIAGRAM`, panelboard branch → `PANELBOARD SCHEDULE` "
     "— with the real `73.xxxx-nn` number pulled from the EDC drawing index when it is "
     "in scope (e.g. 73.1188-05 THREE-LINE MSB1 & ATS1, 73.1188-91 MCC1); (3) any "
     "manufacturer CUT SHEET from the submittals that defines a device's terminals. "
     "Leave a DWG# blank + flag rather than fabricate a number. The extractor "
     "auto-applies (1)+(2)-by-type via `idp_terms.apply_supporting_docs`; pass "
     "`schedule_doc=(dwg,desc,manu)` for the real plan sheet, and fill EDC numbers "
     "from the index. [73.1188 Well2022 every sheet: E-02 STANISLAUS COUNTY + "
     "73.1188-05 AIC three-line.]"),
    ("Ground is classified by the S/D Tag = \"GND\", not the wire Type", "LISA marks a "
     "conductor as GROUND in the conduit fill table ONLY when its S Tag or D Tag reads "
     "\"GND\" (workbook_mapper._derive_fill_slots → _tag_is_gnd); the GND_L/GND_R symbol "
     "and GRN color alone are NOT enough — without the tag LISA renders the green "
     "conductor as POWER and the GND block shows its default placeholder tag. So the "
     "ground conductor must carry **S Tag = D Tag = \"GND\"** with a BLANK terminal (GND "
     "is a tag, not a terminal number). Blanking the term also keeps the device-tag pass "
     "(CB-MAIN/CB-GEN, ATS N/E/L) from landing on the ground row. The extractor now does "
     "this for every ground — a separate GND group, a slot resolving to GND, or any GRN "
     "conductor (`idp_terms._is_ground_group` + `apply_source_info`). Verified through "
     "LISA's own parser: H001/H002/H005 now read 3×POWER + 1×GROUND."),
    ("Every real power feeder gets a ground — no false 'already grounded'", "Verifying "
     "Crows 3 against LISA exposed feeders LEFT UNGROUNDED: (a) motor feeders to a "
     "`MTR_3PH_NoGND` block — the ground detector matched the substring 'GND' inside "
     "'NoGND' and thought a ground already existed (fixed: strip NO[_ ]GND, then match "
     "GND as a token); (b) panel feeders that already carry EDC/panelboard circuit "
     "terminals — the source-info pass skipped them, so the ground group's tag never "
     "became 'GND'. The writer now forces S Tag = D Tag = 'GND' (blank term) on EVERY "
     "ground group unconditionally. Result on 73.1188: 34/34 POWER feeders grounded, 0 "
     "missing. Pull-rope / fiber / empty conduits still correctly get NO ground."),
    ("PLC I/O channels are the control/analog terminations", "The MASTER-PLC-side term "
     "on a control/analog cable is its PLC I/O channel, read from the Main PLC EDC I/O "
     "sheets by matching the field device to a point: a discrete status/interlock → a "
     "DIGITAL INPUT channel on `TBDI-0.<slot>` (e.g. smoke → `TBDI-0.02:15`, door → "
     "`TBDI-0.03:00`); an analog transmitter → an ANALOG INPUT channel on `TBAI-0.<slot>` "
     "(e.g. tank level `LT` → `TBAI-0.08:01`, VFD speed → `TBAI-0.07:02`, TSS/pH → "
     "`TBAI-0.07:00/01`); a PLC→field command → a DIGITAL OUTPUT on `TBDO-0.<slot>`. The "
     "term lands on whichever conduit END is the MASTER PLC. Label reads "
     "`MASTER PLC:TBAI-0.07:02/DEVICE:…`. Slots/channels come from the EDC PLC I/O "
     "drawings (56.1059-05/-06 DI, -10/-11 AI); match by device, never fabricate a "
     "channel. [Moccasin 56.1059 Main PLC EDC]"),
    ("kcmil sizes are MCM, not AWG", "Large conductors sized in kcmil (a bare number "
     ">=250, e.g. 350/500/750 — often written just as `4CC-350` on a cable schedule) are "
     "a DIFFERENT sizing system than AWG. Render them as `NNN MCM`, never `#350AWG`. The "
     "extractor normalizes bare kcmil gauges to `NNN MCM` (`idp_anatomy.apply_conventions`) "
     "and LISA's `_with_awg` skips the AWG suffix when the size contains MCM/KCMIL. AWG "
     "(#10, #4/0) keeps its `#`. [Moccasin 56.1059 C-001 4CC-350 → 350 MCM]"),
    ("Panelboard voltage decides 1Ø vs 3Ø — read the schedule", "Don't assume a "
     "3-conductor branch is 3-phase. A 120/240V 1Ø panelboard's branches are L/N/G "
     "(BLK/WHT/GRN), landing on ONE circuit (`PANEL-4:CKT-9/load:L`, `PANEL-4:N/load:N`); "
     "a 480V 3Ø panel's feeders are ØA/ØB/ØC(+N)+GND on a 3-pole breaker. Read the "
     "PANELBOARD SCHEDULE's VOLTAGE/PHASE header + the circuit→load map to set the "
     "panel-side S Tag = the panel and S Term = the circuit number (1Ø) or phases (3Ø). "
     "AUTOMATIC APPROXIMATION (no schedule/vision): the extractor now infers 1Ø from the "
     "ground pattern — a 2–3-wire POWER group carrying its ground INTEGRALLY (no separate "
     "'W/GND' EGC group, branch gauge) is single-phase → BLK/WHT/GRN + L1/N/G; a group "
     "with a SEPARATE green EGC is 3Ø → ØA/ØB/ØC + GRN (`idp_anatomy._is_single_phase_"
     "branch`). The exact circuit number still needs the schedule/vision. "
     "[Moccasin PANEL-4 = 120/240V 1Ø #56.1059-11; PANEL-3 = 480V 3Ø #56.1059-10]"),
    ("EDC supporting docs are selected by FILL TYPE, any tag scheme", "The AIC EDC "
     "drawing cited per conduit follows its circuit type, regardless of the conduit-tag "
     "naming (P/C/A, K-###, …): POWER feeder → THREE-LINE, POWER branch (panel circuits) "
     "→ PANELBOARD SCHEDULE, CONTROL → PLC I/O DIGITAL, TSP → PLC I/O ANALOG, CAT-6 → "
     "NETWORK LAYOUT, FIBER → FIBER OPTIC LAYOUT — plus the SFPUC/owner design sheets "
     "(one-line, conduit & cable schedules) on every conduit. `idp_terms."
     "apply_supporting_docs` now keys off the fill types (not the tag prefix). Author = "
     "the EDC firm (e.g. W. M. LYLES); real `56.1059-##` numbers come from the EDC index."),
    ("Wire labels follow the FILLWIRELABEL tool grammar, written as a LITERAL", "Each "
     "conductor's Wire Label is the terminal cross-reference the AutoCAD FILLWIRELABEL "
     "tool builds: **SrcName:SrcTAG1:SrcTerm / DstName:DstTAG1:DstTerm** (e.g. "
     "\"NEW T.I.D. XFMER:XFMR:ØA/UGPS:ØA\", \"CHUNGUS:GFR:21/CHUDCENTRAL:TBDI-0.01:CBD6\"). "
     "Rules mirrored from the LISP: the MIDDLE field is the device block's TAG1 (its "
     "baked device designator — XFMR/CB/MTR/DISC/GND from the symbol library "
     "device_default; `idp_terms.device_tag_for_symbol`), shown on EVERY conductor; the "
     "NAME is dropped when it equals the tag; blank fields collapse (join non-empty with "
     "\":\"). Names come from Src_Name01/Dst_Name01. The template stores this as a "
     "TEXTJOIN formula, but openpyxl can't evaluate formulas so the cached value was "
     "blank and LISA (data_only) fell back to printing the wire TYPE (\"POWER\"). The "
     "extractor now writes the computed LITERAL to Wire Label 1-4 (idp_write) and fills a "
     "blank S/D Tag with the symbol's device tag so both the block AND the label carry "
     "it. Hand-built FillIndexes must write literal labels, not un-recalculated formulas."),
    ("Utility service lateral carries NO separate ground (EGC)", "A utility service "
     "lateral — a utility transformer (T.I.D / SMUD / PG&E / TURLOCK / \"UTILITY\") "
     "feeding the service point (UGPS / MSB / switchboard / meter) — has NO equipment "
     "ground conductor in the raceway; grounding is done at the service via the "
     "grounding-electrode conductor. Finished IDP H2201 (T.I.D XFMER → MSB1) draws it as "
     "3×500 MCM phases, no ground. `ensure_ground` now skips these (idp_anatomy."
     "_is_utility_service) while still grounding every real feeder/motor/panel branch. "
     "This is the ONE case where a POWER feeder legitimately has no ground."),
    ("MCM/KCMIL sizes get a stray \"AWG\" suffix (LISA-side note)", "LISA's workbook_"
     "mapper._with_awg appends \"AWG\" to any size whose core starts with a digit, so a "
     "750 MCM aluminum conductor renders as \"#750 MCM ALAWG\". MCM/KCMIL and AWG are "
     "different sizing systems — AWG should not be appended to an MCM size. This is a "
     "LISA rendering issue (the extractor writes the correct \"750 MCM AL\"); the fix is "
     "one guard in _with_awg to skip when the value contains MCM/KCMIL."),
    ("SUPPORTING DOCUMENTS sheet must be named \"Ref Documents\"", "LISA finds the "
     "supporting-docs / deviation-notes sheet ONLY by the exact name \"Ref Documents\" "
     "(or \"Ref Docs\"); the template's \"Ref Documents & Deviations\" is invisible to "
     "it, so the table renders blank no matter how it's filled. The extractor renames "
     "the sheet to \"Ref Documents\" in the output workbook (template untouched). When "
     "building the FillIndex by hand, put the ref docs on a sheet named exactly that."),
    ("ATS terminals by connection role (N/E/L)", "An Automatic Transfer Switch "
     "labels its phase terminals by ROLE: Normal `NA/NB/NC/NN` (from the utility/MSB "
     "source), Emergency `EA/EB/EC/EN` (to/from the generator), Load `LA/LB/LC/LN` "
     "(to the downstream MCC/panel). Only the ATS end is prefixed; the other end "
     "keeps its phases. [73.1188 H2202 dest=NA…, H2203 src=EA…, H2204 src=LA…]. The "
     "extractor auto-applies this (`idp_terms.apply_power_terminals`) — the "
     "device-specific power terminal the EDC three-line carries."),
    ("Pull device data from submittals when needed", "When the conduit schedule / "
     "EDC leaves a value blank, mine the project SUBMITTALS: vendor CUT SHEETS give "
     "a device's terminal names/count, model, and ratings (e.g. a transmitter's "
     "loop terminals, a starter's OL/aux contacts); the `…_SCL_` Submittal Cover "
     "Letter gives project conventions and deviations (conduit material, wire-color "
     "changes, removed/added conduits); vendor quote/BOM docs confirm make/model. "
     "Match a cut sheet to a device once and apply it across every IO point that "
     "shares that make/model (flag the spread). Never fabricate — if the submittal "
     "doesn't state it, leave blank and flag."),
    ("Clean input beats OCR — and normalize OCR text", "The skill/clean path reads "
     "structured text so it has no punctuation noise; the exe OCR path picks up "
     "spacing/dash noise (`MCC1-SEC.2` vs `MCC1-SEC. 2`, em-dash `ZS-01—A`, `@EYE`). "
     "That noise — not logic — is most of the skill-vs-exe accuracy gap on a vector "
     "schedule. The OCR reader now normalizes it (dash→hyphen, letter-period-digit "
     "spacing, `@` spacing) WITHOUT touching unit names or decimals "
     "(`idp_ocr_schedule._normalize_text`). For true skill-level accuracy on a "
     "vector sheet, feed the conduit schedule as an Excel/clean source (the bridge) "
     "rather than relying on OCR; OCR is the ~95% fallback with amber-flagged cells."),
    ("A destination is a terminal-block LANDING by default", "A conduit's DESTINATION "
     "end is almost always a LANDING, not a driven device — so infer it landing-first "
     "and only make it a device when the endpoint text clearly names one. Any enclosure / "
     "panel / switchgear / PLC noun — CONTROL PANEL, CABINET, ENCLOSURE, MCC, SWBD/SWGR, "
     "PANEL/PNL, LCP/MCP/RCP/PCP, PDP/MSB, DISTRIBUTION, PLC/RIO/RTU, TERMINAL BLOCK — is a "
     "terminal-block landing (`TB_Square`), EVEN WHEN a driven load (MOTOR / PUMP / STARTER) "
     "is named alongside it (e.g. `MCC-2 (PUMP 2 STARTER)` or `PUMP CONTROL PANEL` → "
     "`TB_Square`, NOT a motor/starter — the load names what the enclosure FEEDS, not the "
     "landing). Only a bare motor/pump LOAD with no enclosure noun (`WELL PUMP P-01`, `MTR`, "
     "`PMP-`) is `MTR_3PH`. Dedicated endpoint devices still win when named without an "
     "enclosure (VFD, CB/breaker, DISC/disconnect, XFMR, instruments, valves), and GROUND "
     "stays `GND`. Enforced in `symbol_infer.recognize_device` (suppresses the motor/starter "
     "rule when an enclosure noun is present) and reconciled after scan against the real "
     "library / project DWG blocks in `idp_project_symbols.apply_project_symbols` (landing "
     "correction + project confirmation). Fixes destinations that used to come out as motors."),
    ("Meter/Main & service are upstream", "For Source↔Dest orientation, a "
     "METER/MAIN, metered PEDESTAL, service entrance or PDB is UPSTREAM (source) of "
     "panels, PCPs and loads — rank it just below the utility. Fixes feeders like "
     "Meter/Main→PCP reading backwards. (`idp_anatomy._elec_rank`)"),
    ("VFD block (VFD_L/_R) now exists — use it for VFD/RVSS", "A `VFD`/`RVSS`/"
     "soft-start/`nVFDn` bucket feeding a motor now maps to the **`VFD_L/_R`** block "
     "(authored from MTRStrt: POWER, 3 terminals, Tag1='VFD'; registered in the "
     "cascade, contract, BlockIndex, and library DWGs). An across-line / FVNR "
     "starter still uses `MTRStrt`. So distinguish: drive → `VFD`, contactor+OL → "
     "`MTRStrt`. [Crows 3VFD1 → VFD_L]"),
    ("Symbol-library terminal capacity is known", "Each library block's terminal "
     "capacity is cataloged from its DWG (Term1..n attribute defs): 1-term ×36, "
     "2-term ×63, 3-term ×44, 4-term ×26, 0-term ×6. A symbol may only carry a wire "
     "group whose Wire Ct ≤ its terminal capacity — see `symbol_library_catalog.json` "
     "/ the 'Symbol Library (DWG)' sheet in Learned Logic.xlsx."),
    ("Source/Dest info on EVERY row", "Populate the SOURCE and DEST side of every "
     "fill row, not just grounds: the **S/D Symbol** = the device block for that "
     "end (`symbol_infer`); the **S/D Description** = the equipment name from the "
     "ConduitIndex Source/Dest (e.g. `MCC1-SEC. 2`, `1200A MSB`, `PANEL L`) on "
     "EVERY group; the **S/D Term** = the terminal landing — phases `ØA/ØB/ØC`(+`N`) "
     "on power feeders, `GND` on the separate ground conductor, breaker/circuit "
     "numbers on panelboard branches, and the EDC channel/loop terminals on "
     "signal/control. EDC-provided terminals win; convention fills the rest. The "
     "extractor auto-applies this (`idp_terms.apply_source_info` + "
     "`apply_panelboard_circuits`); a finished IDP never leaves a source blank."),
    ("Ground gauge from the schedule", "When the conduit/cable schedule lists a "
     "ground size (its GND column, e.g. `#2/o`, `#6`, `#8`), the `GRN` ground row "
     "uses THAT gauge — never a fabricated default. A schedule that lists grounds "
     "is authoritative: a conduit with none listed (utility service) gets NO "
     "ground. Motor blocks stay `MTR_3PH_NoGND` because the ground is its own row. "
     "(`idp_excel`/`idp_schedule` carry the gauge; `idp_anatomy.ensure_ground` "
     "only synthesizes where the schedule is silent about grounds.)"),
    ("Cable color", "PULL_ROPE / FIBER / CAT-6 carry no wire color → set Color to "
     "`N/A` (never blank, `XXX`, or a descriptor like `ETHERNET`). "
     "[Stratford1: P001, C003, C010–C013, C020A/B]"),
    ("Analog signal type", "A field-instrument analog signal pair (transmitters / "
     "level / flow / pressure — e.g. EFFLUENT FLOW, EFFLUENT PRESSURE, tank LEVEL) "
     "is ONE `TSP` with color `RED/BLK`, NOT two discrete `CONTROL` wires. "
     "[Stratford1: C004, C005, C008, C009, C017] — the extractor auto-applies "
     "this (`idp_anatomy.merge_analog_pairs`); you should too."),
    ("Ground on control/signal conduits", "Every real-circuit conduit carries a "
     "ground conductor shown `GRN`. In the current template it is a separate "
     "`POWER` / Wire-Ct-1 / `GND` / `GRN` fill (the only legal way to carry the "
     "GND symbol; the older sets used a `GROUND` type not in the current dropdown). "
     "[Stratford1: C002, C006, C010–C013, C015, C018] — the extractor auto-applies "
     "this (`idp_anatomy.ensure_ground`); you should too."),
    ("Power phase colors", "3-phase POWER = `BRN / ORG / YEL` (+ `WHT` neutral when "
     "4-wire) + `GRN` ground — confirmed against LISA output (H002, H005)."),
    ("Discrete switch vs. analog transmitter", "A field **switch** (pressure/level/"
     "flow switch — `PSH`/`PSL`/`LSH`/`FSH`/`TSH`/`ZS`, or the word SWITCH) is a "
     "DISCRETE dry contact → `CONTROL` wiring, NOT a `TSP`, even though its name "
     "contains PRESSURE/LEVEL/FLOW. Only a **transmitter** (`PIT`/`LIT`/`FIT`/…`IT` "
     "or the word TRANSMITTER) is analog → `TSP` `RED/BLK`. [Lennar C061 DISCHARGE "
     "PRESSURE SWITCH = 2× CONTROL `BLU`; A071 FLOWMETER = TSP; A051 LEVEL "
     "TRANSMITTER = POWER+TSP] — now enforced in `idp_anatomy.merge_analog_pairs`."),
    ("Loop-powered analog instrument", "A 2-wire loop-powered transmitter (e.g. a "
     "level transmitter `LIT`) lands loop `POWER` (2 conductors) PLUS the `TSP` "
     "signal pair — not TSP alone. [Lennar A051: POWER `#14` `BLU`/`BLK` + TSP "
     "`#18` `RED/BLK`]"),
    ("Utility service entrance", "The utility-transformer→meter/main service conduit "
     "is `POWER` with fill size/color **per utility** (leave the drawing's `TBD` and "
     "flag it — do not invent a gauge/color); it carries `ØA/ØB/ØC/N` and takes NO "
     "separate equipment ground (utility secondary). [Lennar P001]"),
    ("Blocks from the device, terms from the wiring diagram", "The **S/D Symbol** "
     "(block) is inferred from the DEVICE at each end (`symbol_infer`): transformer→"
     "`XFMR_3PH` (incl. a `TX-x` tag), **motor OR pump** (`MOTOR`/`MTR`/`PUMP`/`PMP`/"
     "`PMP-`)→`MTR_3PH` on a 3-conductor feeder (phase resolved by Wire Ct, not just "
     "the name), pressure/level/flow **switch** (`PSH`/`PSL`/`PSHL`/…)→`PressureSwitch`/"
     "`LevelSwitch`/`FlowSwitch`, disconnect→`DISC`, a panel / MCC / MSB / ATS / PLC / "
     "enclosure → `TB_Square` (terminal block, the default landing), ground→`GND`, "
     "pull rope→`Pullrope`, ethernet→`RJ45`. Do NOT leave everything as generic "
     "CB/TB — recognize the real device so the symbol matches the finished IDP. The **S/D Term** (terminal landing) is NOT on the "
     "conduit/cable schedule — read it from the AIC EDC three-line / terminal / "
     "wiring diagrams. **Do not change the workbook's own fill logic (types, counts, "
     "colors) and NEVER relabel the phases: power conductors stay `ØA/ØB/ØC` (+`N`) "
     "on BOTH ends** — the S/D Tag carries the EDC equipment (e.g. `MSB1`, `ATS1`, "
     "`EG1`), the term stays the phase. [Crows Landing 73.1188 DWG 73.1188-05: H005 "
     "`EG1:ØA/ØB/ØC/N → ATS1:ØA/ØB/ØC/N`; H003 `MSB1 → ATS1`, phases ØA/ØB/ØC]. The "
     "conduit/cable schedule (plans) owns the ConduitIndex; only terms/tags/blocks "
     "adhere to the EDC — the fill logic is unchanged."),
    ("Source block from BOTH ends; simple real blocks not generic CB/TB", "Pick the "
     "block from the device at that end AND the other end — an MCC/VFD/RVSS bucket "
     "feeding a **motor/pump** is a **motor STARTER** (`MTRStrt`), not a plain "
     "breaker; a conduit fed **from a transformer** (`TX-x`) shows `XFMR_3PH` on the "
     "source; a **pull section / UGPS / pull box** is `PullBox`; a switchboard main "
     "is breakers (`CB`). Only fall back to the generic terminal-block/breaker "
     "block when the device is genuinely unknown — don't leave everything CB/TB. "
     "(`idp_anatomy.refine_source_symbols` does this from the src+dest pair.)"),
    ("Tags vs. descriptions", "The **S/D Tag** columns are for a real device/terminal "
     "tag (e.g. an ISA point `YA-10`) — NOT the equipment name. The equipment NAME/"
     "description (e.g. `PANEL L`, `MCC1-SEC. 2`, `UGPS`) goes in the **S/D "
     "Description** columns; the wire label already pulls the name from the "
     "ConduitIndex Source/Dest Name. Leave the Tag blank when there's no sub-tag."),
    ("Where each terminal comes from (term source map)", "Read `S/D Term` from the "
     "matching EDC sheet by circuit kind: **3-phase power feeders** → the three-line "
     "(terms = phases `ØA/ØB/ØC`(+`N`/`GND`), tag = equipment); **120/208V branch "
     "circuits** → the **PANELBOARD SCHEDULE** (S Tag = the panel, e.g. `PANEL L`; "
     "S Term = the breaker/circuit numbers from the `L1-2,4,6`-style note; S Rating = "
     "the breaker amps) — and cross-check the circuit's panelboard description against "
     "the conduit's destination to catch a wrong circuit ref; **discrete control** "
     "(status/interlocks) → the **PLC DI sheets** (D Term = the PLC channel address "
     "`0.2.dd.dd`, tag = the ISA point, e.g. generator `YA/YH/YR/YF-10` → `0.2.01.05-08`); "
     "**analog instruments** (flow/level/pressure transmitters) → the **analog-input "
     "(AI) sheets** (term = the AI channel `0.2.0x.xx` and subpanel `TB` pair, tag = the "
     "loop, e.g. `FIT P-04`→`FOP-04`@`0.2.06.06`, `LIT-4001`@`0.2.05.01`). A pump "
     "**PSHL** pressure switch is usually a hardwired VFD protection interlock on the "
     "pump ELEMENTARY sheet (VFD terminals), not a PLC point. [Crows Landing 73.1188 "
     "DWG -05/-51/-54/-65/-69/-70/-71 + Panel L panelboard schedule]"),
    ("Extract EDC terminal landings — what generation needs, and how each is read",
     "Generation needs the S/D TERM (the landing) on every conduit, and it comes from "
     "the AIC EDC submittal by circuit kind — NOT from the plans: **3-phase feeders** → "
     "the three-line, terms = phases ØA/ØB/ØC(+N/GND); **120/240V branches** → the "
     "PANELBOARD SCHEDULE, panel-side term = the circuit CKT-## (tag = the panel); "
     "**discrete control** → the PLC DIGITAL-INPUT/OUTPUT block (term tag = TBDI-0.0n / "
     "TBDO-0.0n); **analog instruments** → the PLC ANALOG block (term tag = TBAI-0.0n). "
     "How the extractor reads them, in order: (1) TEXT-LAYER channel addresses "
     "(`0.r.ss.cc`) when the EDC exports them as text (`idp_edc.parse_io_sheets`); (2) "
     "OFFLINE POSITIONAL read of AIC LADDER sheets — the common form — where the channel "
     "NUMBER is graphical but the device DESCRIPTIONS are real text and channels run "
     "sequentially (INPUT 00,01,…): read the description blocks in the sheet's READING "
     "order (via the page rotation matrix — these sheets are rotated 270°), keep SPARE "
     "rungs + disambiguating digits (VFD 1 vs 2), drop the TB-tag/module/title blocks, and "
     "the POSITION IS the channel → exact TB:channel with NO OCR and NO vision, in ~0.5s "
     "(`idp_edc.parse_io_positional`/`match_io_positional`, hard-filtered analog↔digital). "
     "(2b) TB-only ladder fallback flags `edc_io_tb=…(verify channel#)` for anything the "
     "positional read can't resolve; (3) PANELBOARD by LOAD — derive "
     "CKT-## by matching the conduit's load name to the schedule's circuit descriptions "
     "when the note has no explicit circuit (`idp_terms._match_load_to_ckts`, flag "
     "`panelboard_circuits_by_load(verify ckt#)`); (4) VISION ESCALATION — when the EDC "
     "sheet is pure vector (channel/CKT numbers are geometric on rotated ladder drawings, "
     "so NOT reliably text-readable), render the sheet and transcribe via API key or the "
     "ASK_CLAUDE_EDC.md packet to get exact channels. Never fabricate a channel/circuit "
     "number — land the TB/panel tag and flag the number for a quick check. Feed the EDC "
     "PDFs to the run (`apply_edc_terms_from_paths`); without them the terms stay blank. "
     "GOTCHA: when a CABLES-IN-CONDUIT list is folded into a conduit's note, do NOT read "
     "its cable IDs as panelboard circuits — `C-021` is cable 21, not circuit 21. Strip "
     "the cable list before parsing circuit numbers (`idp_terms._strip_cable_list`); "
     "otherwise a panel branch lands on a bogus circuit and blocks its real phase terms. "
     "[56.1059 Main PLC: TBDI-0.02/0.03, TBAI-0.07/0.08 matched by device; channels via "
     "vision. 73.1188: text-layer channel addresses parsed directly.]"),
    ("Fill precision needs the CABLE schedule — the conduit schedule alone under-"
     "specifies", "The conduit schedule (MC-E-8 / E-3) lists WHICH cables run in each"
     "conduit and a coarse TYP (POWER/CONTROL), but NOT the per-conductor specs. The "
     "conductor count, gauge, insulation/type (TSP vs CAT-6 vs FIBER vs a mfg cable), "
     "and — critically — whether a run carries a green EGC all live on the **cable "
     "schedule** (MC-E-9 / the C-### cable list). Comparing the Moccasin 56.1059 "
     "auto-run (conduit schedule only) against the curated fill made the gap concrete: "
     "(a) **over-grounding** — with no ground column, the writer synthesized a ground "
     "on every circuit (61 grounds vs the 25 the cable schedule actually lists), so "
     "control/signal/spare runs got a false EGC; (b) **coarse types** — everything "
     "collapsed to POWER/CONTROL, losing the real TSP/CAT-6/FIBER/MFG_CABLE "
     "distinctions. So: derive fill from the CABLE schedule when one exists; the "
     "conduit-schedule TYP is only a fallback seed. When you must run conduit-schedule-"
     "only, mark the records NOT ground-authoritative and flag the synthesized grounds "
     "+ coarse types (`ocr_conduit_only_fill_from_cables`) so the confidence score "
     "reflects the assumption rather than reading as complete. [56.1059 skill-v9 "
     "curated=25 grounds/precise types vs exe-v1 auto=61 grounds/POWER-CONTROL only]"),
]

# ── ConduitExtractor-domain conventions (conduit metadata) ───────────────────
CONDUIT_CONVENTIONS = [
    ("A finished AIC IDP is ONE CONDUIT PER SHEET — read it by OCR, not the text layer", "An "
     "AIC interconnection-diagram finished IDP (title 'INTERCONNECTION DIAGRAMS') draws one "
     "conduit per sheet, and the identity + fill are GRAPHICAL (the PDF text layer carries "
     "almost none of it — the text-only reader recovers ~7 of 24). Read each sheet by OFFLINE "
     "OCR instead (idp_idp_ocr.read_interconnection_idp): the FIELD header gives `NAME: <tag>` "
     "(the conduit), `TYPE: <PVC/RGS/…>`, `SIZE: <4\"/2\"/3-4\">`; the SOURCE label (top-left) "
     "and DESTINATION label (top-right) give from/to; and the FILL TYPE table "
     "(FILL TYPE | SIZE | COLOR | QUANTITY) gives the conductors. Fill vocabulary → LISA type: "
     "POWER→POWER, CONTROL→CONTROL, GROUND→ground group (S/D Tag GND), TSP→TSP (RED/BLK), "
     "PULL ROPE→PULL_ROPE, ETHERNET→CAT-6, fiber (OM3/SMFO)→FIBER, NONE→empty conduit. "
     "QUANTITY is the conductor count; SIZE like `#10AWG`/`400KCMIL`/`#18` is the gauge; a "
     "`XXX`/`XX` NAME/TYPE/SIZE means the sheet itself flags it TBD (don't invent). Learned "
     "from City of Gonzales Industrial WWTF (56.1113): 24 conduits offline vs 7 by text."),
    ("Conduit tags can be ANY scheme — circuit class lives in the TYP column", "Don't "
     "assume AIC's P/C/A prefix. Some owners tag every conduit with one prefix (e.g. "
     "Moccasin 56.1059 uses `K-001…K-067` for power, control, and ground alike) and put "
     "the circuit class in a separate TYP column (POWER / CONTROL / GROUND / PWR-CONTRL). "
     "Take `conduit_name` verbatim from the schedule; read the class from TYP (or the "
     "fill), never from the tag prefix."),
    ("SIZE column often combines trade size + material", "A conduit-schedule SIZE cell "
     "like `2\" PVC`, `1.25\" RGS`, `3/4\" RGS` carries BOTH the trade size and the "
     "material. Split it: `conduit_size` = the trade size (`2\"`), `conduit_type` = the "
     "material normalized to the enum (PVC / RGS). A COMMENT of \"EXPOSED/SURFACE MOUNT "
     "TO BE RGS\" means the run is PVC buried but RGS where exposed — keep the design "
     "material as `conduit_type` and note the RGS-exposed nuance. [Moccasin MC-E-8]"),
    ("Generator stays the destination (keep schedule direction)", "A generator "
     "(`EG#`, ENGINE GENERATOR) connected to an ATS is drawn by AIC as the "
     "DESTINATION, not the source — the finished IDP keeps the conduit schedule's "
     "FROM/TO (`ATS→EG1`), NOT the electrical direction. [73.1188 H2203: "
     "SOURCE=ATS1, DEST=EG1]. Do not flip a generator conduit to make the generator "
     "upstream. (`idp_anatomy._elec_rank` ranks EG as a load on every conduit.)"),
    ("Index-less IDP packages", "Some finished IDP PDFs have NO drawing-index page "
     "and interleave client spec pages with the diagrams (e.g. San Rafael 56.1125, "
     "conduits named S1/E05). The reader enumerates conduits directly off each "
     "diagram page — anchoring the conduit id to the NAME: label (rotation-agnostic) "
     "with a broadened id pattern (`[A-Z]{1,3}\\d{1,4}`). Such packages parse only "
     "partially (non-standard layout/geometry) — verify the conduit count and flag "
     "gaps rather than assume completeness. (`idp_idp_pdf._parse_page_fills`)"),
    ("Only real, documented conduits", "Emit one ConduitIndex row per physically "
     "documented run — no speculative or duplicated conduits. Training on finished "
     "sets showed over-production (Stratford: ~72 emitted vs ~42 real). If a run "
     "isn't in the schedule or schematic, flag it — never invent it."),
    ("Conduit Type normalization", "Normalize the source MATERIAL / TYPE to the "
     "`Conduit Type` enum: `PVC-40` / `PVC-80` → `PVC`; `GRC` / rigid galvanized → "
     "`RGS`; `RMC` stays `RMC`. An unknown/absent type → `XXX` + a flag, never "
     "blank and never a guessed material."),
    ("Wire fill stays out of ConduitIndex", "A conductor / wire-fill column in the "
     "schedule is Phase 1b (FillIndex) input, not a conduit field. Retain the text "
     "separately for FillIndex and note anything unusual in `deviations_notes`."),
    ("Target fast + remember layouts — don't deep-scan every file", "On a big project "
     "folder, don't OCR/full-scan every file. TARGET like the skill does: skip irrelevant "
     "subtrees by folder (estimating, quotes-received, photos, correspondence, PMO) and "
     "extension without opening them; decide cover letters (`_SCL_`/`_CL_`/transmittal) and "
     "vendor cut sheets (datasheet/catalog, or a Cut-Sheets folder) by NAME/FOLDER — vendor "
     "datasheets resemble PLC-I/O content, so never classify them by content. Recognize a "
     "document by its LAYOUT SIGNATURE from the first page (`idp_layouts`: seed token sets "
     "per role + tokens LEARNED from confirmed examples, confidence gated on the SEED "
     "overlap so one over-fit example can't mislabel a related doc). Only relevant files "
     "whose type isn't recognized by signature fall through to the authoritative full-"
     "document finders — and each confirmed type is LEARNED so the next project is faster. "
     "[56.1059: skipped Estimating/Quotes; Eaton Dwg_Panel→panelboard, moxa datasheet→cut_"
     "sheet by name, EDC PLC→edc_plc_io.]"),
    ("Route a whole project folder — each field knows its source", "Given an entire "
     "project folder, classify EVERY file by ROLE and populate each workbook field from "
     "the source that actually carries it, instead of hand-picking files "
     "(`idp_router.classify`). Roles: CONDUIT SCHEDULE (conduit name/src/dst/size/type), "
     "CABLE SCHEDULE (fill type/gauge/colors + grounds), EDC THREE-LINE (3Ø feeder phase "
     "terms), PANELBOARD SCHEDULE (branch CKT-##), EDC PLC I/O (TBDI/TBAI control+analog "
     "terms), CUT SHEETS + project DWGs (S/D symbols), finished IDP (reference / conduit "
     "source of LAST resort only), cover letters + cut sheets (NO conduit data — kept out "
     "of conduit extraction). Detect by the authoritative full-document finders "
     "(find_schedule_pages / find_cable_schedule_pages / find_edc_sheets / is_idp_package), "
     "NOT just head-text — a conduit schedule can sit on page 37 of a 55-page plan set. "
     "Pick the LATEST revision per single-source role (R##/date in the filename). Log the "
     "field→source map so it's auditable. [56.1059: Plans→conduit_schedule, EDC Main PLC→"
     "edc_plc_io, Eaton Dwg_Panel→panelboard, *_SCL_*→cover_letter, datasheet→cut_sheet]"),
    ("Hand FillIndex the CABLE schedule, not just the conduit schedule", "When a "
     "project splits its schedules — a **conduit schedule** (MC-E-8: id / TYP / FROM / "
     "TO / SIZE / cables-in-conduit) and a separate **cable schedule** (MC-E-9: C-### / "
     "SIZE / TYPE-INSUL / FROM / TO / which conduits it routes through) — capture BOTH "
     "and pass the cable schedule through for Phase 1b. The conduit schedule alone lets "
     "FillIndex seed only a coarse POWER/CONTROL fill and forces it to synthesize "
     "grounds blindly; the cable schedule is what carries the per-conductor gauge, "
     "insulation/type, and green-EGC presence. Note the cable-schedule sheet id in the "
     "conduit record so FillIndex knows where the real fill data is. [Moccasin 56.1059 "
     "MC-E-8 conduit + MC-E-9 cable]"),
    ("Prefer the finished IDP package as source", "When a project has AIC's own "
     "finished IDP submittal PDF (`…_Prj_IDP_<date>.pdf` — a drawing-index page + one "
     "interconnection-diagram page per conduit), extract from THAT, not the "
     "marked-up contract plans, which are usually scanned/vector with no "
     "text-extractable schedule. [Lennar: `Plans (Marked-Up).pdf` → nothing; the "
     "finished IDP → all 7 conduits] The extractor auto-detects this format "
     "(`idp_idp_pdf`)."),
    ("`_SCL_` is a Submittal Cover Letter, not a schedule", "A `…_IDP_SCL_<date>.pdf` "
     "is the cover letter / review-comments log — do NOT parse it for conduits. Mine "
     "it for project conventions and RFI/deviation responses (conduit material, wire "
     "color changes, removed conduits). [Lennar SCL: all conduits PVC SCH40; C061 "
     "control changed to BLU; spares X001–X003 removed]"),
    ("Conduit-tag prefix hints at circuit class", "AIC tags `P###` = power feeders, "
     "`C###` = control, `A###` = analog/instrument. Use as a weak type hint — always "
     "confirm against the actual fill, never as the sole basis."),
    ("Source/Destination follow electrical direction", "Orient each conduit so "
     "**Source = the upstream end** (where power/signal originates) and "
     "**Destination = the downstream end** (the load / field device), per the "
     "distribution hierarchy: utility → transformer → UGPS → MSB → ATS → MCC → "
     "panel → VFD/starter → motor/load; PLC/controller is the source for a "
     "control conduit. A schedule's FROM/TO may be listed backwards — fix it. "
     "Dual-role gotcha: a **generator** is the SOURCE on its main feeder (it feeds "
     "the ATS, so `ATS→EG` becomes `EG→ATS`) but the DESTINATION for its auxiliaries "
     "(block heater / charger / controls, fed by an MCC or panel); disambiguate by "
     "the fill — a main feeder uses MCM/large conductors. [Crows Landing 73.1188: "
     "H005 corrected to EG1→ATS] — enforced by `idp_anatomy.orient_by_electrical_direction`."),
]

# Bounded region inside each SKILL.md that export_to_skills owns and rewrites.
_BLK_BEGIN = ("<!-- BEGIN AUTO-LEARNED (managed by IDP Extractor — do not edit "
              "inside the markers) -->")
_BLK_END = "<!-- END AUTO-LEARNED -->"


def _inject_block(md_path, block_body):
    """Replace the text between the AUTO-LEARNED markers in a SKILL.md, leaving
    everything else untouched. No-op (returns False) if the file or markers are
    absent — we never create the section blind, so a hand-managed skill is safe."""
    try:
        with open(md_path, encoding="utf-8") as fh:
            text = fh.read()
    except OSError:
        return False
    i = text.find(_BLK_BEGIN)
    j = text.find(_BLK_END)
    if i == -1 or j == -1 or j < i:
        return False
    new = text[:i + len(_BLK_BEGIN)] + "\n" + block_body.rstrip() + "\n" + text[j:]
    if new == text:
        return True
    try:
        with open(md_path, "w", encoding="utf-8") as fh:
            fh.write(new)
        return True
    except OSError:
        return False


def _fill_reference_body(rules):
    """Full FillIndex reference: device→symbol table + fill conventions."""
    lines = ["# Learned conventions — FillIndex (auto-generated by the IDP Extractor)",
             "",
             "Device-name → symbol-token associations harvested from finished IDP",
             "drawings and from plans↔finished↔generated training comparisons, plus",
             "the wire-attribute conventions confirmed in the same runs. The",
             "extractor regenerates this file; treat it as accumulated ground truth",
             "when inferring `S Symbol` / `D Symbol` and wire attributes.",
             "",
             f"_{len(rules)} learned device→symbol rule(s)._",
             "",
             "| Device name (match) | Symbol token (result) |",
             "|---|---|"]
    for r in sorted(rules, key=lambda x: x.get("match", "")):
        lines.append(f"| {r.get('match','')} | {r.get('result','')} |")
    lines += ["", "## Wire-fill conventions learned from finished-vs-generated training", ""]
    for title, rule in FILL_CONVENTIONS:
        lines.append(f"- **{title}.** {rule}")
    return "\n".join(lines) + "\n"


def _conduit_reference_body():
    """Full ConduitExtractor reference: conduit-domain conventions only."""
    lines = ["# Learned conventions — ConduitExtractor (auto-generated by the IDP Extractor)",
             "",
             "Conduit-domain ground truth confirmed against finished IDP drawings and",
             "plans↔finished↔generated training comparisons. The extractor regenerates",
             "this file. It never carries wire-fill or symbol conventions — those",
             "belong to the FillIndex skill.",
             "",
             "## Conduit conventions learned from finished-vs-generated training", ""]
    for title, rule in CONDUIT_CONVENTIONS:
        lines.append(f"- **{title}.** {rule}")
    return "\n".join(lines) + "\n"


def _fill_skill_block(rules):
    """Compact block injected into idp-fillindex/SKILL.md. Curated conventions
    inline (always in context); the long device table stays in the reference."""
    lines = []
    for title, rule in FILL_CONVENTIONS:
        lines.append(f"- **{title}.** {rule}")
    lines.append("")
    if rules:
        lines.append(f"_{len(rules)} learned device→symbol rule(s) available — see "
                     "`references/learned-conventions.md` for the full table._")
    else:
        lines.append("_No device→symbol rules learned yet — see "
                     "`references/learned-conventions.md`._")
    return "\n".join(lines)


def _conduit_skill_block():
    """Compact block injected into idp-conduit-extractor/SKILL.md."""
    lines = [f"- **{title}.** {rule}" for title, rule in CONDUIT_CONVENTIONS]
    lines.append("")
    lines.append("_See `references/learned-conventions.md` for the full conduit-domain set._")
    return "\n".join(lines)


def export_to_skills(skills_dir=None):
    """Route learned conventions to the skill whose schema/end-goal each serves,
    writing BOTH the on-demand reference file and the always-in-context SKILL.md
    block — to EVERY skills location (project source copy AND the installed
    ~/.claude/skills the Skill tool runs), so the two never drift apart again.
    FillIndex gets device→symbol rules + wire-fill conventions; ConduitExtractor
    gets conduit-metadata conventions only. Returns the list of files written."""
    dirs = [skills_dir] if skills_dir else _find_skills_dirs()
    try:
        import logic_store
        rules = [r for r in logic_store.load().get("rules", [])
                 if r.get("type") == "symbol_keyword" and r.get("result")]
    except Exception:
        rules = []

    plan = {
        "idp-fillindex": (_fill_reference_body(rules), _fill_skill_block(rules)),
        "idp-conduit-extractor": (_conduit_reference_body(), _conduit_skill_block()),
    }
    written = []
    for sd in dirs:
        for skill, (ref_body, skill_block) in plan.items():
            refdir = os.path.join(sd, skill, "references")
            if os.path.isdir(refdir):
                path = os.path.join(refdir, "learned-conventions.md")
                try:
                    with open(path, "w", encoding="utf-8") as fh:
                        fh.write(ref_body)
                    written.append(path)
                except OSError:
                    pass
            md_path = os.path.join(sd, skill, "SKILL.md")
            if _inject_block(md_path, skill_block):
                written.append(md_path)
    # also refresh the human-readable "Learned Logic" workbook next to the tool
    try:
        xl = export_learned_logic_xlsx()
        if xl:
            written.append(xl)
    except Exception:
        pass
    return written


def export_learned_logic_xlsx(out_path=None):
    """Write a human-readable 'Learned Logic.xlsx' beside the tool: built-in +
    learned device→symbol rules, fill & conduit conventions, and the offline
    OCR/accuracy logic. Refreshed on every export_to_skills() call so it stays in
    sync with what the extractor actually applies. Returns the path (or None)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import symbol_infer
        import logic_store
    except Exception:
        return None
    if out_path is None:
        out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Learned Logic.xlsx")
    wb = openpyxl.Workbook()
    HDR = PatternFill("solid", fgColor="16263D"); HF = Font(color="FFFFFF", bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    def sheet(name, headers, rows, widths):
        ws = wb.create_sheet(name[:31])
        for c, h in enumerate(headers, 1):
            x = ws.cell(1, c, h); x.fill = HDR; x.font = HF
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v).alignment = WRAP
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    sheet("Device to Symbol (built-in)", ["Device / meaning", "Keyword pattern", "LISA symbol"],
          [(dsc, kw, tok) for kw, tok, dsc in symbol_infer._RULES], [46, 52, 20])
    d = logic_store.load()
    lr = [(r.get("match", ""), r.get("result", ""), r.get("note", "") or r.get("source", ""))
          for r in d.get("rules", []) if r.get("type") == "symbol_keyword"]
    sheet("Device to Symbol (learned)", ["Device keyword", "Symbol", "Source / note"], lr, [26, 20, 60])
    sheet("Fill & Wire Conventions", ["Convention", "Rule"], list(FILL_CONVENTIONS), [34, 120])
    sheet("Conduit Conventions", ["Convention", "Rule"], list(CONDUIT_CONVENTIONS), [34, 120])
    ocr = [
        ("Vector schedule OCR", "Locate table upper-left, render ~500 DPI, OCR, bin to columns. Offline.", "idp_ocr_schedule"),
        ("High-accuracy 2-pass", "Re-crop tight + re-OCR 560 DPI. ~86%->95%.", "read_schedule refine"),
        ("ID sequence repair", "Implausible conduit # snaps to running sequence.", "_repair_ids"),
        ("Name self-consistency", "Rare name misread snaps to frequent canonical; guards distinguishing digits.", "_repair_names"),
        ("Prime/quote normalize", "Curly quotes/primes unified (vault names).", "read_schedule"),
        ("MSB boundary", "1000AMSB -> CB-CB-CB not TB.", "refine_source_symbols"),
        ("Notes column", "Notes -> panel circuit numbers.", "idp_excel/idp_schedule"),
        ("Source info every row", "Symbol+Term+Description on every group, not just grounds.", "apply_source_info"),
        ("Ground from schedule", "Real GND gauge; authoritative (utility gets none).", "ensure_ground"),
    ]
    sheet("OCR & Accuracy Logic", ["Logic", "What it does", "Where"], ocr, [26, 90, 40])
    # symbol-library catalog learned from the block DWGs (if a catalog exists)
    cat_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "symbol_library_catalog.json")
    if os.path.exists(cat_path):
        try:
            import json as _json
            cat = _json.load(open(cat_path, encoding="utf-8"))
            rows = []
            for nm in sorted(cat):
                v = cat[nm]
                if "error" in v:
                    continue
                g = ", ".join(f"{k}:{n}" for k, n in sorted(v.get("geom", {}).items()))
                rows.append((nm, v.get("device_default", ""), v.get("terminals", ""),
                             v.get("tag_slots", ""), g))
            sheet("Symbol Library (DWG)",
                  ["Block", "Device tag", "Terminals", "Tag slots", "Geometry"],
                  rows, [30, 14, 12, 12, 40])
        except Exception:
            pass
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(out_path)
    return out_path


# ── orchestrator ─────────────────────────────────────────────────────────────
def run_training(plans=None, finished=None, generated=None, learn=True, log=print):
    report = {"conduits_compared": 0, "gaps": [], "learned": [], "uncertain": [],
              "summary": ""}
    ground = _ground_truth_records(finished or [])
    ours = _our_records(generated or [])
    log(f"ground truth: {len(ground)} conduits from finished IDPs; "
        f"ours: {len(ours)} conduits from generated output")
    if not ground or not ours:
        report["summary"] = ("Need BOTH a finished IDP (drawings) and our generated "
                             "output (workbook or DWGs) to compare.")
        return report
    gaps = diff_records(ground, ours)
    report["gaps"] = gaps
    report["conduits_compared"] = len({_norm(r["name"]) for r in ground}
                                      & {_norm(r["name"]) for r in ours})
    # ── PROJECT-MISMATCH GUARD ──────────────────────────────────────────────────────────
    # If both sides carry a real set of conduits but they share almost NONE, the finished IDP
    # and the generated output are DIFFERENT PROJECTS (the wrong pair was loaded). Every
    # conduit then shows as missing/extra — which would flood "uncertainties" with false gaps
    # and could teach garbage. Refuse to learn or escalate; tell the user to fix the pair.
    n_ground = len({_norm(r.get("name")) for r in ground if _is_real_conduit(r.get("name"))})
    n_ours = len({_norm(r.get("name")) for r in ours if _is_real_conduit(r.get("name"))})
    smaller = min(n_ground, n_ours)
    if smaller >= 5 and report["conduits_compared"] <= max(1, int(0.10 * smaller)):
        report["mismatch"] = True
        report["learned"] = []
        report["uncertain"] = []
        report["summary"] = (
            f"⚠ PROJECT MISMATCH — the finished IDP ({n_ground} conduits) and the generated "
            f"output ({n_ours} conduits) share only {report['conduits_compared']} conduit "
            f"name(s), so they look like DIFFERENT projects. Nothing was learned. Load the "
            f"SAME project's finished IDP AND its generated workbook, then train again.")
        log(report["summary"])
        if plans:
            report["plans"] = list(plans)
        return report
    if learn:
        report["learned"] = _learn_from_gaps(ground, gaps)
        if report["learned"]:
            export_to_skills()
    report["uncertain"] = _uncertain_from_gaps(gaps)
    report["summary"] = (f"{report['conduits_compared']} conduits compared · "
                        f"{len(gaps)} gap(s) · {len(report['learned'])} rule(s) learned · "
                        f"{len(report['uncertain'])} item(s) need Claude")
    if plans:
        report["plans"] = list(plans)
    return report


if __name__ == "__main__":
    import sys
    fin = [sys.argv[1]] if len(sys.argv) > 1 else []
    gen = [sys.argv[2]] if len(sys.argv) > 2 else []
    rep = run_training(finished=fin, generated=gen)
    print(rep["summary"])
    for g in rep["gaps"][:25]:
        print(f"  [{g['conduit']}] {g['field']}: ground={g['ground']!r} ours={g['ours']!r} ({g['kind']})")
