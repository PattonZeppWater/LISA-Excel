"""
IDP Excel writer.

Takes parsed records (from idp_extract) and writes them into a copy of the IDP
workbook template (.xlsm), preserving VBA/macros and formatting.

Layout target: IDP_Workbook_CurrentWIP_3.xlsm (matches the sample filled
workbook "05_IDP_Workbook_CurrentWIP_3.xlsm").

Sheets written:
  ConduitIndex : A=Conduit Name, B-D=Source Name 1-3, E-G=Destination Name 1-3,
                 H=Conduit Size, I=Conduit Type, J=Ref Documents (indices),
                 K=Deviations Notes
  Ref Documents & Deviations : A=index #, B=DWG#, C=Description, D=Manufacturer
                 (deduped, global; header row 2, data row 3+)
  FillIndex : A=Conduit, B=Wire Ct, C=Type, D=Wire Gauge, E=S Symbol,
                 F-I=Color 1-4, J-M=S Tag 1-4, R-U=S Term 1-4,
                 AF/AI/AJ/AK=D Tag 1-4, AP-AS=D Term 1-4,
                 BA-BD=Wire Label 1-4, BE-BH=WL Mode (header row 2, data row 3+)
"""

import os
import re
import shutil
import warnings

warnings.filterwarnings("ignore")


def versioned_path(path):
    """Never overwrite: if `path` exists, return path with a _vN suffix (v2, v3…)
    so each extraction keeps its own file and previous results are preserved."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 2
    while os.path.exists(f"{base}_v{n}{ext}"):
        n += 1
    return f"{base}_v{n}{ext}"

import openpyxl
import idp_terms
from openpyxl.formula.translate import Translator
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

# Uncertain-value flag: light amber fill + a comment explaining why. Distinct
# from the grey-out (FF808080) so the de-grey pass leaves flags intact.
_FLAG_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

# ---- ConduitIndex columns (1-based) ----
CI_NAME, CI_SRC1, CI_SRC2, CI_SRC3 = 1, 2, 3, 4
CI_DST1, CI_DST2, CI_DST3 = 5, 6, 7
CI_SIZE, CI_TYPE, CI_REF, CI_DEV = 8, 9, 10, 11
CI_NCOLS = 11

# Optional parsed panelboard map {ckt: {"breaker","desc"}} — set by the ingest
# layer (idp_panelboard.find_and_parse) so idp_terms can add breaker ratings +
# cross-check branch circuits. None = circuit-number terms only (still correct).
_panelboard_map = None

# ---- FillIndex columns (1-based) ----
FI_CONDUIT, FI_WIRECT, FI_TYPE, FI_GAUGE, FI_SSYM = 1, 2, 3, 4, 5
FI_DSYM = 29                            # AC  D Symbol
FI_COLOR = (6, 7, 8, 9)                 # F-I
FI_STAG = (10, 11, 12, 13)              # J-M
FI_STERM = (18, 19, 20, 21)            # R-U
FI_DTAG = (32, 35, 36, 37)             # AF, AI, AJ, AK
FI_DTERM = (42, 43, 44, 45)            # AP-AS
# S/D ISATag_Loop#/ElementID/Element#/FunctionID — per LISA_workbook_mapper.py,
# these only apply to instrument blocks (symbol name contains "Inst"); LISA maps
# them onto the instrument bubble (ISATag_LoopNum/ElementIdent/ElemNum/FunctIdent).
FI_SISATAG = (14, 15, 16, 17)           # N-Q
FI_DISATAG = (38, 39, 40, 41)           # AL-AO
FI_SRATING, FI_SFUSE = 22, 23           # V, W — device Rating/Fuse Rating (general)
FI_DRATING, FI_DFUSE = 46, 47           # AT, AU
FI_SDESC = (24, 25, 26)                 # X-Z  — free-text description (general)
FI_DDESC = (48, 49, 50)                 # AV-AX
# S/D Type + S/D Quantity — per LISA_workbook_mapper.py, these are SPARE-BLOCK-ONLY
# attributes (Src_SpareType/Qty, Dst_SpareType/Qty on Spare_L/Spare_R blocks), NOT
# a general device type/quantity field. Only write them when the symbol is Spare.
FI_STYPE, FI_SQTY = 27, 28              # AA, AB
FI_DTYPE, FI_DQTY = 51, 52              # AY, AZ
FI_LABELS = {53: "BA", 54: "BB", 55: "BC", 56: "BD"}   # Wire Label 1-4
FI_MODES = (57, 58, 59, 60)            # WL1..WL4 Mode
FI_NCOLS = 60


def _flag_cell(ws, row, col, msg):
    """Mark a populated cell as needs-review: amber highlight + comment.
    Appends to an existing comment (rather than overwriting) so a cell that
    accumulates multiple flags keeps every reason visible."""
    cell = ws.cell(row, col)
    cell.fill = _FLAG_FILL
    prior = cell.comment.text if cell.comment else ""
    lines = [re.sub(r"^REVIEW: ", "", ln).strip() for ln in prior.split("\n") if ln.strip()]
    if msg not in lines:
        lines.append(msg)
    c = Comment("REVIEW: " + "\nREVIEW: ".join(lines), "IDP Extractor")
    c.width, c.height = 320, 120
    cell.comment = c


def _suppress_repeat_tags(tags):
    """Wire-label tag rule: within a row's 4 tag slots, show a tag only when it
    differs from the previous shown tag; blank consecutive repeats. So
    ['TB-5','TB-5','TB-5','CB-13'] -> ['TB-5','','','CB-13'] and four identical
    tags -> only the first is shown. Terminals are never suppressed."""
    out, prev = [], None
    for t in tags:
        t = (t or "").strip()
        if t and t == prev:
            out.append("")
        else:
            out.append(t)
            if t:
                prev = t
    return out


def _last_data_row(ws, cols, start_row):
    """Return the last row (>= start_row-1) that has any value in the given cols."""
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in cols):
            last = r
    return last


_BAD_TEMPLATE_MARKERS = ("_filled", "_nogrey", "_idp_filled")
_VERSION_SUFFIX_RE = re.compile(r"_v\d+$")


def check_template_sane(path):
    """Guard against feeding a PRIOR OUTPUT back in as the template.

    A de-greyed (`_NoGrey`) or already-filled (`_FILLED`) copy can carry a
    corrupted or truncated vbaProject.bin — confirmed the hard way: a stray
    `..._FILLED_NoGrey.xlsm` sample had a vbaProject.bin 86KB smaller than the
    real template's, and reusing it as a template silently propagated that
    corruption into a brand-new workbook that Excel then refused to open
    ("we found a problem with some content"). `degrey()`/`write_workbook()`
    themselves preserve the VBA project byte-for-byte when run against a good
    template — the corruption came in with whatever file was fed as the
    template, not from processing it. Raises ValueError with the fix.
    """
    base = os.path.splitext(os.path.basename(path))[0].lower()
    base = _VERSION_SUFFIX_RE.sub("", base)
    if any(m in base for m in _BAD_TEMPLATE_MARKERS):
        raise ValueError(
            f"'{os.path.basename(path)}' looks like a PREVIOUS OUTPUT "
            f"(matches _FILLED / _NoGrey), not the original template. A "
            f"de-greyed or already-filled copy can carry a damaged macro "
            f"project (grey-out relies on VBA) — reusing it as the template "
            f"propagates that damage into every new workbook, which is "
            f"exactly what makes Excel refuse to open the result. Point "
            f"'Template workbook' at the original IDP_Workbook_CurrentWIP_3.xlsm "
            f"instead (never a _FILLED or _NoGrey file).")


def _sanitize_formula_leaks(ws, start_row, end_row, ncols, formula_cols=()):
    """Force any cell openpyxl auto-detected as a FORMULA back to literal text,
    except columns we deliberately write real formulas into (formula_cols).

    openpyxl treats ANY string value starting with '=' as a formula the moment
    it's assigned (confirmed: `cell.value = "="` -> `cell.data_type == 'f'`,
    even for plain extracted text). This is exactly what corrupted a real
    workbook: a CAD-symbol-font PDF misread a device name into garbled text
    that happened to start with '=' ('= - 27 - = -'); openpyxl stored it as a
    formula; Excel's formula parser choked on the invalid syntax on open
    ("we found a problem with some content"). Extracted text is never
    intentionally a formula here — the ONLY real formulas we write are the
    Wire Label TEXTJOIN cells, which the caller excludes via formula_cols.
    """
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            if c in formula_cols:
                continue
            cell = ws.cell(r, c)
            if cell.data_type == "f" and isinstance(cell.value, str):
                cell.data_type = "s"


def degrey(in_path, out_path=None):
    """Strip the grey-out (FF808080) fills. Returns count of de-greyed cells."""
    wb = openpyxl.load_workbook(in_path, keep_vba=True)
    nofill = PatternFill(fill_type=None)
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.fill
                if f and f.patternType == "solid" and str(getattr(f.fgColor, "rgb", "")) == "FF808080":
                    c.fill = nofill
                    n += 1
    wb.save(out_path or in_path)
    return n


def write_workbook(records, template_path, out_path, clear_rows=True, add_flags=True,
                   lisa_gate=True, anatomy=True, schedule_doc=None, clear_deviations=False):
    _schedule_doc = schedule_doc
    check_template_sane(template_path)   # never build on top of a prior output
    # LISA-readiness gate: remap raw conductor counts to legal connection counts
    # and snap symbols to legal dropdown blocks so LISA can generate the IDP.
    # Mutates records in place; degrades gracefully if the contract is unavailable.
    if lisa_gate:
        try:
            import lisa_contract, idp_anatomy as _anat
            lisa_contract.normalize_types(records)        # off-dropdown types -> legal
            # convention fixes that add/retype fills BEFORE symbol gating, so the
            # new ground (POWER/GND) and TSP symbols get validated + snapped too
            if anatomy:
                _anat.orient_by_electrical_direction(records)  # source=upstream, dest=downstream
                _anat.merge_analog_pairs(records)         # analog 2×CONTROL -> 1 TSP
                _anat.ensure_ground(records)              # add GRN ground to real circuits
            lisa_contract.normalize_connections(records)  # raw counts -> connections
        except Exception:
            pass
    # Finished-IDP conventions: fill missing wire colors/gauge (power phases
    # BRN/ORG/YEL, control RED, TSP RED/BLK, ...) so sheets render correctly.
    # check_archetypes surfaces drawing-fidelity gaps (missing ground, flipped
    # symbol side, no fill rows, ...) as amber flags on the ConduitIndex row —
    # findings that used to exist only as an unused diagnostic function.
    archetype_notes = {}
    if anatomy:
        try:
            import idp_anatomy
            idp_anatomy.apply_conventions(records)
            idp_anatomy.refine_source_symbols(records)   # device-specific source blocks
            for issue in idp_anatomy.check_archetypes(records):
                archetype_notes.setdefault(issue["conduit"], []).append(issue["note"])
        except Exception:
            pass
    # EDC-CONFIRMED SYMBOL GATE: a device S/D Symbol carrying NO inference confidence is an
    # unvalidated name-echo (e.g. VFD_R lifted straight from a "VFD(PERMEATE PUMP)" label) —
    # not a symbol the drawing/EDC actually confirms. Leave it BLANK; only scored/confirmed
    # symbols (and grounds) are written. Matches the curated skill (which fills confirmed
    # device symbols but never fabricates one from a bare equipment name).
    try:
        _blanked = 0
        for rec in records or []:
            for g in rec.get("fill", []) or []:
                if g.get("is_ground"):
                    continue
                for sk, ck in (("s_symbol", "s_symbol_conf"), ("d_symbol", "d_symbol_conf")):
                    sym = g.get(sk)
                    if (sym and g.get(ck) is None
                            and str(sym).upper().split("_")[0] != "GND"):
                        g[sk] = None
                        g.setdefault("flags", []).append(sk + "_unconfirmed_blanked")
                        _blanked += 1
    except Exception:
        pass
    # PICK FROM THE PICKLIST: every S/D Symbol we write must be a value that actually exists
    # in the workbook's dropdown for its (type, wire ct, side). Snap any off-list value to
    # the closest valid pick (flagged for review) so we never write a symbol the dropdown /
    # LISA can't offer. No-op on symbols that are already valid picks.
    try:
        import lisa_contract as _lc
        for rec in records or []:
            for g in rec.get("fill", []) or []:
                if g.get("is_ground"):
                    continue
                ct = g.get("wire_ct") or g.get("count") or 1
                for sk, side in (("s_symbol", "L"), ("d_symbol", "R")):
                    cur = g.get(sk)
                    if not cur:
                        continue
                    best, exact = _lc.snap_symbol(g.get("type"), ct, side, cur)
                    if best and not exact:
                        g[sk] = best
                        g.setdefault("flags", []).append(sk + "_snapped_to_picklist")
    except Exception:
        pass
    # OFFLINE terminal generators (run after colors are set so the phase 4th
    # slot resolves GND vs N; only populate S/D Term, never touch fill logic):
    #   panelboard branch circuits -> breaker/circuit numbers on the panel side
    #   3-phase feeders            -> ØA/ØB/ØC (+GND/N) phases on both ends
    try:
        import idp_terms
        idp_terms.apply_panelboard_circuits(records, panelboard=_panelboard_map)
        idp_terms.apply_source_info(records)   # terms + descriptions on every group
        idp_terms.apply_power_terminals(records)  # device-specific ATS N/E/L terminals
        idp_terms.apply_supporting_docs(records, schedule_doc=_schedule_doc)  # SUPPORTING DOCUMENTS refs
    except Exception:
        pass

    if template_path != out_path:
        shutil.copyfile(template_path, out_path)

    wb = openpyxl.load_workbook(out_path, keep_vba=True, data_only=False)
    ci = wb["ConduitIndex"]
    rd = wb["Ref Documents & Deviations"]
    fi = wb["FillIndex"]

    # honor the add_flags option: no-op the flagger when disabled
    flag = _flag_cell if add_flags else (lambda *a, **k: None)

    # Capture the per-row Wire Label TEXTJOIN formulas from the first data row so
    # we can re-apply them (row-translated) to each row we write — computed labels
    # survive the clear. If the blank template has none, the workbook's own VBA
    # fills them on open in Excel.
    label_tmpl = {}
    for cidx, col in FI_LABELS.items():
        v = fi.cell(3, cidx).value
        if isinstance(v, str) and v.startswith("="):
            label_tmpl[cidx] = (col, v)

    def _clear(ws, start_row, ncols):
        last = _last_data_row(ws, list(range(1, ncols + 1)), start_row)
        for r in range(start_row, last + 1):
            for c in range(1, ncols + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).value = None

    if clear_rows:
        _clear(ci, 2, CI_NCOLS)    # ConduitIndex A..K
        _clear(fi, 3, FI_NCOLS)    # FillIndex A..BH
        _clear(rd, 3, 4)           # Ref Documents A..D (drop template's stale refs)
        ci_start, fi_start = 2, 3
    else:
        ci_start = _last_data_row(ci, list(range(1, CI_NCOLS + 1)), 2) + 1
        fi_start = _last_data_row(fi, list(range(1, FI_NCOLS + 1)), 3) + 1

    # ---- deduped global Ref Documents & Deviations, numeric index in col A ----
    doc_index = {}
    rd_start = 3
    rd_last = _last_data_row(rd, [1, 2, 3, 4], rd_start)
    next_idx = 1
    for r in range(rd_start, rd_last + 1):
        dwg = rd.cell(r, 2).value or ""
        desc = rd.cell(r, 3).value or ""
        manu = rd.cell(r, 4).value or ""
        idx = rd.cell(r, 1).value
        if (dwg, desc, manu) != ("", "", ""):
            doc_index[(str(dwg), str(desc), str(manu))] = idx
            try:
                next_idx = max(next_idx, int(idx) + 1)
            except (TypeError, ValueError):
                next_idx += 1
    rd_write = rd_last + 1

    def ref_index(doc):
        nonlocal rd_write, next_idx
        key = (str(doc[0]), str(doc[1]), str(doc[2]))
        if key in doc_index:
            return doc_index[key]
        idx = next_idx
        rd.cell(rd_write, 1).value = idx
        rd.cell(rd_write, 2).value = doc[0]
        rd.cell(rd_write, 3).value = doc[1]
        rd.cell(rd_write, 4).value = doc[2]
        doc_index[key] = idx
        next_idx += 1
        rd_write += 1
        return idx

    ci_row = ci_start
    fi_row = fi_start

    for rec in records:
        # schedule-layout rows: mirror the source schedule's own formatting.
        #   layout="blank"   -> an empty separator row between letter-groups
        #   layout="notused" -> a reserved-but-unused conduit tag (e.g. L004)
        # These carry no fill, so the FillIndex is untouched.
        lay = rec.get("layout")
        if lay == "blank":
            ci_row += 1
            continue
        if lay == "notused":
            ci.cell(ci_row, CI_NAME).value = rec.get("name", "")
            note = (rec.get("source") or ["NOT USED"])[0] or "NOT USED"
            ci.cell(ci_row, CI_SRC1).value = note
            ci_row += 1
            continue

        idxs = [str(ref_index(d)) for d in rec.get("docs", [])]
        src = list(rec["source"]) + [""] * 3
        dst = list(rec["dest"]) + [""] * 3
        ci.cell(ci_row, CI_NAME).value = rec["name"]
        ci.cell(ci_row, CI_SRC1).value = src[0]
        ci.cell(ci_row, CI_SRC2).value = src[1]
        ci.cell(ci_row, CI_SRC3).value = src[2]
        ci.cell(ci_row, CI_DST1).value = dst[0]
        ci.cell(ci_row, CI_DST2).value = dst[1]
        ci.cell(ci_row, CI_DST3).value = dst[2]
        ci.cell(ci_row, CI_SIZE).value = rec["size"]
        ci.cell(ci_row, CI_TYPE).value = rec["ctype"]
        ci.cell(ci_row, CI_REF).value = ", ".join(idxs)
        ci.cell(ci_row, CI_DEV).value = "" if clear_deviations else rec.get("deviations", "")

        # --- flag uncertain ConduitIndex values ---
        rflags = set(rec.get("flags", []))
        if str(rec["ctype"]).strip() in ("XXX", "PER SPEC", ""):
            flag(ci, ci_row, CI_TYPE,
                       "Conduit Type not stated on the source document; placeholder set. "
                       "Confirm against the spec / general note before import.")
        if not str(rec["size"]).strip():
            flag(ci, ci_row, CI_SIZE, "Conduit Size not found on the source; verify.")
        if "derived_from_cable_schedule" in rflags:
            flag(ci, ci_row, CI_NAME,
                       "Conduit derived from the cable schedule ROUTING column "
                       "(no conduit schedule present); verify it is a real conduit.")
        if "endpoint_bleed" in rflags:
            flag(ci, ci_row, CI_SRC1,
                       "Source name recovered from garbled/wrapped schedule text; verify.")
            flag(ci, ci_row, CI_DST1,
                       "Destination name recovered from garbled/wrapped schedule text; verify.")
        for note in archetype_notes.get(rec["name"], []):
            flag(ci, ci_row, CI_NAME, "ARCHETYPE: " + note)
        ci_row += 1

        wires = list(rec.get("wires", []))
        wi = 0
        for grp in rec["fill"]:
            colors = (grp.get("colors") or []) + [""] * 4
            fi.cell(fi_row, FI_CONDUIT).value = rec["name"]
            fi.cell(fi_row, FI_WIRECT).value = grp.get("wire_ct") or grp.get("count")
            fi.cell(fi_row, FI_TYPE).value = grp["type"]
            fi.cell(fi_row, FI_GAUGE).value = grp["gauge"]
            if grp.get("s_symbol"):
                fi.cell(fi_row, FI_SSYM).value = grp["s_symbol"]
            if grp.get("d_symbol"):
                fi.cell(fi_row, FI_DSYM).value = grp["d_symbol"]
            for k in range(4):
                fi.cell(fi_row, FI_COLOR[k]).value = colors[k]

            # consume EXACTLY the wires stored for this group. When "slots" is
            # set (idp_excel workbook round-trip), it's the real per-group wire
            # count and may be 0 — honor it exactly so later groups don't shift.
            # Otherwise (drawing/heuristic records with an empty wire list) fall
            # back to the color-count proxy, min 1.
            if "slots" in grp:
                nslots = max(0, min(int(grp["slots"]), 4))
            else:
                nslots = min(max(len(grp.get("colors") or []), 1), 4)
            gw = wires[wi:wi + nslots]
            # FILLWIRELABEL uses the device block's TAG1 as the label's middle field.
            # Fill any conductor whose tag is blank with the symbol's device tag
            # (XFMR/CB/MTR/…) so both the block tag AND the wire label carry it —
            # a specific tag already set (CB-MAIN, GND, panel circuit) is preserved.
            _sdev = idp_terms.device_tag_for_symbol(grp.get("s_symbol"))
            _ddev = idp_terms.device_tag_for_symbol(grp.get("d_symbol"))
            raw_s = [((w["src"][1] or _sdev) or "") for w in gw]
            raw_d = [((w["dst"][1] or _ddev) or "") for w in gw]
            # block Tag cells: suppress repeats (a 3-pole device shows its tag once)
            s_tags = _suppress_repeat_tags(raw_s)
            d_tags = _suppress_repeat_tags(raw_d)
            for k, w in enumerate(gw[:4]):
                fi.cell(fi_row, FI_STAG[k]).value = s_tags[k]
                fi.cell(fi_row, FI_STERM[k]).value = w["src"][2]
                fi.cell(fi_row, FI_DTAG[k]).value = d_tags[k]
                fi.cell(fi_row, FI_DTERM[k]).value = w["dst"][2]
            # A ground group ALWAYS carries "GND" as its S/D Tag with a blank term —
            # that is the ONLY signal LISA uses to classify a fill as GROUND. Force it
            # here (the reliable single point that knows the group boundary), because
            # conduits that already have EDC/panelboard terminals skip wire synthesis
            # in apply_source_info and would otherwise leave the ground untagged.
            if idp_terms._is_ground_group(grp):
                for k in range(min(max(nslots, 1), 4)):
                    fi.cell(fi_row, FI_STAG[k]).value = "GND"
                    fi.cell(fi_row, FI_STERM[k]).value = None
                    fi.cell(fi_row, FI_DTAG[k]).value = "GND"
                    fi.cell(fi_row, FI_DTERM[k]).value = None
            wi += nslots

            # --- Rating / Fuse Rating / Description (general — any device) ---
            if grp.get("s_rating"):
                fi.cell(fi_row, FI_SRATING).value = grp["s_rating"]
            if grp.get("s_fuse_rating"):
                fi.cell(fi_row, FI_SFUSE).value = grp["s_fuse_rating"]
            for k, v in enumerate(grp.get("s_desc", [])[:3]):
                if v:
                    fi.cell(fi_row, FI_SDESC[k]).value = v
            if grp.get("d_rating"):
                fi.cell(fi_row, FI_DRATING).value = grp["d_rating"]
            if grp.get("d_fuse_rating"):
                fi.cell(fi_row, FI_DFUSE).value = grp["d_fuse_rating"]
            for k, v in enumerate(grp.get("d_desc", [])[:3]):
                if v:
                    fi.cell(fi_row, FI_DDESC[k]).value = v

            # --- ISATag_Loop#/ElementID/Element#/FunctionID — INSTRUMENT blocks only
            # (LISA_workbook_mapper.py: these map onto the instrument bubble; writing
            # them on a non-instrument symbol is inert in LISA, but gate anyway so the
            # data only appears where it's meaningful) ---
            if "inst" in str(grp.get("s_symbol") or "").lower():
                for col, key in zip(FI_SISATAG, ("s_isa_loop", "s_isa_elem",
                                                  "s_isa_elemnum", "s_isa_func")):
                    if grp.get(key):
                        fi.cell(fi_row, col).value = grp[key]
            if "inst" in str(grp.get("d_symbol") or "").lower():
                for col, key in zip(FI_DISATAG, ("d_isa_loop", "d_isa_elem",
                                                  "d_isa_elemnum", "d_isa_func")):
                    if grp.get(key):
                        fi.cell(fi_row, col).value = grp[key]

            # --- S/D Type + S/D Quantity — SPARE blocks only (LISA_workbook_mapper.py:
            # Src_SpareType/Qty, Dst_SpareType/Qty on Spare_L/Spare_R; NOT a general
            # device type/quantity field) ---
            if "spare" in str(grp.get("s_symbol") or "").lower():
                if grp.get("s_spare_type"):
                    fi.cell(fi_row, FI_STYPE).value = grp["s_spare_type"]
                if grp.get("s_spare_qty"):
                    fi.cell(fi_row, FI_SQTY).value = grp["s_spare_qty"]
            if "spare" in str(grp.get("d_symbol") or "").lower():
                if grp.get("d_spare_type"):
                    fi.cell(fi_row, FI_DTYPE).value = grp["d_spare_type"]
                if grp.get("d_spare_qty"):
                    fi.cell(fi_row, FI_DQTY).value = grp["d_spare_qty"]

            # Wire Label = the terminal cross-reference the finished IDPs print on each
            # conductor: "SrcName:STag:STerm / DstName:DTag:DTerm". The template carries
            # this as a TEXTJOIN formula, but openpyxl can't evaluate it, so its cached
            # value is blank and LISA (which reads cached values) fell back to printing
            # the wire TYPE ("POWER"). Write the computed LITERAL instead — mirrors the
            # template's TEXTJOIN(":",TRUE,…) skip-blanks + drop-name-when-equal-to-tag.
            # Wire Label per conductor = FILLWIRELABEL's grammar:
            #   SrcName:SrcTag:SrcTerm / DstName:DstTag:DstTerm
            # tag on EVERY conductor (device tag, NOT the suppressed block tag — the
            # finished IDPs show XFMR/LANDING LUG on each row); name dropped when it
            # equals the tag; blank fields collapse (join non-empty with ":").
            _src_nm = str((rec.get("source") or [""])[0] or "").strip()
            _dst_nm = str((rec.get("dest") or [""])[0] or "").strip()
            _lbl_cols = sorted(FI_LABELS)
            _gnd_grp = idp_terms._is_ground_group(grp)
            for k in range(min(max(nslots, 0), 4)):
                w = gw[k] if k < len(gw) else {"src": ("", "", ""), "dst": ("", "", "")}
                _stag = "GND" if _gnd_grp else (raw_s[k] if k < len(raw_s) else "")
                _dtag = "GND" if _gnd_grp else (raw_d[k] if k < len(raw_d) else "")
                _sterm = "" if _gnd_grp else w["src"][2]
                _dterm = "" if _gnd_grp else w["dst"][2]
                _left = ":".join(str(p) for p in (
                    (_src_nm if _src_nm and _src_nm != _stag else ""),
                    _stag, _sterm) if p not in (None, ""))
                _right = ":".join(str(p) for p in (
                    (_dst_nm if _dst_nm and _dst_nm != _dtag else ""),
                    _dtag, _dterm) if p not in (None, ""))
                _label = (_left + "/" + _right) if (_left or _right) else None
                fi.cell(fi_row, _lbl_cols[k]).value = _label
            for mc in FI_MODES:
                fi.cell(fi_row, mc).value = "DEFAULT"

            # --- flag uncertain FillIndex values ---
            if str(grp.get("type")) == "MFG_CABLE":
                flag(fi, fi_row, FI_TYPE,
                           "MFG_CABLE assumed from conductor count; may instead be "
                           "individual CONTROL/POWER wires — verify vs the wiring diagram.")
            if not gw:
                flag(fi, fi_row, FI_STAG[0],
                           "Terminations (S/D Tag & Term) not available from the cable "
                           "schedule; populate from the wiring diagram / FILLWIRELABEL.")
            if grp.get("s_symbol") and grp.get("s_symbol_conf", 1) < 0.6:
                flag(fi, fi_row, FI_SSYM,
                           "S Symbol inferred from the device name / library; verify vs the drawing.")
            if grp.get("d_symbol") and grp.get("d_symbol_conf", 1) < 0.6:
                flag(fi, fi_row, FI_DSYM,
                           "D Symbol inferred from the device name / library; verify vs the drawing.")
            if grp.get("connection_remodel"):
                flag(fi, fi_row, FI_WIRECT,
                           "Connection-count remodel (raw conductor count is not a "
                           "valid Wire Ct): " + grp["connection_remodel"] +
                           ". Confirm connections vs the wiring diagram.")
            if grp.get("color_note"):
                flag(fi, fi_row, FI_COLOR[0], grp["color_note"])
            if grp.get("gauge_note"):
                flag(fi, fi_row, FI_GAUGE, grp["gauge_note"])
            if grp.get("ground_note"):
                flag(fi, fi_row, FI_TYPE, grp["ground_note"])
            if grp.get("type_note"):
                flag(fi, fi_row, FI_TYPE, grp["type_note"])
            if grp.get("s_desc_note"):
                flag(fi, fi_row, FI_SDESC[0], grp["s_desc_note"])
            if grp.get("d_desc_note"):
                flag(fi, fi_row, FI_DDESC[0], grp["d_desc_note"])

            fi_row += 1

    # Sweep every cell we just wrote for accidental formula-leaks (see
    # _sanitize_formula_leaks) before saving. FI_LABELS columns carry real
    # TEXTJOIN formulas and are deliberately excluded.
    _sanitize_formula_leaks(ci, ci_start, ci_row - 1, CI_NCOLS)
    _sanitize_formula_leaks(fi, fi_start, fi_row - 1, FI_NCOLS, formula_cols=set(FI_LABELS.keys()))
    _sanitize_formula_leaks(rd, rd_start, rd_write - 1, 4)

    # LISA's parser finds the supporting-docs sheet ONLY by the exact name
    # "Ref Documents" / "Ref Docs" (parser._parse_ref_docs + _parse_deviation_notes);
    # the template names it "Ref Documents & Deviations", so LISA returns ref_docs={}
    # and every SUPPORTING DOCUMENTS table renders blank. Rename it in the OUTPUT copy
    # (template untouched) so LISA reads both ref docs and deviation notes. Re-point the
    # one defined name (RefDoc_List dropdown source) so the workbook stays self-consistent.
    _OLD_RD, _NEW_RD = "Ref Documents & Deviations", "Ref Documents"
    if _OLD_RD in wb.sheetnames and _NEW_RD not in wb.sheetnames:
        wb[_OLD_RD].title = _NEW_RD
        try:
            dn = wb.defined_names.get("RefDoc_List")
            if dn is not None and _OLD_RD in str(dn.value):
                dn.value = str(dn.value).replace(f"'{_OLD_RD}'", f"'{_NEW_RD}'")
        except Exception:
            pass

    wb.save(out_path)
    return out_path
