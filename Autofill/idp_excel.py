"""
idp_excel.py — read Excel sources into extractor `records`.

Handles two shapes:
  1. An existing IDP workbook (has ConduitIndex + FillIndex sheets) -> read
     directly, highest fidelity (conduits + fill + tags/terms).
  2. A tabular conduit/cable LIST (e.g. AIC "IDP - CONDUIT LIST" xlsx) -> map the
     two-row header to canonical fields, group multi-row conduits, and build
     conduit-level records + best-effort fill.

Output records match idp_write's contract (name/source/dest/size/ctype/docs/
deviations/flags/fill/wires), so they flow straight through the gated writer
(normalize_types -> normalize_connections -> anatomy -> flag -> versioned write),
which is what makes LISA render them like the finished IDPs.
"""
from __future__ import annotations

import os
import openpyxl

try:
    import idp_extract
except Exception:
    idp_extract = None


# Conduit-name values that mean "not a real conduit yet" — spare/placeholder rows on a
# conduit LIST/SCHEDULE (e.g. trailing TBD rows). They must not become emitted conduits.
_PLACEHOLDER_NAMES = {"", "TBD", "TBC", "XXX", "XXXX", "N/A", "NA", "-", "--", "?"}


def _placeholder_name(name):
    return str(name or "").strip().upper() in _PLACEHOLDER_NAMES


# ── public entry ─────────────────────────────────────────────────────────────
def read_source(path):
    """Read an Excel file into records. Auto-detects IDP-workbook vs tabular."""
    keep = path.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=keep)
    names = {n.lower(): n for n in wb.sheetnames}
    if "conduitindex" in names and "fillindex" in names:
        return _read_idp_workbook(wb, names)
    # tabular: pick the sheet that looks most like a conduit list
    ws = _best_tabular_sheet(wb)
    return _read_tabular(ws) if ws else []


# ── shape 1: an existing IDP workbook ────────────────────────────────────────
def _read_idp_workbook(wb, names):
    ci = wb[names["conduitindex"]]
    fi = wb[names["fillindex"]]

    # ConduitIndex: cols A-K, data from row 2
    conduits = {}
    order = []
    for r in range(2, ci.max_row + 1):
        name = ci.cell(r, 1).value
        if _placeholder_name(name):        # skip blank + TBD/XXX placeholder rows
            continue
        name = str(name).strip()
        conduits[name] = {
            "name": name,
            "source": [ci.cell(r, 2).value or "", ci.cell(r, 3).value or "", ci.cell(r, 4).value or ""],
            "dest": [ci.cell(r, 5).value or "", ci.cell(r, 6).value or "", ci.cell(r, 7).value or ""],
            "size": ci.cell(r, 8).value or "",
            "ctype": _norm_ctype(ci.cell(r, 9).value or "XXX"),
            "docs": [], "deviations": ci.cell(r, 11).value or "",
            "flags": ["from_excel_idp"], "fill": [], "wires": [],
        }
        order.append(name)

    # FillIndex: data from row 3
    S_TAG, S_TERM = (10, 11, 12, 13), (18, 19, 20, 21)
    D_TAG, D_TERM = (32, 35, 36, 37), (42, 43, 44, 45)
    for r in range(3, fi.max_row + 1):
        cond = fi.cell(r, 1).value
        if cond in (None, ""):
            continue
        cond = str(cond).strip()
        rec = conduits.get(cond)
        if rec is None:
            rec = conduits[cond] = {"name": cond, "source": ["", "", ""], "dest": ["", "", ""],
                                    "size": "", "ctype": "XXX", "docs": [], "deviations": "",
                                    "flags": ["from_excel_idp", "fill_without_conduit"],
                                    "fill": [], "wires": []}
            order.append(cond)
        ct = fi.cell(r, 2).value or 1
        try:
            ct = int(ct)
        except (TypeError, ValueError):
            ct = 1
        colors = [fi.cell(r, c).value for c in (6, 7, 8, 9)]
        colors = [str(c).strip() for c in colors if c not in (None, "")]
        grp = {
            "type": fi.cell(r, 3).value or "CONTROL",
            "count": ct, "wire_ct": ct,
            "gauge": fi.cell(r, 4).value or "",
            "colors": colors,
            "s_symbol": fi.cell(r, 5).value or "", "d_symbol": fi.cell(r, 29).value or "",
            "s_symbol_conf": 1.0, "d_symbol_conf": 1.0,
        }
        # tags/terms -> wires (consumed nslots-per-group by idp_write)
        s_name = rec["source"][0] if rec["source"] else ""
        d_name = rec["dest"][0] if rec["dest"] else ""
        slots = 0
        for k in range(4):
            st = fi.cell(r, S_TAG[k]).value
            dt = fi.cell(r, D_TAG[k]).value
            se = fi.cell(r, S_TERM[k]).value
            de = fi.cell(r, D_TERM[k]).value
            if any(v not in (None, "") for v in (st, dt, se, de)):
                rec["wires"].append({"src": (s_name, str(st or ""), str(se or "")),
                                     "dst": (d_name, str(dt or ""), str(de or ""))})
                slots += 1
        # store the ACTUAL wire count (may be 0). write_workbook consumes the flat
        # wire list positionally by this count — forcing a minimum of 1 here on a
        # 0-wire group would consume the next group's wire and shift every later
        # group's terminations by one (silent corruption on workbook round-trip).
        grp["slots"] = slots
        rec["fill"].append(grp)

    return [conduits[n] for n in order]


# ── shape 2: a tabular conduit/cable list ────────────────────────────────────
_ROLE_KEYS = {
    "conduit_name": ["name", "conduit name", "conduit", "cond name", "tag"],
    "notes": ["notes", "note", "remarks", "remark", "comments", "comment"],
    "fill_type": ["fill type", "wire type", "conductor type", "cable type"],
    "count": ["qty", "quantity", "count", "no. of conductors", "# of conductors"],
    "source_1": ["source line 1", "source name 1", "source 1", "from"],
    "source_2": ["source line 2", "source name 2", "source 2"],
    "source_3": ["source line 3", "source name 3", "source 3"],
    "dest_1": ["destination line 1", "destination name 1", "dest line 1", "to"],
    "dest_2": ["destination line 2", "destination name 2", "dest line 2"],
    "dest_3": ["destination line 3", "destination name 3", "dest line 3"],
    "conduit_type": ["type", "conduit type", "cond type"],
    "conduit_size": ["size", "conduit size", "trade size"],
    "s_tag": ["source tag field 1", "source tag", "s tag 1", "source tag 1"],
    "d_tag": ["destination tag field 1", "destination tag", "d tag 1", "dest tag 1"],
    "color": ["color", "colour", "wire color"],
    "wire_size": ["size", "wire size", "gauge", "awg"],
    "s_component": ["source side component", "source component", "source symbol"],
    "d_component": ["destination side component", "destination component", "destination symbol"],
    "s_desc": ["source side description", "source description"],
    "d_desc": ["destination side description", "destination description"],
}


def _best_tabular_sheet(wb):
    best, score = None, -1
    for ws in wb.worksheets:
        s = 0
        for r in range(1, min(ws.max_row, 4) + 1):
            row = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 30) + 1)).upper()
            for kw in ("NAME", "SOURCE", "DESTINATION", "TYPE", "SIZE", "CONDUIT"):
                if kw in row:
                    s += 1
        if s > score:
            best, score = ws, s
    return best if score > 0 else (wb.worksheets[0] if wb.worksheets else None)


def _build_roles(ws, hdr_top, two_row):
    """Assign a role to each column; two_row merges the sub-header row into the label."""
    roles, seen_color = {}, False
    for c in range(1, ws.max_column + 1):
        g = str(ws.cell(hdr_top, c).value or "").strip()
        sub = str(ws.cell(hdr_top + 1, c).value or "").strip() if two_row else ""
        combined = (g + " " + sub).strip().lower()
        if not combined:
            continue
        role = _match_role(combined, seen_color)
        if role == "color":
            seen_color = True
        if role and role not in roles.values():
            roles[c] = role
        elif role in ("s_tag", "d_tag"):   # allow multiple tag fields collapsed
            roles[c] = role
    return roles


def _colmap(ws):
    """Map column index -> role. Auto-detects a 1-row (clean export) vs 2-row
    (merged CAD group+sub) header by keeping whichever maps more columns."""
    hdr_top = 1
    for r in range(1, min(ws.max_row, 6) + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)).upper()
        if "NAME" in joined or ("SOURCE" in joined and "DESTINATION" in joined):
            hdr_top = r
            break
    one = _build_roles(ws, hdr_top, two_row=False)
    two = _build_roles(ws, hdr_top, two_row=True)
    if len(two) > len(one):
        return two, hdr_top + 2
    return one, hdr_top + 1


def _match_role(text, seen_color):
    # gauge/AWG is always the wire gauge, regardless of column order
    if text in ("gauge", "awg") or text.endswith(" gauge") or "awg" in text:
        return "wire_size"
    # resolve "*type*" columns explicitly (else "conduit type" is grabbed by the
    # loose "conduit" alias of conduit_name and dropped as a duplicate)
    if "type" in text:
        if any(k in text for k in ("fill", "wire", "conductor", "cable")):
            return "fill_type"
        return "conduit_type"
    # SIZE is ambiguous: a wire/conductor size is the gauge; conduit size is the trade size;
    # a bare "size" resolves by position (after COLOR = gauge, before = conduit).
    if text == "size" or text.endswith(" size"):
        if text.startswith("wire") or text.startswith("conductor"):
            return "wire_size"
        if text.startswith("conduit") or text.startswith("trade"):
            return "conduit_size"
        return "wire_size" if seen_color else "conduit_size"
    for role, keys in _ROLE_KEYS.items():
        if role in ("conduit_size", "wire_size"):
            continue
        for k in keys:
            if text == k or text.startswith(k) or k in text:
                return role
    return None


def _read_tabular(ws):
    roles, data_start = _colmap(ws)
    inv = {}
    for c, role in roles.items():
        inv.setdefault(role, c)
    if "conduit_name" not in inv:
        return []   # not a recognizable conduit list

    records, cur = [], None

    def cell(r, role):
        c = inv.get(role)
        v = ws.cell(r, c).value if c else None
        return "" if v in (None,) else str(v).strip()

    def flush(rec):
        if rec:
            _finalize_tabular(rec)
            records.append(rec)

    for r in range(data_start, ws.max_row + 1):
        name = cell(r, "conduit_name")
        s_tag, d_tag = cell(r, "s_tag"), cell(r, "d_tag")
        color, wsize = cell(r, "color"), cell(r, "wire_size")
        s_comp, d_comp = cell(r, "s_component"), cell(r, "d_component")
        ftype, fcount = cell(r, "fill_type"), cell(r, "count")
        line_has_wire = any([s_tag, d_tag, color, wsize, s_comp, d_comp, ftype, fcount])
        if _placeholder_name(name) and str(name or "").strip():
            # a NAMED placeholder row (TBD/XXX): end the current conduit and DROP this row
            # entirely — its wire cells must NOT be appended to the previous real conduit.
            flush(cur); cur = None
            continue
        if name:                       # new conduit
            flush(cur)
            cur = {
                "name": name,
                "source": [cell(r, "source_1"), cell(r, "source_2"), cell(r, "source_3")],
                "dest": [cell(r, "dest_1"), cell(r, "dest_2"), cell(r, "dest_3")],
                "size": cell(r, "conduit_size"),
                "ctype": _norm_ctype(cell(r, "conduit_type") or "XXX"),
                "docs": [], "deviations": cell(r, "notes"),   # carry panel-circuit notes
                "flags": ["from_excel_list"], "fill": [], "wires": [], "_lines": [],
            }
        if cur is None:
            continue
        if line_has_wire:
            cur["_lines"].append({"s_tag": s_tag, "d_tag": d_tag, "color": color,
                                  "wire_size": wsize, "s_comp": s_comp, "d_comp": d_comp,
                                  "fill_type": ftype, "count": fcount,
                                  "s_desc": cell(r, "s_desc"), "d_desc": cell(r, "d_desc")})
    flush(cur)
    # If the schedule spells out grounds for any conduit it is authoritative about
    # grounds for all of them — a conduit with none listed genuinely has none.
    if any(g.get("is_ground") for r in records for g in r.get("fill", [])):
        for r in records:
            r["ground_authoritative"] = True
    return records


def _finalize_tabular(rec):
    """Turn accumulated wire lines into fill groups + wires; infer type/symbols."""
    lines = rec.pop("_lines", [])
    s_name = rec["source"][0] if rec["source"] else ""
    d_name = rec["dest"][0] if rec["dest"] else ""
    if not lines:
        rec["flags"].append("no_fill_lines")   # ConduitIndex only; flag for review
        return
    for ln in lines:
        typ = _norm_fill_type(ln.get("fill_type")) or _infer_fill_type(
            rec["name"], ln.get("s_comp"), ln.get("d_comp"),
            ln.get("s_desc"), ln.get("d_desc"))
        colors = [ln["color"]] if ln.get("color") else []
        try:
            cnt = int(float(str(ln.get("count")).strip())) if str(ln.get("count") or "").strip() else 1
        except ValueError:
            cnt = 1
        cnt = max(cnt, 1)
        if typ == "GROUND":
            # LISA has no GROUND *type* — the legal representation of an equipment
            # ground is POWER / wire-ct 1 / GND_L·GND_R / GRN (LISA labels it
            # "GROUND" in the diagram from the GND symbol). Keep the schedule's
            # real ground gauge (#2/o, #6, ...) instead of a fabricated default.
            grp = {"type": "POWER", "count": 1, "wire_ct": 1,
                   "gauge": ln.get("wire_size") or "", "colors": ["GRN"], "slots": 1,
                   "s_symbol": "GND_L", "d_symbol": "GND_R",
                   "s_symbol_conf": 0.95, "d_symbol_conf": 0.95, "is_ground": True}
            rec["fill"].append(grp)
            continue
        if typ == "NEUTRAL" and not colors:
            colors = ["WHT"]
        grp = {"type": typ, "count": cnt, "wire_ct": cnt, "gauge": ln.get("wire_size") or "",
               "colors": colors, "slots": cnt}
        # symbols: prefer explicit components, else infer from device names
        if idp_extract is not None:
            try:
                idp_extract._attach_symbols([grp], ln.get("s_comp") or s_name,
                                            ln.get("d_comp") or d_name)
            except Exception:
                pass
        rec["fill"].append(grp)
        rec["wires"].append({"src": (s_name, ln.get("s_tag", ""), ""),
                             "dst": (d_name, ln.get("d_tag", ""), "")})
    rec["flags"].append("terms_from_diagram")   # conduit lists lack terminals


def _infer_fill_type(name, s_comp, d_comp, s_desc, d_desc):
    blob = " ".join(x for x in (s_comp, d_comp, s_desc, d_desc) if x).upper()
    if any(k in blob for k in ("TSP", "SHLD", "TWISTED", "4-20", "ANALOG", "TRANSMITTER")):
        return "TSP"
    if any(k in blob for k in ("CAT", "ETHERNET", "NETWORK", "RJ45")):
        return "CAT-6"
    if any(k in blob for k in ("FIBER", "FIBRE", "SC", "LC", "MTP")):
        return "FIBER"
    if "PULL" in blob and "ROPE" in blob:
        return "PULL_ROPE"
    n = (name or "").upper()
    if n[:1] in ("P", "H", "L"):
        return "POWER"
    if n[:1] in ("C",):
        return "CONTROL"
    if n[:1] in ("A",):
        return "TSP"
    return "CONTROL"


def _norm_fill_type(s):
    """Normalize an explicit fill-type cell to the canonical group type names."""
    if not s:
        return ""
    t = str(s).strip().upper().replace(" ", "")
    aliases = {
        "PULLROPE": "PULL_ROPE", "PULLROPES": "PULL_ROPE", "PULL_ROPE": "PULL_ROPE",
        "2C/16STP": "TSP", "STP": "TSP", "TSP": "TSP", "SHLD": "TSP",
        "CAT-6SHLD": "CAT-6", "CAT6": "CAT-6", "CAT-6": "CAT-6",
        "MFRCABLE": "MFG_CABLE", "MFGCABLE": "MFG_CABLE", "MFG_CABLE": "MFG_CABLE",
        "PWR": "POWER", "POWER": "POWER", "CONTROL": "CONTROL", "FIBER": "FIBER",
        "GND": "GROUND", "GROUND": "GROUND", "GRND": "GROUND", "EGC": "GROUND",
        "NEUTRAL": "NEUTRAL", "NEUT": "NEUTRAL", "N": "NEUTRAL",
    }
    return aliases.get(t, str(s).strip().upper())


def _norm_ctype(s):
    if idp_extract is not None:
        try:
            return idp_extract._norm_ctype(s)
        except Exception:
            pass
    return str(s or "XXX").strip().upper()


if __name__ == "__main__":
    import sys
    p = sys.argv[1]
    recs = read_source(p)
    print(f"{os.path.basename(p)} -> {len(recs)} conduits, "
          f"{sum(len(r['fill']) for r in recs)} fill rows")
    for rec in recs[:5]:
        print(f"  {rec['name']:<8} {rec['ctype']:<6} src={rec['source'][0]!r} "
              f"dst={rec['dest'][0]!r} fill={len(rec['fill'])}")
