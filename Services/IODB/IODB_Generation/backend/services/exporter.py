"""
exporter.py — Build an IOList xlsx from a validated io_list payload.
"""

import glob
import io
import os

import openpyxl

TEMPLATE_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), '..', '..', '..', '_Templates')
)

COLUMNS = [
    "ProcIdent", "EquipProcLoc", "Descriptor", "LoopNumber",
    "IOType", "Comment", "RangeMin", "Range Max", "Units",
]

KEY_TO_HEADER = {
    "proc_ident":     "ProcIdent",
    "equip_proc_loc": "EquipProcLoc",
    "descriptor":     "Descriptor",
    "loop_number":    "LoopNumber",
    "io_type":        "IOType",
    "comment":        "Comment",
    "range_min":      "RangeMin",
    "range_max":      "Range Max",
    "units":          "Units",
}


def _latest_template() -> str:
    pattern = os.path.join(TEMPLATE_DIR, "Template_IODB_*.xlsx")
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No template found in {TEMPLATE_DIR}")
    return matches[-1]


def export(payload: dict) -> io.BytesIO:
    template_path = _latest_template()
    wb = openpyxl.load_workbook(template_path)

    if "IOList" not in wb.sheetnames:
        raise ValueError("Template does not contain an 'IOList' sheet")

    ws = wb["IOList"]

    header_row_idx = None
    col_positions: dict = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell == "ProcIdent" for cell in row if cell is not None):
            header_row_idx = row_idx
            for col_idx, cell in enumerate(row, start=1):
                if cell in COLUMNS:
                    col_positions[cell] = col_idx
            break

    if header_row_idx is None:
        raise ValueError("Could not locate header row ('ProcIdent') in IOList sheet")

    if ws.max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, ws.max_row - header_row_idx)

    points = payload.get("io_list", [])
    for point in points:
        row_data = {}
        for json_key, header in KEY_TO_HEADER.items():
            col_idx = col_positions.get(header)
            if col_idx is not None:
                row_data[col_idx] = point.get(json_key)
        if row_data:
            max_col = max(col_positions.values())
            flat = [row_data.get(c) for c in range(1, max_col + 1)]
            ws.append(flat)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
