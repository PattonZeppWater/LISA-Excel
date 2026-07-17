"""
parser.py — IODB xlsx workbook → JSON sheet data
"""

import io
import base64
import openpyxl
from datetime import datetime, date

HIDDEN_SHEETS = {
    "obj_classdescriptions",
    "obj_masterparameterlist",
    "validation",
    "document control",
}

READONLY_SHEETS = set()

PART_TYPES = [
    "CPU", "CPS", "Rack", "DICard", "DOCard",
    "AICard", "AOCard", "CommunicationCard", "Other", "Spare",
]


def _serialize(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def _parse_cover_as_kv(ws) -> dict:
    rows: list = []
    process_areas: list = []
    in_pa = False

    for row in ws.iter_rows(values_only=True):
        col_a  = row[0] if len(row) > 0 else None
        label  = row[1] if len(row) > 1 else None
        abbrev = row[2] if len(row) > 2 else None
        name   = row[3] if len(row) > 3 else None

        if col_a is not None and isinstance(col_a, str) and col_a.strip():
            break

        if isinstance(label, str) and label.strip() == "Process Areas":
            in_pa = True
            continue

        if in_pa:
            if label is not None and isinstance(label, str) and label.strip():
                in_pa = False
            else:
                a = str(abbrev).strip() if abbrev is not None else ""
                n = str(name).strip()   if name   is not None else ""
                if not (a in ("", "-") and n in ("", "-")):
                    process_areas.append([a, n])
                continue

        if label is None or not isinstance(label, str):
            continue
        label = label.strip()
        if len(label) <= 1 or label.startswith("-"):
            continue
        rows.append([label, _serialize(abbrev) if abbrev is not None else ""])

    return {
        "headers":       ["Label", "Value"],
        "rows":          rows,
        "process_areas": process_areas,
        "readonly":      False,
        "hidden":        False,
    }


def _extract_parts_catalog(wb) -> dict:
    if "Validation" not in wb.sheetnames:
        return {}

    ws = wb["Validation"]
    catalog: dict[str, list] = {}
    in_table = False

    for row in ws.iter_rows(values_only=True):
        non_null = [c for c in row if c is not None]
        if not non_null:
            in_table = False
            continue

        if isinstance(non_null[0], str) and non_null[0].strip() == "PartNumber":
            in_table = True
            continue

        if not in_table:
            continue

        part_num       = row[0] if len(row) > 0 else None
        classification = row[1] if len(row) > 1 else None
        description    = row[2] if len(row) > 2 else None
        io_size        = row[3] if len(row) > 3 else None

        if not part_num or not classification:
            continue
        if not isinstance(part_num, (str, int, float)):
            continue
        if not isinstance(classification, str):
            continue

        key = classification.strip()
        if key not in catalog:
            catalog[key] = []
        catalog[key].append({
            "partNumber":  str(part_num).strip(),
            "description": str(description).strip() if description is not None else "",
            "ioSize":      int(io_size) if isinstance(io_size, (int, float)) else 0,
        })

    return catalog


def parse_workbook(wb_bytes: bytes, filename: str = "workbook.xlsx") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)

    sheets = {}
    sheet_order = []
    parts_catalog = _extract_parts_catalog(wb)

    for sheet_name in wb.sheetnames:
        key      = sheet_name.lower().strip()
        hidden   = key in HIDDEN_SHEETS
        readonly = key in READONLY_SHEETS

        ws = wb[sheet_name]

        if key == "cover":
            sheets[sheet_name] = _parse_cover_as_kv(ws)
            sheet_order.append(sheet_name)
            continue

        all_rows = list(ws.iter_rows(values_only=True))

        header_idx = next(
            (i for i, r in enumerate(all_rows) if any(c is not None for c in r)),
            None,
        )

        if header_idx is None:
            headers = []
            rows    = []
        else:
            headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]
            rows = []
            for raw in all_rows[header_idx + 1:]:
                if all(c is None for c in raw):
                    continue
                rows.append([_serialize(v) for v in raw])

        sheet_entry = {
            "headers":  headers,
            "rows":     rows,
            "readonly": readonly,
            "hidden":   hidden,
        }

        if key == "plcequipment":
            sheet_entry["validations"] = {"PartType": PART_TYPES}

        sheets[sheet_name] = sheet_entry
        sheet_order.append(sheet_name)

    wb.close()

    return {
        "filename":      filename,
        "original_b64":  base64.b64encode(wb_bytes).decode("utf-8"),
        "sheet_order":   sheet_order,
        "sheets":        sheets,
        "parts_catalog": parts_catalog,
    }
