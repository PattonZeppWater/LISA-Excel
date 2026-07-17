"""
downloader.py — JSON sheets state + original workbook bytes → updated xlsx
"""

import io
import base64
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_TAGDB_SECTION_NAMES = {"READ", "WRITE", "CONFIG", "INTERNAL"}
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="000000")
_SECTION_FONT = Font(color="FFFFFF", bold=True)
_DATA_FONT    = Font(color="000000")
_NO_FILL      = PatternFill(fill_type=None)

_SECTION_E_SEEDS = {"READ": 400001, "WRITE": 401001, "CONFIG": 402101, "INTERNAL": 403101}
_TAGDB_SEEDS     = {"READ": 400001, "WRITE": 401001, "CONFIG": 402101, "INTERNAL": 403101}


def _d_formula(r):
    return (
        f'=_xlfn.IFS(E{r}="","",C{r}<>"BOOL",'
        f'"%MW"&TEXT(E{r}-400001,"0000"),'
        f'C{r}="BOOL","%MW"&TEXT(E{r}-400001,"0000.00"))'
    )


def _e_chain_formula(r):
    p = r - 1
    return (
        f'=IF(C{r}="","",_xlfn.LET('
        f'_xlpm.sz_p,IFERROR(VLOOKUP(C{p},DataTypes_Table,2,FALSE),0),'
        f'_xlpm.sz_n,IFERROR(VLOOKUP(C{r},DataTypes_Table,2,FALSE),2),'
        f'_xlpm.base,IF(C{p}="BOOL",ROUND(E{p},0)+1,E{p}+_xlpm.sz_p),'
        f'IF(C{r}="BOOL",'
        f'IF(C{p}="BOOL",IF(MOD(E{p},1)>=0.15,ROUND(E{p},0)+1,E{p}+0.01),_xlpm.base),'
        f'IF(_xlpm.sz_n>=2,IF(MOD(_xlpm.base,2)=0,_xlpm.base+1,_xlpm.base),_xlpm.base))))'
    )


def _load_datatypes_table(sheets):
    val_key = next((k for k in sheets if k.lower().strip() == "validation"), None)
    if val_key is None:
        return {}
    rows   = sheets[val_key].get("rows", [])
    dt_idx = None
    for i, row in enumerate(rows):
        first = str(row[0] if row else "").strip() if row else ""
        if first.upper() in ("DATATYPES", "DATATYPES_TABLE"):
            dt_idx = i
            break
    if dt_idx is None:
        return {}
    sizes: dict = {}
    for row in rows[dt_idx + 2:]:
        if not row or row[0] is None or str(row[0]).strip() == "":
            break
        name = str(row[0]).strip()
        try:
            size = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0
        except (TypeError, ValueError):
            size = 0.0
        if name:
            sizes[name] = size
    return sizes


def compute_tagdb_addresses(sheets: dict) -> list:
    sizes = _load_datatypes_table(sheets)

    tagdb_key = next((k for k in sheets if k.lower().strip() == "tagdb"), None)
    if tagdb_key is None:
        return []

    sheet   = sheets[tagdb_key]
    headers = sheet.get("headers", [])
    rows    = sheet.get("rows", [])

    h_idx     = {name: i for i, name in enumerate(headers)}
    tag_idx   = h_idx.get("TAG",           0)
    type_idx  = h_idx.get("TYPE",          2)
    plc_idx   = h_idx.get("PLC ADDRESS",   3)
    scada_idx = h_idx.get("SCADA ADDRESS", 4)
    min_len   = max(plc_idx, scada_idx) + 1

    prev_scada = None
    prev_type  = None

    result = []
    for raw in rows:
        row = list(raw)
        while len(row) < min_len:
            row.append(None)

        tag_val  = str(row[tag_idx]  or "").strip().upper()
        type_val = str(row[type_idx] or "").strip()

        if tag_val in _TAGDB_SEEDS:
            scada          = _TAGDB_SEEDS[tag_val]
            row[plc_idx]   = f"%MW{scada - 400001:04d}"
            row[scada_idx] = scada
            prev_scada     = float(scada)
            prev_type      = tag_val
            result.append(row)
            continue

        if not type_val or prev_scada is None:
            result.append(row)
            continue

        sz_p         = sizes.get(prev_type, 0)
        sz_n         = sizes.get(type_val,  2)
        prev_is_bool = prev_type.upper() == "BOOL" if prev_type else False
        curr_is_bool = type_val.upper() == "BOOL"

        if prev_is_bool:
            base = round(prev_scada) + 1
        else:
            base = prev_scada + sz_p

        if curr_is_bool:
            if prev_is_bool:
                bit_num = round((prev_scada % 1) * 100)
                scada = float(round(prev_scada) + 1) if bit_num >= 15 else round(prev_scada + 0.01, 10)
            else:
                scada = float(base)
        else:
            scada = int(base) + 1 if sz_n >= 2 and int(base) % 2 == 0 else int(base)

        if curr_is_bool:
            int_part = int(scada)
            bit_part = round((scada - int_part) * 100)
            plc = f"%MW{int_part - 400001:04d}.{bit_part:02d}"
        else:
            plc = f"%MW{int(scada) - 400001:04d}"

        row[plc_idx]   = plc
        row[scada_idx] = scada
        prev_scada     = scada
        prev_type      = type_val
        result.append(row)

    return result


def _write_tagdb_rows(ws, rows):
    data_c_cells = []

    for row in rows:
        tag_val    = str(row[0]).strip().upper() if row and row[0] is not None else ""
        is_section = tag_val in _TAGDB_SECTION_NAMES

        ws.append(row)
        r = ws.max_row

        ws.cell(row=r, column=4).value = _d_formula(r)
        if is_section:
            ws.cell(row=r, column=5).value = _SECTION_E_SEEDS.get(tag_val)
        else:
            ws.cell(row=r, column=5).value = _e_chain_formula(r)
            data_c_cells.append(f"C{r}")

        for cell in ws[r]:
            if is_section:
                cell.fill = _SECTION_FILL
                cell.font = _SECTION_FONT
            else:
                cell.fill = _NO_FILL
                cell.font = _DATA_FONT

    if data_c_cells:
        dv = DataValidation(
            type="list",
            formula1='INDIRECT("Table7[DataTypes]")',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = " ".join(data_c_cells)
        ws.add_data_validation(dv)


SKIP_ON_WRITE = {
    "obj_classdescriptions", "obj_masterparameterlist",
    "validation", "document control", "scratch",
}


def _write_cover(ws, sheet_data):
    rows          = sheet_data.get("rows", [])
    process_areas = sheet_data.get("process_areas", None)

    label_to_row: dict = {}
    pa_label_row = None

    for excel_row_idx, ws_row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
    ):
        col_b = ws_row[1] if len(ws_row) > 1 else None
        if col_b is not None and isinstance(col_b, str) and col_b.strip():
            label_to_row[col_b.strip()] = excel_row_idx
            if col_b.strip() == "Process Areas":
                pa_label_row = excel_row_idx

    for pair in rows:
        if not pair or len(pair) < 2:
            continue
        label = str(pair[0]).strip() if pair[0] is not None else ""
        value = pair[1]
        if not label or label not in label_to_row:
            continue
        ws.cell(row=label_to_row[label], column=3).value = value

    if process_areas is not None and pa_label_row is not None:
        pa_start = pa_label_row + 1
        pa_end   = pa_start
        for r in range(pa_start, ws.max_row + 1):
            col_b = ws.cell(row=r, column=2).value
            if col_b is not None and str(col_b).strip():
                break
            pa_end = r

        original_slots = pa_end - pa_start + 1

        for i, (abbrev, name) in enumerate(process_areas):
            r = pa_start + i
            ws.cell(row=r, column=3).value = abbrev
            ws.cell(row=r, column=4).value = name

        for i in range(len(process_areas), original_slots):
            r = pa_start + i
            ws.cell(row=r, column=3).value = None
            ws.cell(row=r, column=4).value = None


def build_workbook(original_b64: str, sheets: dict, filename: str = "IODB.xlsx") -> bytes:
    wb_bytes = base64.b64decode(original_b64)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))

    for sheet_name, sheet_data in sheets.items():
        key = sheet_name.lower().strip()

        if key in SKIP_ON_WRITE:
            continue
        if sheet_data.get("hidden"):
            continue
        if sheet_name not in wb.sheetnames:
            continue

        ws   = wb[sheet_name]
        rows = sheet_data.get("rows", [])

        if key == "cover":
            _write_cover(ws, sheet_data)
            continue

        if sheet_data.get("readonly"):
            continue

        headers = sheet_data.get("headers", [])
        if not headers:
            continue

        header_row = 1
        for i, ws_row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
        ):
            if any(str(c).strip() == headers[0] for c in ws_row if c is not None):
                header_row = i
                break

        if ws.max_row >= header_row + 1:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)

        if key == "tagdb":
            _write_tagdb_rows(ws, rows)
        else:
            for row in rows:
                ws.append(row)
                r = ws.max_row
                for cell in ws[r]:
                    cell.fill = _NO_FILL

            if headers and ws.tables:
                end_col = get_column_letter(len(headers))
                end_row = header_row + len(rows)
                for tbl in ws.tables.values():
                    start_cell = tbl.ref.split(":")[0] if ":" in tbl.ref else f"A{header_row}"
                    tbl.ref = f"{start_cell}:{end_col}{end_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
