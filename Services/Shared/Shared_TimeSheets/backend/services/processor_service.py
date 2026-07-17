"""
processor_service.py — Timesheet comparison and highlighting logic.

Supports EXEMPT and NON-EXEMPT AIC timesheets (auto-detected from cells A2/B2).

Detection:
  B2 contains 'NON-EXEMPT'  → non-exempt layout
  A2 contains 'EXEMPT'      → exempt layout

EXEMPT layout:
  Col C (3)  = Job    |  Col D (4)   = Phase  |  Col L (12) = Total Hours
  Data rows: 14–35    |  Preview label row: 13  |  Preview cols: A–O (15)

NON-EXEMPT layout:
  Col D (4)  = Job    |  Col E (5)   = Phase  |  Col R (18) = Total Hours
  Data rows: 14–36    |  Preview label row: 12  |  Preview cols: A–R (18)

Highlight rules (per row, stops at first failure):
  RED    on Job cell   : Job not found in remaining-units data.
  RED    on Phase cell : Job exists but (Job, Phase) pair does not.
  YELLOW on Hours cell : Pair exists but hours exceed RemainingUnits.
"""

import zipfile
from collections import namedtuple
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")


# ── Layout configs ────────────────────────────────────────────────────────────

TimesheetConfig = namedtuple("TimesheetConfig", [
    "type_name",
    "col_job",
    "col_phase",
    "col_hours",
    "data_start",
    "data_end",
    "preview_first",
    "preview_cols",
])

EXEMPT_CONFIG = TimesheetConfig(
    type_name    = "exempt",
    col_job      = 3,
    col_phase    = 4,
    col_hours    = 12,
    data_start   = 14,
    data_end     = 35,
    preview_first= 13,
    preview_cols = 15,
)

NONEXEMPT_CONFIG = TimesheetConfig(
    type_name    = "nonexempt",
    col_job      = 4,
    col_phase    = 5,
    col_hours    = 18,
    data_start   = 14,
    data_end     = 36,
    preview_first= 12,
    preview_cols = 18,
)


def detect_config(ts_bytes: bytes) -> TimesheetConfig:
    """Return the layout config for a timesheet. Raises ValueError if unrecognised."""
    wb = load_workbook(BytesIO(ts_bytes), data_only=True, read_only=True)
    ws = wb.active
    a2 = str(ws.cell(row=2, column=1).value or "").strip().upper()
    b2 = str(ws.cell(row=2, column=2).value or "").strip().upper()
    wb.close()

    if "NON" in b2.replace("-", "").replace(" ", ""):
        return NONEXEMPT_CONFIG
    if "EXEMPT" in a2:
        return EXEMPT_CONFIG

    raise ValueError(
        f"Cannot detect timesheet type: A2={a2!r}, B2={b2!r}. "
        "Expected 'EXEMPT' in A2 or 'NON-EXEMPT' in B2."
    )


# ── Normalisation ─────────────────────────────────────────────────────────────

def normalize(val) -> str:
    """Canonical form for job/phase lookup — handles floats, trailing periods, spaces."""
    if val is None:
        return ""
    s = str(val).strip().rstrip(".").strip()
    if not s:
        return ""
    if s.count(".") <= 1:
        try:
            s = f"{float(s):.10g}"
        except (ValueError, OverflowError):
            pass
    return s


# ── Lookup builder from API rows ──────────────────────────────────────────────

def build_lookup_from_rows(rows: list) -> tuple[dict, set]:
    """
    Build lookup tables from pre-fetched remaining-units rows
    (list of {"Job", "Phase", "RemainingUnits"} dicts from data_service).

    Returns:
        lookup     : { (norm_job, norm_phase): remaining_units_float }
        known_jobs : { norm_job }
    """
    lookup     = {}
    known_jobs = set()

    for item in rows:
        nj  = normalize(item["Job"])
        np_ = normalize(item["Phase"])
        key = (nj, np_)
        ru  = item.get("RemainingUnits", 0)
        lookup[key] = float(ru) if isinstance(ru, (int, float)) else 0.0
        known_jobs.add(nj)

    return lookup, known_jobs


# ── Media patcher ─────────────────────────────────────────────────────────────

def _patch_media(orig_bytes: BytesIO, mod_bytes: BytesIO) -> BytesIO:
    """Restore xl/media/* files openpyxl may have dropped (embedded images)."""
    orig_bytes.seek(0)
    mod_bytes.seek(0)

    try:
        with zipfile.ZipFile(orig_bytes, "r") as orig_zip:
            orig_names = set(orig_zip.namelist())

            if not any(n.startswith("xl/media/") for n in orig_names):
                mod_bytes.seek(0)
                return mod_bytes

            with zipfile.ZipFile(mod_bytes, "r") as mod_zip:
                mod_names = set(mod_zip.namelist())
                missing = [
                    n for n in orig_names
                    if n not in mod_names and (
                        n.startswith("xl/media/") or
                        n.startswith("xl/drawings/") or
                        ("_rels" in n and "drawing" in n.lower())
                    )
                ]

                if not missing:
                    mod_bytes.seek(0)
                    return mod_bytes

                result = BytesIO()
                with zipfile.ZipFile(result, "w", zipfile.ZIP_DEFLATED) as out_zip:
                    for name in mod_zip.namelist():
                        out_zip.writestr(name, mod_zip.read(name))
                    for name in missing:
                        out_zip.writestr(name, orig_zip.read(name))

                result.seek(0)
                return result

    except Exception:
        mod_bytes.seek(0)
        return mod_bytes


# ── Preview ───────────────────────────────────────────────────────────────────

def get_preview_rows(ts_bytes: bytes, cfg: TimesheetConfig) -> list:
    """Return preview rows as a list of string lists (row 0 = header)."""
    wb = load_workbook(BytesIO(ts_bytes), data_only=True)
    ws = wb.active
    result = []
    for row_num in range(cfg.preview_first, cfg.data_end + 1):
        row_cells = []
        for col_num in range(1, cfg.preview_cols + 1):
            val = ws.cell(row=row_num, column=col_num).value
            if val is None:
                row_cells.append("")
            elif isinstance(val, float):
                row_cells.append(str(int(val)) if val == int(val) else str(round(val, 4)))
            else:
                row_cells.append(str(val))
        result.append(row_cells)
    wb.close()
    return result


def get_preview_with_highlights(
    ts_bytes: bytes,
    lookup: dict,
    known_jobs: set,
) -> tuple[list, list, dict]:
    """
    Return (rows, highlights, stats).

    rows:       list of row-lists (strings); row 0 = label/header
    highlights: list of {"row": 0-based, "col": 0-based, "color": "red"|"yellow"}
    stats:      {"red": int, "yellow": int, "ok": int}
    """
    cfg  = detect_config(ts_bytes)
    rows = get_preview_rows(ts_bytes, cfg)

    highlights = []
    ok_count   = 0

    wb_vals = load_workbook(BytesIO(ts_bytes), data_only=True)
    ws_vals = wb_vals.active

    for row_num in range(cfg.data_start, cfg.data_end + 1):
        row_idx   = row_num - cfg.preview_first  # 0-based into rows
        job_val   = ws_vals.cell(row=row_num, column=cfg.col_job).value
        phase_val = ws_vals.cell(row=row_num, column=cfg.col_phase).value
        hours_val = ws_vals.cell(row=row_num, column=cfg.col_hours).value

        if job_val is None and phase_val is None:
            continue

        norm_job   = normalize(job_val)
        norm_phase = normalize(phase_val)

        if not norm_job and not norm_phase:
            continue

        hours = float(hours_val) if isinstance(hours_val, (int, float)) else 0.0
        key   = (norm_job, norm_phase)

        if norm_job and norm_job not in known_jobs:
            highlights.append({"row": row_idx, "col": cfg.col_job - 1,   "color": "red"})
        elif key not in lookup:
            highlights.append({"row": row_idx, "col": cfg.col_phase - 1, "color": "red"})
        elif hours > lookup[key]:
            highlights.append({"row": row_idx, "col": cfg.col_hours - 1, "color": "yellow"})
        else:
            ok_count += 1

    wb_vals.close()

    red_count    = sum(1 for h in highlights if h["color"] == "red")
    yellow_count = sum(1 for h in highlights if h["color"] == "yellow")

    return rows, highlights, {"red": red_count, "yellow": yellow_count, "ok": ok_count}


# ── Core processor ────────────────────────────────────────────────────────────

def process_timesheets(
    ts_bytes: bytes,
    lookup: dict,
    known_jobs: set,
) -> tuple[BytesIO, dict]:
    """
    Apply cell highlights to the timesheet xlsx and return (BytesIO, stats_dict).
    """
    cfg = detect_config(ts_bytes)

    wb_vals = load_workbook(BytesIO(ts_bytes), data_only=True)
    wb_out  = load_workbook(BytesIO(ts_bytes), data_only=False)

    ws_vals = wb_vals.active
    ws_out  = wb_out.active

    stats = {"red": 0, "yellow": 0, "ok": 0}

    for row_num in range(cfg.data_start, cfg.data_end + 1):
        job_val   = ws_vals.cell(row=row_num, column=cfg.col_job).value
        phase_val = ws_vals.cell(row=row_num, column=cfg.col_phase).value
        hours_val = ws_vals.cell(row=row_num, column=cfg.col_hours).value

        if job_val is None and phase_val is None:
            continue

        norm_job   = normalize(job_val)
        norm_phase = normalize(phase_val)

        if not norm_job and not norm_phase:
            continue

        hours = float(hours_val) if isinstance(hours_val, (int, float)) else 0.0
        key   = (norm_job, norm_phase)

        if norm_job and norm_job not in known_jobs:
            ws_out.cell(row=row_num, column=cfg.col_job).fill = RED_FILL
            stats["red"] += 1
        elif key not in lookup:
            ws_out.cell(row=row_num, column=cfg.col_phase).fill = RED_FILL
            stats["red"] += 1
        elif hours > lookup[key]:
            ws_out.cell(row=row_num, column=cfg.col_hours).fill = YELLOW_FILL
            stats["yellow"] += 1
        else:
            stats["ok"] += 1

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    output = _patch_media(BytesIO(ts_bytes), output)
    return output, stats
