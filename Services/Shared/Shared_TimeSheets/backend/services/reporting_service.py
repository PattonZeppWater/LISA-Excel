"""
reporting_service.py — Process a zip of AIC timesheet xlsx files into a
Vista-import-ready CSV.

Supports two sheet layouts (auto-detected):
  EXEMPT    (salaried)  — A2 == "EXEMPT"
  NON-EXEMPT (hourly)   — B2 contains "NON-EXEMPT"

CSV column layout (11 columns):
  1  employee   2  PostDate   3  Job   4  Phase   5  EarnCode
  6  Hours      7  (blank)    8  EMCo  9  (blank) 10 (blank)  11 (blank)

EarnCodes:
  12 — Salary/Jobs  (EXEMPT)
  50 — Reg Time     (NON-EXEMPT straight time)
  51 — Overtime     (NON-EXEMPT overtime, row omitted if 0)
"""

import io
import csv
import zipfile
from datetime import datetime, date


EMCO = 16


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip().rstrip(".")
    return s if s else None


def _format_date(val) -> str:
    if isinstance(val, (datetime, date)):
        return val.strftime("%m/%d/%y")
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%y")
        except ValueError:
            pass
    raise ValueError(f"Unrecognised date value: {val!r} — expected MM/DD/YY or MM/DD/YYYY")


def _cell_float(ws, row: int, col: int) -> float:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


# ── EXEMPT extractor ──────────────────────────────────────────────────────────

def _extract_rows_exempt(ws) -> list[dict]:
    """
    Employee: I8  |  PostDate: I10
    Data rows 14–35: Job=col C(3), Phase=col D(4), Hours=col L(12)
    EarnCode: 12
    """
    employee  = ws["I8"].value
    post_date = _format_date(ws["I10"].value)
    rows = []

    for r in range(14, 36):
        job   = _normalize(ws.cell(row=r, column=3).value)
        phase = _normalize(ws.cell(row=r, column=4).value)
        hours = _cell_float(ws, r, 12)

        if not job or not phase or hours == 0.0:
            continue

        rows.append({
            "employee":  employee,
            "post_date": post_date,
            "job":       job,
            "phase":     phase,
            "earn_code": 12,
            "hours":     hours,
        })

    return rows


# ── NON-EXEMPT extractor ──────────────────────────────────────────────────────

_OT_COLS = [9, 11, 13, 15, 17]  # Mon=I, Tue=K, Wed=M, Thu=O, Fri=Q


def _extract_rows_nonexempt(ws) -> list[dict]:
    """
    Employee: N8  |  PostDate: N10
    Data rows 14–36: Job=col D(4), Phase=col E(5)
      ST hours: col R(18)  → EarnCode 50
      OT hours: sum I,K,M,O,Q → EarnCode 51 (omitted if 0)
    """
    employee  = ws["N8"].value
    post_date = _format_date(ws["N10"].value)
    rows = []

    for r in range(14, 37):
        job   = _normalize(ws.cell(row=r, column=4).value)
        phase = _normalize(ws.cell(row=r, column=5).value)

        if not job or not phase:
            continue

        st_hours = _cell_float(ws, r, 18)
        ot_hours = sum(_cell_float(ws, r, c) for c in _OT_COLS)

        if st_hours != 0.0:
            rows.append({
                "employee":  employee,
                "post_date": post_date,
                "job":       job,
                "phase":     phase,
                "earn_code": 50,
                "hours":     st_hours,
            })

        if ot_hours > 0.0:
            rows.append({
                "employee":  employee,
                "post_date": post_date,
                "job":       job,
                "phase":     phase,
                "earn_code": 51,
                "hours":     ot_hours,
            })

    return rows


# ── Sheet detector ────────────────────────────────────────────────────────────

def _detect_and_extract(ws) -> list[dict]:
    if str(ws["A2"].value or "").strip().upper() == "EXEMPT":
        return _extract_rows_exempt(ws)
    if "NON-EXEMPT" in str(ws["B2"].value or "").upper():
        return _extract_rows_nonexempt(ws)
    raise ValueError(f"Sheet '{ws.title}' is neither EXEMPT nor NON-EXEMPT")


# ── CSV builder ───────────────────────────────────────────────────────────────

def _build_csv(all_rows: list[dict]) -> bytes:
    buf    = io.StringIO()
    writer = csv.writer(buf)
    for r in all_rows:
        writer.writerow([
            r["employee"],    # 1
            r["post_date"],   # 2
            r["job"] + ".",   # 3
            r["phase"] + ".", # 4
            r["earn_code"],   # 5
            r["hours"],       # 6
            "",               # 7
            EMCO,             # 8
            "",               # 9
            "",               # 10
            "",               # 11
        ])
    return buf.getvalue().encode("utf-8")


# ── PDF converter ─────────────────────────────────────────────────────────────

def _xlsx_to_pdf(xlsx_bytes: bytes) -> bytes | None:
    """Convert xlsx → PDF via Excel COM. Returns None on any failure."""
    import os, tempfile
    try:
        import pythoncom
        import win32com.client
        pythoncom.CoInitialize()
    except ImportError:
        return None

    fd_x, tmp_x = tempfile.mkstemp(suffix=".xlsx")
    fd_p, tmp_p = tempfile.mkstemp(suffix=".pdf")
    os.close(fd_x)
    os.close(fd_p)
    try:
        os.unlink(tmp_p)
    except OSError:
        pass

    excel = None
    try:
        with open(tmp_x, "wb") as f:
            f.write(xlsx_bytes)

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible       = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(tmp_x)

        for ws in wb.Sheets:
            ps = ws.PageSetup
            ps.Zoom             = False
            ps.FitToPagesWide   = 1
            ps.FitToPagesTall   = 1

        wb.ExportAsFixedFormat(0, tmp_p)  # 0 = xlTypePDF
        wb.Close(False)

        with open(tmp_p, "rb") as f:
            return f.read()
    except Exception:
        return None
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        for p in (tmp_x, tmp_p):
            try:
                os.unlink(p)
            except OSError:
                pass


# ── Filename helper ───────────────────────────────────────────────────────────

def _renamed_stem(original_basename: str, employee_id) -> str:
    """Insert employee ID into the filename stem.

    AIC_Timesheet_BolleS_20260417.xlsx + 6021 → AIC_Timesheet_6021-BolleS_20260417
    """
    stem   = original_basename.rsplit(".", 1)[0]
    emp    = str(employee_id).strip()
    prefix = "AIC_Timesheet_"
    if stem.startswith(prefix):
        return prefix + emp + "-" + stem[len(prefix):]
    return emp + "-" + stem


# ── Main entry point ──────────────────────────────────────────────────────────

def process_reporting_zip(zip_bytes: bytes) -> tuple[bytes, dict, dict, dict]:
    """
    Process a zip of timesheet xlsx files.

    Returns:
        (csv_bytes, xlsx_files, per_file_csvs, per_file_pdfs)
    """
    import openpyxl

    all_rows      = {}
    orig_bytes    = {}
    per_file_rows = {}
    errors        = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            basename = name.split("/")[-1]
            if not basename or basename.startswith("~$"):
                continue
            ext = basename.rsplit(".", 1)[-1].lower() if "." in basename else ""
            if ext != "xlsx":
                continue

            file_bytes         = zf.read(name)
            orig_bytes[basename] = file_bytes

            try:
                wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws   = wb.active
                rows = _detect_and_extract(ws)
                all_rows[basename] = rows
                per_file_rows[basename] = rows
                wb.close()
            except Exception as e:
                errors.append(f"{basename}: {e}")

    if errors:
        raise ValueError("Errors processing file(s):\n" + "\n".join(errors))

    flat_rows     = [r for rows in all_rows.values() for r in rows]
    xlsx_files    = {}
    per_file_csvs = {}
    per_file_pdfs = {}

    for orig_name, rows in per_file_rows.items():
        emp_id = rows[0]["employee"] if rows else "unknown"
        stem   = _renamed_stem(orig_name, emp_id)
        fbytes = orig_bytes[orig_name]

        xlsx_files[stem + ".xlsx"]   = fbytes
        per_file_csvs[stem + ".csv"] = _build_csv(rows)

        pdf = _xlsx_to_pdf(fbytes)
        if pdf is not None:
            per_file_pdfs[stem + ".pdf"] = pdf

    return _build_csv(flat_rows), xlsx_files, per_file_csvs, per_file_pdfs
