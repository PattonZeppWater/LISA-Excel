"""
parser.py — Parse IDP Excel workbook into conduit_index, fill_index, and block_lib.

Supports .xlsx and .xlsm (keep_vba=True to preserve macros).
Column mapping is always by header name.

Sheet structures:
  ConduitIndex — row 1 = title (skip), row 2 = headers, row 3+ = data
  FillIndex    — row 1 = title (skip), row 2 = headers, row 3+ = data
  BlockLib_ACAD— row 1 = headers (no title row), row 2+ = data
  PickList     — row 1 = title (skip), row 2 = headers, row 3+ = data

Dynamic workbook support:
  Human-readable column names (e.g. "S Symbol", "Wire Gauge") are translated to
  internal names via workbook_mapper.  Cond_Ident and Fill_Ident are auto-generated
  when absent so the dynamic workbook needs no identifier columns.
"""

import io
import os
import re
import json
import base64
from copy import copy
from datetime import datetime, date

import openpyxl

from . import workbook_mapper


# Required sheets
_REQUIRED_SHEETS = {"ConduitIndex", "FillIndex"}

# Minimum columns that must be present after alias expansion.
# Cond_Ident and Fill_Ident are excluded because they are auto-generated.
_CONDUIT_REQUIRED_COLS = {"Cond_Tag"}
_FILL_REQUIRED_COLS    = {"Cond_Tag", "Src_TermBlockDesc", "Dst_TermBlockDesc"}


def _serialize(v):
    """Convert openpyxl cell value to a JSON-safe Python type."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _build_col_map(header_row: tuple) -> dict:
    """
    Return {header_name: zero_based_index} from a header row tuple.

    Duplicate header names are suffixed _2, _3, … so that all columns are
    addressable.  The first occurrence keeps the plain name; the second becomes
    "Name_2", the third "Name_3", etc.
    """
    result: dict = {}
    seen:   dict = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip()
        if not key:
            continue
        if key in seen:
            seen[key] += 1
            result[f"{key}_{seen[key]}"] = idx
        else:
            seen[key] = 1
            result[key] = idx
    return result


def _apply_aliases(col_map: dict, aliases: dict) -> dict:
    """
    Extend col_map with internal-name keys for any workbook column that has an alias.
    The original workbook key is preserved so write-back via _write_sheet still works.
    Alias keys are stripped before lookup to match how _build_col_map normalises headers.
    """
    result = dict(col_map)
    for wb_name, internal_name in aliases.items():
        normalised = wb_name.strip()
        if normalised in col_map and internal_name not in result:
            result[internal_name] = col_map[normalised]
    return result


def _row_to_dict(row: tuple, col_map: dict) -> dict:
    """Convert a data row tuple to a dict keyed by column header."""
    return {
        name: _serialize(row[idx] if idx < len(row) else None)
        for name, idx in col_map.items()
    }


def _parse_ref_docs(wb) -> dict:
    """
    Parse the 'Ref Documents' sheet into a dict keyed by ref-document name.

    Sheet layout: col A = name key, col B = DWG#, col C = Description, col D = Manufacturer.
    ConduitIndex col L ("Ref Documents") holds a comma-separated list of these name keys.
    """
    sheet_name = next(
        (s for s in wb.sheetnames
         if s.lower() in ("ref documents", "ref docs", "ref documents & deviations")),
        None,
    )
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        upper = {str(c).strip().upper() for c in row if c}
        if "REF DOCUMENT" in upper or "DWG#" in upper:
            header_idx = i
            break
    if header_idx is None:
        return {}

    headers = [str(c).strip().upper() if c else "" for c in rows[header_idx]]

    def _col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    ci_name = _col("REF DOCUMENT")
    ci_dwg  = _col("DWG#")
    ci_desc = _col("DESCRIPTION")
    ci_mfr  = _col("MANUFACTURER")

    result = {}
    for row in rows[header_idx + 1:]:
        if not any(c for c in row):
            continue
        name = row[ci_name] if ci_name is not None and ci_name < len(row) else None
        if not name:
            continue
        dwg  = row[ci_dwg]  if ci_dwg  is not None and ci_dwg  < len(row) else None
        desc = row[ci_desc] if ci_desc is not None and ci_desc < len(row) else None
        mfr  = row[ci_mfr]  if ci_mfr  is not None and ci_mfr  < len(row) else None
        result[str(name).strip()] = {
            "dwg_num":      str(dwg).strip()  if dwg  else None,
            "description":  str(desc).strip() if desc else None,
            "manufacturer": str(mfr).strip()  if mfr  else None,
        }
    return result


# Module cache of block heights (from BlockIndex) so /generate — which works from
# parsed JSON, not the workbook — can drive grid spacing without a frontend change.
_BLOCK_HEIGHTS: dict = {}


_MEASURED_PATH = os.path.join(os.path.dirname(__file__), "measured_geometry.json")
_MEASURED_CACHE = None


def _load_measured() -> dict:
    """Exact per-block vertical geometry measured once from AutoCAD (below/above the wire
    terminal, in drawing units), keyed by block name. Ground truth that overrides the
    BlockIndex's estimated/padded Dimension_Length + Insertion_ShiftY, which were
    unreliable (some Insertion_ShiftY values exceeded the block's own height)."""
    global _MEASURED_CACHE
    if _MEASURED_CACHE is None:
        try:
            with open(_MEASURED_PATH, "r") as f:
                _MEASURED_CACHE = json.load(f)
        except Exception:
            _MEASURED_CACHE = {}
    return _MEASURED_CACHE


def get_block_heights() -> dict:
    """Block heights captured by the last parse_workbook() call, with measured geometry
    merged in.  Keys: "NAME|VISIBILITY" and bare "NAME" (max over states) -> height;
    "__SHIFTY__|NAME[|VIS]" -> signed below-wire extent.

    Measured geometry (measured_geometry.json) overrides the workbook for FIXED
    single-state blocks only (multi-state switches/instruments keep their per-state
    workbook values). The measured "_L" geometry is mirror-identical to "_R", so it's
    applied to both sides."""
    out = dict(_BLOCK_HEIGHTS)
    for nm, d in _load_measured().items():
        try:
            if int(d.get("vstates", 1)) != 1:      # leave multi-state blocks to the workbook
                continue
            h = float(d["height"]); below = float(d["below"])
        except Exception:
            continue
        up = str(nm).strip().upper()
        variants = [up]
        if up.endswith("_L"):
            variants.append(up[:-2] + "_R")         # mirror: same vertical extents
        for v in variants:
            out[f"{v}|NA"] = h
            out[v] = max(out.get(v, 0.0), h)
            out[f"__SHIFTY__|{v}|NA"] = -below
            out[f"__SHIFTY__|{v}"] = -below
    return out


def _parse_block_index(wb) -> dict:
    """Read BlockIndex into {name|vis: height, name: height}.  Height is the block's
    vertical size, column 'Dimension_Length'.  Used for grid spacing."""
    if "BlockIndex" not in wb.sheetnames:
        return {}
    rows = list(wb["BlockIndex"].iter_rows(values_only=True))
    if not rows:
        return {}
    # find the header row (has 'Block_Name')
    hdr_idx = None
    for i in range(min(4, len(rows))):
        cells = [str(c).strip().lower() if c is not None else "" for c in rows[i]]
        if "block_name" in cells:
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}
    header = [str(c).strip().lower() if c is not None else "" for c in rows[hdr_idx]]
    try:
        ni = header.index("block_name")
    except ValueError:
        return {}
    hi = header.index("dimension_length") if "dimension_length" in header else None
    vi = header.index("visibility state") if "visibility state" in header else None
    si = header.index("insertion_shifty") if "insertion_shifty" in header else None
    if hi is None:
        return {}
    out: dict = {}
    for row in rows[hdr_idx + 1:]:
        if ni >= len(row):
            continue
        name = row[ni]
        h = row[hi] if hi < len(row) else None
        if not name or not isinstance(h, (int, float)):
            continue
        nm = str(name).strip().upper()
        vs = (str(row[vi]).strip().upper() if vi is not None and vi < len(row)
              and row[vi] is not None else "NA")
        out[f"{nm}|{vs}"] = float(h)
        out[nm] = max(out.get(nm, 0.0), float(h))   # name-only fallback = tallest state
        # Insertion_ShiftY = how far the block sits BELOW its wire terminal; the block
        # reaches height-|ShiftY| ABOVE the wire (e.g. an HOA switch's box). Stored under a
        # prefixed key in the same map so grid spacing can reserve that top overhang.
        if si is not None and si < len(row) and isinstance(row[si], (int, float)):
            out[f"__SHIFTY__|{nm}|{vs}"] = float(row[si])
            out.setdefault(f"__SHIFTY__|{nm}", float(row[si]))
    return out


def _parse_project_desc(wb) -> dict:
    """Read the optional 'Project Description' sheet into ordered title-block lines.

    Layout is a horizontal form: a row of labels (Owner / Job Title / Content /
    Proj No. / Status / Date / Engineer / Drafter) with the values in the row
    directly beneath. These columns map POSITIONALLY to ACADE's project-description
    lines LINE1..LINE24 (i.e. the .wdp *[1]..*[N] fields): the k-th labeled column
    is LINE k. That mirrors AutoCAD Electrical's "Update Title Block", which fills
    every drawing's title block from these lines.

    Returns {"lines": [v1, v2, ...], "labels": [l1, l2, ...]} in column order, or
    {} when the sheet is absent / has no recognizable labels (so the workbook just
    falls back to the .wdp template defaults).
    """
    name = next((s for s in wb.sheetnames if s.strip().lower() == "project description"), None)
    if not name:
        return {}
    ws = wb[name]
    known = {"owner", "job title", "content", "proj no.", "proj no", "project no.",
             "project number", "status", "date", "engineer", "drafter"}
    ncols = min(int(ws.max_column or 1), 30)
    label_row = None
    for r in range(1, min(int(ws.max_row or 1), 10) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ncols + 1)]
        if any(v in known for v in vals):
            label_row = r
            break
    if label_row is None:
        return {}
    labels, lines = [], []
    for c in range(1, ncols + 1):
        label = str(ws.cell(label_row, c).value or "").strip()
        if not label:
            continue
        val = ws.cell(label_row + 1, c).value
        labels.append(label)
        lines.append("" if val is None else str(val).strip())
    if not labels:
        return {}
    return {"lines": lines, "labels": labels}


def parse_workbook(file_bytes: bytes, filename: str = "workbook.xlsx") -> dict:
    """
    Parse an IDP Excel workbook.

    Returns:
        {
            "filename":      str,
            "original_b64":  str,
            "conduit_index": [ {col_name: value, ...}, ... ],
            "fill_index":    [ {col_name: value, ...}, ... ],
            "block_lib":     [ {block_name, visibility_state, category, ..., attributes: [...]}, ... ],
            "ref_docs":      [ {dwg_num, description, manufacturer, conduits: [...]}, ... ],
        }

    Raises:
        ValueError on missing sheets or missing required columns.
    """
    keep_vba = filename.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=keep_vba)

    missing = _REQUIRED_SHEETS - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing required sheet(s): {', '.join(sorted(missing))}")

    conduit_index = _parse_sheet(
        wb["ConduitIndex"], _CONDUIT_REQUIRED_COLS,
        workbook_mapper.CONDUIT_COL_ALIASES,
    )
    # Drop blank conduit rows: formatted-but-empty ConduitIndex rows can carry
    # stray cell content (dropdown defaults, etc.) but no Conduit Name, and would
    # otherwise be offered for generation as hundreds of blank conduits.
    conduit_index = [
        r for r in conduit_index if str(r.get("Cond_Tag") or "").strip()
    ]
    fill_index = _parse_sheet(
        wb["FillIndex"], _FILL_REQUIRED_COLS,
        workbook_mapper.FILL_COL_ALIASES,
    )
    block_lib = _parse_block_lib(wb["BlockLib_ACAD"]) if "BlockLib_ACAD" in wb.sheetnames else []
    ref_docs  = _parse_ref_docs(wb)
    deviation_notes = _parse_deviation_notes(wb)
    project_desc = _parse_project_desc(wb)

    global _BLOCK_HEIGHTS
    _BLOCK_HEIGHTS = _parse_block_index(wb)

    wb.close()

    parsed = {
        "filename":      filename,
        "original_b64":  base64.b64encode(file_bytes).decode("utf-8"),
        "conduit_index": conduit_index,
        "fill_index":    fill_index,
        "block_lib":     block_lib,
        "ref_docs":      ref_docs,
        "deviation_notes": deviation_notes,
        "block_heights": _BLOCK_HEIGHTS,
        "project_desc": project_desc,
    }
    return workbook_mapper.apply_workbook_mapping(parsed)


def _find_header_row(all_rows: list, aliases: dict) -> int:
    """
    Return the 0-based index of the header row by scanning rows 0-4.

    A row is considered the header row when, after alias expansion, at least
    one of its cells matches a known alias key.  Falls back to row index 1
    (the original assumption) if nothing matches.
    """
    known = {k.strip() for k in (aliases or {})}
    for idx in range(min(5, len(all_rows))):
        cells = {str(c).strip() for c in all_rows[idx] if c is not None}
        if cells & known:
            return idx
    return 1  # original default


def _parse_sheet(ws, required_cols: set, aliases: dict = None) -> list:
    """
    Parse a sheet, auto-detecting which row holds the column headers.

    Scans rows 1-5 for a row whose cells overlap known alias keys; falls back
    to row 2 (index 1) for sheets with no aliases.  Data rows start immediately
    after the header row.

    aliases — optional dict of {workbook_col_name: internal_name}.  Matching
    entries are added to col_map so each row dict carries both the original
    workbook key and the internal-name key pointing to the same value.
    Required column validation runs against the expanded col_map, which means
    either the workbook name or the internal name satisfies the check.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 1:
        raise ValueError(f"Sheet '{ws.title}' is empty.")

    hdr_idx = _find_header_row(all_rows, aliases)
    col_map = _build_col_map(all_rows[hdr_idx])

    if aliases:
        col_map = _apply_aliases(col_map, aliases)

    missing = required_cols - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Sheet '{ws.title}' is missing required column(s): {', '.join(sorted(missing))}"
        )

    rows = []
    for raw in all_rows[hdr_idx + 1:]:  # data rows start after header
        if all(c is None for c in raw):
            continue
        # Stop at the END sentinel row
        first = raw[0]
        if first is not None and str(first).strip().upper() == "END":
            break
        rows.append(_row_to_dict(raw, col_map))

    return rows



def _parse_deviation_notes(wb) -> dict:
    """
    Parse the Ref Documents sheet's deviation-notes lookup:
    col 'Deviations Notes #' (the number) -> col 'Deviations Notes' (the note text).
    Independent of the ref-doc rows (which are keyed by name in cols A-D).
    Returns { "1": "Terminal Unknown", "2": "Bitch", ... } keyed by number-as-string.
    """
    sheet_name = next(
        (s for s in wb.sheetnames
         if s.lower() in ("ref documents", "ref docs", "ref documents & deviations")),
        None,
    )
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        up = {str(c).strip().upper() for c in row if c}
        if any(h in up for h in ("DEVIATIONS NOTES #", "DEVIATION NOTES #")) or "REF DOCUMENT" in up:
            header_idx = i
            break
    if header_idx is None:
        return {}
    headers = [str(c).strip().upper() if c else "" for c in rows[header_idx]]

    def _col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    ci_num = _col("DEVIATIONS NOTES #", "DEVIATION NOTES #", "DEVIATIONS #", "NOTE #")
    ci_txt = _col("DEVIATIONS NOTES", "DEVIATION NOTES")
    if ci_num is None or ci_txt is None:
        return {}

    result = {}
    for row in rows[header_idx + 1:]:
        num = row[ci_num] if ci_num < len(row) else None
        txt = row[ci_txt] if ci_txt < len(row) else None
        if num is None or str(num).strip() == "":
            continue
        key = str(num).strip()
        try:
            key = str(int(float(key)))   # normalise "1.0" -> "1"
        except ValueError:
            pass
        result[key] = str(txt).strip() if txt is not None else ""
    return result

def _parse_block_lib(ws) -> list:
    """
    Parse BlockLib_ACAD — row 1 is headers (no title row).
    Returns a list of block definition dicts, each with a compact `attributes` list.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    col_map = _build_col_map(all_rows[0])  # row 1 = headers
    blocks  = []

    for raw in all_rows[1:]:
        if all(c is None for c in raw):
            continue
        row = _row_to_dict(raw, col_map)

        block_name = row.get("Block_Name")
        if not block_name:
            continue

        att_count = int(row.get("Att_Count") or 0)
        attributes = []
        for i in range(1, att_count + 1):
            val = row.get(f"Att_{i:03d}_Name")
            if val:
                attributes.append(str(val).strip())

        blocks.append({
            "block_name":       block_name,
            "visibility_state": row.get("VisibilityState"),
            "category":         row.get("Block_Category"),
            "manufacturer":     row.get("Block_Manufacturer"),
            "part_number":      row.get("Block_PartNumber"),
            "complete":         row.get("COMPLETE"),
            "dim_width":        row.get("Dimension_Width"),
            "dim_length":       row.get("Dimension_Length"),
            "insert_shift_x":   row.get("Insertion_ShiftX"),
            "insert_shift_y":   row.get("Insertion_ShiftY"),
            "att_count":        att_count,
            "attributes":       attributes,
        })

    return blocks


# ── Write-back ────────────────────────────────────────────────────────────────

def write_workbook(file_bytes: bytes, filename: str, conduit_index: list, fill_index: list) -> bytes:
    """
    Write conduit_index and fill_index back into the workbook and return the
    modified bytes.  All other sheets (BlockLib_ACAD, PickList, etc.) are
    preserved unchanged.
    """
    keep_vba = filename.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_vba=keep_vba)

    if "ConduitIndex" in wb.sheetnames:
        _write_sheet(wb["ConduitIndex"], conduit_index)
    if "FillIndex" in wb.sheetnames:
        _write_sheet(wb["FillIndex"], fill_index)

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _adjust_row_refs(formula: str, from_row: int, to_row: int) -> str:
    """Rewrite cell row-number references in a formula (e.g. B3 → B5)."""
    if from_row == to_row:
        return formula
    return re.sub(
        rf'(\$?[A-Z]+){from_row}\b',
        lambda m: m.group(1) + str(to_row),
        formula,
    )


def _write_sheet(ws, rows: list):
    """
    Write data rows into a sheet, preserving the END sentinel, cell fills,
    row formulas, and data validations from the template.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 1:
        return

    # Use the same header-row detection as _parse_sheet so write-back is consistent
    from . import workbook_mapper as _wm
    _aliases = {**_wm.CONDUIT_COL_ALIASES, **_wm.FILL_COL_ALIASES}
    hdr_idx = _find_header_row(all_rows, _aliases)

    col_map = _build_col_map(all_rows[hdr_idx])
    if not col_map:
        return

    INSERT_AT = hdr_idx + 2  # 1-based first data row (header is hdr_idx+1 in 1-based)

    template_fills    = {}
    template_formulas = {}
    for cell in ws[INSERT_AT]:
        if cell.fill and cell.fill.fill_type == "solid":
            template_fills[cell.column] = copy(cell.fill)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            template_formulas[cell.column] = cell.value

    end_row = None
    for i, raw in enumerate(all_rows[INSERT_AT - 1:], start=INSERT_AT):
        first_val = raw[0] if raw else None
        if first_val is not None and str(first_val).strip().upper() == "END":
            end_row = i
            break

    existing_slots = (end_row - INSERT_AT) if end_row is not None else max(0, ws.max_row - 2)
    n = len(rows)

    if n > existing_slots:
        ws.insert_rows(INSERT_AT + existing_slots, n - existing_slots)
    elif n < existing_slots:
        ws.delete_rows(INSERT_AT + n, existing_slots - n)

    formula_col_set = set(template_formulas)
    for r_idx, row_dict in enumerate(rows):
        excel_row = INSERT_AT + r_idx
        for name, col_idx in col_map.items():
            col_1 = col_idx + 1
            if col_1 not in formula_col_set:
                ws.cell(row=excel_row, column=col_1, value=row_dict.get(name))
        for col_1, formula in template_formulas.items():
            ws.cell(row=excel_row, column=col_1,
                    value=_adjust_row_refs(formula, INSERT_AT, excel_row))
        for col_1, fill in template_fills.items():
            ws.cell(row=excel_row, column=col_1).fill = copy(fill)

    if n > 0:
        last_data_row = INSERT_AT + n - 1
        for dv in ws.data_validations.dataValidation:
            new_sqref = re.sub(
                r'([A-Z]+)\d+:([A-Z]+)\d+',
                lambda m: f"{m.group(1)}{INSERT_AT}:{m.group(2)}{last_data_row}",
                str(dv.sqref),
            )
            dv.sqref = new_sqref


# ── Row helpers ────────────────────────────────────────────────────────────────

def get_conduit_row(conduit_index: list, conduit_ident: int) -> dict | None:
    """Return the ConduitIndex row dict matching conduit_ident, or None."""
    for row in conduit_index:
        val = row.get("Cond_Ident")
        if val is not None and int(val) == conduit_ident:
            return row
    return None


def get_fill_rows(fill_index: list, cond_tag: str) -> list:
    """Return all FillIndex rows matching cond_tag (preserves row order)."""
    return [
        row for row in fill_index
        if row.get("Cond_Tag") is not None and str(row["Cond_Tag"]) == str(cond_tag)
    ]


def build_conduit_data(conduit_row: dict) -> dict:
    """
    Extract conduit block attributes from a ConduitIndex row dict.
    Fill slots: include Fill01..Fill30 only while Fill##_Type is non-empty.
    """
    data = {
        "Cdt_Name":  conduit_row.get("Cond_Tag"),
        "Cdt_Type":  conduit_row.get("Cond_Type"),
        "Cdt_Size":  conduit_row.get("Cond_Size"),
        "Src_Jbox":  conduit_row.get("Src_Jbox"),
        "Dst_Jbox":  conduit_row.get("Dst_Jbox"),
        "Src_Name01": conduit_row.get("Src_Name01"),
        "Src_Name02": conduit_row.get("Src_Name02"),
        "Src_Name03": conduit_row.get("Src_Name03"),
        "Dst_Name01": conduit_row.get("Dst_Name01"),
        "Dst_Name02": conduit_row.get("Dst_Name02"),
        "Dst_Name03": conduit_row.get("Dst_Name03"),
    }

    for i in range(1, 31):
        slot = f"{i:02d}"
        fill_type = conduit_row.get(f"Fill{slot}_Type")
        if not fill_type:
            break
        data[f"Fill{slot}_Type"]     = fill_type
        data[f"Fill{slot}_Color"]    = conduit_row.get(f"Fill{slot}_Color")
        data[f"Fill{slot}_Size"]     = conduit_row.get(f"Fill{slot}_Size")
        data[f"Fill{slot}_Quantity"] = conduit_row.get(f"Fill{slot}_Quantity")

    return data


def build_wire_labels(fill_index: list) -> bytes:
    """
    Build a wire-labels Excel file from fill_index.

    One row per conductor (up to Wire_Count per fill row).
    Column A = Wire{n}_SrcLabel, Column B = Wire{n}_DstLabel, Column C = Wire{n}_Size.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.column_dimensions["A"].width = 35.7
    ws.column_dimensions["B"].width = 35.9
    ws.column_dimensions["C"].width = 12.0

    wire_fields = [
        ("Wire1_SrcLabel", "Wire1_DstLabel", "Wire1_Size"),
        ("Wire2_SrcLabel", "Wire2_DstLabel", "Wire2_Size"),
        ("Wire3_SrcLabel", "Wire3_DstLabel", "Wire3_Size"),
        ("Wire4_SrcLabel", "Wire4_DstLabel", "Wire4_Size"),
    ]

    for fill_row in fill_index:
        wire_count = int(fill_row.get("Wire_Count") or 0)
        for i, (src_field, dst_field, size_field) in enumerate(wire_fields):
            if i >= wire_count:
                break
            src  = (fill_row.get(src_field)  or "").strip() or None
            dst  = (fill_row.get(dst_field)  or "").strip() or None
            size = (fill_row.get(size_field) or "").strip() or None
            if src or dst or size:
                ws.append([src, dst, size])

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _split_block_name(name):
    """Split a symbol display name into (base block name, visibility state).

    The dynamic workbook stores e.g. "Inst_4W_R (Field_4Term)" where the real
    AutoCAD block is "Inst_4W_R" and "Field_4Term" is a dynamic-block visibility
    state.  Names without parentheses pass through unchanged.
    """
    if name is None:
        return (None, None)
    s = str(name).strip()
    if "(" in s and ")" in s:
        base = s.split("(", 1)[0].strip()
        vis = s[s.index("(") + 1:].split(")", 1)[0].strip()
        return (base or None, vis or None)
    return (s or None, None)


def build_loop_list(fill_rows: list) -> list:
    """Convert FillIndex rows into loop_data dicts for the DWG generator."""
    loops = []
    for row in fill_rows:
        _sb, _sv = _split_block_name(row.get("Src_TermBlockDesc"))
        _db, _dv = _split_block_name(row.get("Dst_TermBlockDesc"))
        # The instrument bubble's ISA tag comes from whichever side holds the
        # instrument. Each workbook column maps 1:1 to its like-named block attr:
        #   FunctionID -> FunctIdent, ElementID -> ElementIdent,
        #   Element#   -> ElementNum, Loop#     -> LoopNum.
        # (FunctionID arrives under the legacy alias name Loop_Src/DstDesc.)
        _si = "inst" in (_sb or "").lower()
        _di = "inst" in (_db or "").lower()
        if _si:
            _isa_elem, _isa_loop, _isa_num, _isa_func = (row.get("Src_ISAElem"),
                row.get("Src_ISALoop"), row.get("Src_ISAElemNum"),
                row.get("Loop_SrcDesc"))
        elif _di:
            _isa_elem, _isa_loop, _isa_num, _isa_func = (row.get("Dst_ISAElem"),
                row.get("Dst_ISALoop"), row.get("Dst_ISAElemNum"),
                row.get("Loop_DstDesc"))
        else:
            _isa_elem = _isa_loop = _isa_num = _isa_func = None
        loop = {
            "src_block": _sb,
            "dst_block": _db,
            "src_block_visibility": row.get("Src_TermBlockVisibilityState") or _sv,
            "dst_block_visibility": row.get("Dst_TermBlockVisibilityState") or _dv,
            # FunctionID (e.g. AE) fills the function bubble; ElementID (e.g. AIT)
            # fills the element bubble; Element# and Loop# fill their numbers.
            "ISATag_FunctIdent":   _isa_func,
            "ISATag_ElementIdent": _isa_elem,
            "ISATag_ElementNum":   _isa_num,
            "ISATag_LoopNum":      _isa_loop,
            "_cond_tag":           row.get("Cond_Tag"),
            "Wire_Count":    row.get("Wire_Count"),
            "Wire_Type":     row.get("Wire_Type"),
            "Wire1_Color":      row.get("Wire1_Color"),
            "Wire1_Size":       row.get("Wire1_Size"),
            "Wire1_SrcTermBlk": row.get("Wire1_SrcTermBlk"),
            "Wire1_SrcTermNum": row.get("Wire1_SrcTermNum"),
            "Wire1_DstTermBlk": row.get("Wire1_DstTermBlk"),
            "Wire1_DstTermNum": row.get("Wire1_DstTermNum"),
            "Wire1_SrcLabel":   row.get("Wire1_SrcLabel"),
            "Wire1_DstLabel":   row.get("Wire1_DstLabel"),
            "Wire2_Color":      row.get("Wire2_Color"),
            "Wire2_Size":       row.get("Wire2_Size"),
            "Wire2_SrcTermBlk": row.get("Wire2_SrcTermBlk"),
            "Wire2_SrcTermNum": row.get("Wire2_SrcTermNum"),
            "Wire2_DstTermBlk": row.get("Wire2_DstTermBlk"),
            "Wire2_DstTermNum": row.get("Wire2_DstTermNum"),
            "Wire2_SrcLabel":   row.get("Wire2_SrcLabel"),
            "Wire2_DstLabel":   row.get("Wire2_DstLabel"),
            "Wire3_Color":      row.get("Wire3_Color"),
            "Wire3_Size":       row.get("Wire3_Size"),
            "Wire3_SrcTermBlk": row.get("Wire3_SrcTermBlk"),
            "Wire3_SrcTermNum": row.get("Wire3_SrcTermNum"),
            "Wire3_DstTermBlk": row.get("Wire3_DstTermBlk"),
            "Wire3_DstTermNum": row.get("Wire3_DstTermNum"),
            "Wire3_SrcLabel":   row.get("Wire3_SrcLabel"),
            "Wire3_DstLabel":   row.get("Wire3_DstLabel"),
            "Wire4_Color":      row.get("Wire4_Color"),
            "Wire4_Size":       row.get("Wire4_Size"),
            "Wire4_SrcTermBlk": row.get("Wire4_SrcTermBlk"),
            "Wire4_SrcTermNum": row.get("Wire4_SrcTermNum"),
            "Wire4_DstTermBlk": row.get("Wire4_DstTermBlk"),
            "Wire4_DstTermNum": row.get("Wire4_DstTermNum"),
            "Wire4_SrcLabel":   row.get("Wire4_SrcLabel"),
            "Wire4_DstLabel":   row.get("Wire4_DstLabel"),
            "Loop_SrcDesc":  row.get("Loop_SrcDesc"),
            "Loop_DstDesc":  row.get("Loop_DstDesc"),
            "Loop_Category": row.get("Loop_Catagory"),
            "Src_Desc1":     row.get("Src_Desc1"),
            "Src_Desc2":     row.get("Src_Desc2"),
            "Src_Desc3":     row.get("Src_Desc3"),
            "Dst_Desc1":     row.get("Dst_Desc1"),
            "Dst_Desc2":     row.get("Dst_Desc2"),
            "Dst_Desc3":     row.get("Dst_Desc3"),
            # Device ratings (S Rating / D Rating) -> Rating attr on device blocks
            "Src_Rating":    row.get("Src_Rating"),
            "Dst_Rating":    row.get("Dst_Rating"),
            # Per-term "Hide from Generation" tokens (CSV like "S1,D3") from the
            # hidden "Hidden Terms" column; consumed just below.
            "Hidden_Terms":  row.get("Hidden_Terms"),
            # Per-conductor tag names (S Tag / D Tag columns) → block Tag1-4 attrs
            "Wire1_SrcTag":  row.get("Wire1_SrcTag"),
            "Wire2_SrcTag":  row.get("Wire2_SrcTag"),
            "Wire3_SrcTag":  row.get("Wire3_SrcTag"),
            "Wire4_SrcTag":  row.get("Wire4_SrcTag"),
            "Wire1_DstTag":  row.get("Wire1_DstTag"),
            "Wire2_DstTag":  row.get("Wire2_DstTag"),
            "Wire3_DstTag":  row.get("Wire3_DstTag"),
            "Wire4_DstTag":  row.get("Wire4_DstTag"),
            # Spare-block attributes -> Type/Quantity on Spare_L / Spare_R
            "Src_SpareType": row.get("Src_SpareType"),
            "Dst_SpareType": row.get("Dst_SpareType"),
            "Src_SpareQty":  row.get("Src_SpareQty"),
            "Dst_SpareQty":  row.get("Dst_SpareQty"),
            # Wire-label display toggle ("Auto" / "None")
            "Wire1_LabelMode": row.get("Wire1_LabelMode"),
            "Wire2_LabelMode": row.get("Wire2_LabelMode"),
            "Wire3_LabelMode": row.get("Wire3_LabelMode"),
            "Wire4_LabelMode": row.get("Wire4_LabelMode"),
            "is_continuation": False,
        }
        # "Hide from Generation" tokens (CSV) from the hidden helper column. Terms
        # (S#/D#) are stamped with a sentinel so the generator blanks just that
        # terminal number (slot preserved). Tags (SG#/DG#) are blanked outright --
        # _collapse_tags then clears that slot's block-default text. The workbook
        # cells keep their real values either way.
        _ht = str(row.get("Hidden_Terms") or "")
        if _ht.strip():
            _toks = {t.strip().upper() for t in _ht.split(",") if t.strip()}
            for _w in range(1, 5):
                if f"S{_w}" in _toks:
                    loop[f"Wire{_w}_SrcTermNum"] = "##HIDETERM##"
                if f"D{_w}" in _toks:
                    loop[f"Wire{_w}_DstTermNum"] = "##HIDETERM##"
                if f"SG{_w}" in _toks:
                    loop[f"Wire{_w}_SrcTag"] = ""
                if f"DG{_w}" in _toks:
                    loop[f"Wire{_w}_DstTag"] = ""
        loops.append(loop)

    # Mark instrument continuation rows: a row whose instrument side (the side
    # holding an "Inst…" block) repeats the previous row's instrument on the same
    # conduit. The DWG generator inserts the instrument once per group and stacks
    # all the group's wires against that single block.
    for k in range(1, len(loops)):
        cur, prev = loops[k], loops[k - 1]
        cd, pd = (cur.get("dst_block") or ""), (prev.get("dst_block") or "")
        cs, ps = (cur.get("src_block") or ""), (prev.get("src_block") or "")
        same_conduit = cur.get("_cond_tag") == prev.get("_cond_tag")
        dst_inst_same = bool(cd) and ("inst" in cd.lower()) and cd == pd
        src_inst_same = bool(cs) and ("inst" in cs.lower()) and cs == ps
        # A row carrying its OWN colors is the start of a new loop (the anchor),
        # never a continuation -- continuation rows share the anchor's Color 1-4 and
        # have none of their own.  Without this, two identical back-to-back loops
        # (e.g. a pasted duplicate of the same instrument) merge into one group, so
        # the 2nd instrument never inserts and its wires lose their colours.
        has_own_colors = any(
            cur.get(f"Wire{n}_Color") not in (None, "") for n in range(1, 5)
        )
        if same_conduit and (dst_inst_same or src_inst_same) and not has_own_colors:
            cur["is_continuation"] = True

    return loops
