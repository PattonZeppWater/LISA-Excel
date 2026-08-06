"""
generate.py â€” IDP Generation Flask routes.

Endpoints:
  GET  /api/health                 Flask health check (also on app.py)
  POST /api/parse                  Upload + parse Excel workbook
  GET  /api/parse-template         Parse the latest template from _Templates/
  GET  /api/download-template      Serve the latest IDP workbook template for download
  GET  /api/autocad-status         Check if AutoCAD is running on this machine
  GET  /api/browse-folder          Open a tkinter folder picker; return chosen path
  POST /api/generate               Generate one DWG for a given conduit_ident
  POST /api/finalize-project       Assemble the project ONCE (GENERAL sheets + drawing index + .wdp/.aepx)
  POST /api/wire-labels            Generate wire-labels Excel from fill_index
  POST /api/download               Re-export current workbook state to Excel
"""

import io
import os
import base64
import zipfile
import tkinter as tk
from tkinter import filedialog

from flask import Blueprint, request, jsonify, send_file

from ..services import parser as svc_parser
from ..services import autocad_bridge
from ..services import workbook_mapper as svc_mapper
from ..services import wdp_writer
from ..services import note_refs as svc_note_refs

idp_gen_bp = Blueprint("idp_gen", __name__)


# â”€â”€ Template directory â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_TEMPLATE_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "_Templates")
)


# â”€â”€ Parse â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/parse", methods=["POST"])
def parse():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return jsonify({"error": "Only .xlsx and .xlsm files are supported"}), 400

    try:
        data = svc_parser.parse_workbook(f.read(), f.filename)
        return jsonify(data)
    except zipfile.BadZipFile:
        return jsonify({"error": "Cannot read the file â€” it may be open in Excel. Close it and try again."}), 400
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# â”€â”€ Parse template â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/parse-template", methods=["GET"])
def parse_template():
    """Load and parse the most recently modified workbook template from _Templates/."""
    try:
        candidates = [
            f for f in os.listdir(_TEMPLATE_DIR)
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
        ]
    except FileNotFoundError:
        return jsonify({"error": f"Template directory not found: {_TEMPLATE_DIR}"}), 404

    if not candidates:
        return jsonify({"error": "No workbook template files found in _Templates/"}), 404

    latest = max(
        candidates,
        key=lambda f: os.path.getmtime(os.path.join(_TEMPLATE_DIR, f))
    )
    template_path = os.path.join(_TEMPLATE_DIR, latest)

    try:
        with open(template_path, "rb") as fh:
            file_bytes = fh.read()
        data = svc_parser.parse_workbook(file_bytes, latest)
        return jsonify(data)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# â”€â”€ Download workbook template â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/download-template", methods=["GET"])
def download_template():
    """Serve the latest IDP workbook template from _Templates/ for download."""
    try:
        candidates = [
            f for f in os.listdir(_TEMPLATE_DIR)
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
        ]
    except FileNotFoundError:
        return jsonify({"error": f"Template directory not found: {_TEMPLATE_DIR}"}), 404

    if not candidates:
        return jsonify({"error": "No workbook template files found in _Templates/"}), 404

    latest = max(
        candidates,
        key=lambda f: os.path.getmtime(os.path.join(_TEMPLATE_DIR, f))
    )
    template_path = os.path.join(_TEMPLATE_DIR, latest)
    mimetype = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if latest.lower().endswith(".xlsm")
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return send_file(
        template_path,
        mimetype=mimetype,
        as_attachment=True,
        download_name=latest,
    )


# â”€â”€ AutoCAD status â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/autocad-status", methods=["GET"])
def autocad_status():
    """Check whether a running AutoCAD instance is accessible on this machine. Runs the COM
    check on autocad_bridge's dedicated COM thread -- a raw GetActiveObject on a Flask worker
    thread (threaded server) would fail 'CoInitialize has not been called' and wrongly report
    AutoCAD as not running."""
    try:
        return jsonify(autocad_bridge.autocad_status())
    except Exception:
        return jsonify({"running": False, "version": None})


# â”€â”€ Browse folder â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/browse-folder", methods=["GET"])
def browse_folder():
    """Open a tkinter folder picker on the server machine and return the chosen path."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", True)
        folder = filedialog.askdirectory(title="Select Output Folder")
        root.destroy()
        return jsonify({"path": folder or None})
    except Exception as e:
        return jsonify({"path": None, "error": str(e)})


# ── Export conduit-list CSV ──────────────────────────────────────────────────

@idp_gen_bp.route("/export-conduit-list", methods=["POST"])
def export_conduit_list():
    """Save a conduit-list CSV to a location the user picks (native Save-As dialog).

    A browser-style download does not work inside the LISA desktop webview, so the
    frontend hands us the finished CSV text and we write it to disk here.

    Body: { "csv": "<text>", "filename": "name.csv", "default_dir": "C:\\..." }
    Returns { ok: true, path } | { ok: false, cancelled: true } | { ok: false, error }.
    """
    body = request.get_json(silent=True) or {}
    csv_text = body.get("csv", "")
    if not csv_text:
        return jsonify({"ok": False, "error": "No CSV content to save."}), 400
    filename = body.get("filename") or "conduit_list.csv"
    default_dir = body.get("default_dir") or ""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", True)
        path = filedialog.asksaveasfilename(
            title="Save Conduit List CSV",
            initialfile=filename,
            initialdir=default_dir or None,
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        root.destroy()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    if not path:
        return jsonify({"ok": False, "cancelled": True})
    try:
        # utf-8-sig writes the BOM so Excel opens the CSV with the right encoding;
        # newline="" keeps the CSV's own \r\n line endings intact.
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            fh.write(csv_text)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not write the file: {e}"}), 500
    return jsonify({"ok": True, "path": path})


# â”€â”€ Generate one DWG â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/generate", methods=["POST"])
def generate():
    """
    Generate a DWG for one conduit.

    Request body (JSON):
    {
        "conduit_ident":  int,
        "conduit_index":  [ {...row dicts...} ],
        "fill_index":     [ {...row dicts...} ],
        "output_folder":  "C:\\path\\to\\output"
    }

    Response (JSON):
    { "success": bool, "output_path": str, "warnings": [...], "error": str|null }
    """
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "JSON body required"}), 400

    conduit_ident = body.get("conduit_ident")
    conduit_index = body.get("conduit_index", [])
    fill_index    = body.get("fill_index", [])
    output_folder = body.get("output_folder", "")
    ref_docs       = body.get("ref_docs", {})
    deviation_notes = body.get("deviation_notes", {})   # { "1": "note text", ... }
    project_desc   = body.get("project_desc", {}) or {}  # Project Description sheet -> .wdp fields
    project_number = str(body.get("project_number") or "").strip()
    seq_num        = body.get("seq_num")
    sheet_max      = body.get("sheet_max")   # total sheets in the project -> title block "N OF <max>"
    # File suffix appended after the sequence number (default "e", e.g. 56.1077-01e)
    file_suffix    = body.get("file_suffix")
    file_suffix    = "e" if file_suffix is None else str(file_suffix).strip()
    # "Make it a project": assemble a full AIC project (GENERAL sheets + a sectioned .wdp/.aepx)
    # instead of just listing the conduit drawings. Requires a project number (names the files).
    make_project   = bool(body.get("make_project"))

    if conduit_ident is None:
        return jsonify({"error": "No conduit was selected to generate. Pick a conduit and try again."}), 400
    if not output_folder:
        return jsonify({"error": "No output folder is set. Choose where the drawings should be saved, then try again."}), 400
    if make_project and not project_number:
        return jsonify({"error": "A project number is required to generate a project. "
                                 "Enter one in the Project number field or uncheck 'Make it a project'."}), 400

    conduit_row = svc_parser.get_conduit_row(conduit_index, int(conduit_ident))
    if conduit_row is None:
        return jsonify({"error": f"The selected conduit (#{conduit_ident}) wasn't found in the "
                                 f"loaded workbook. Re-load the workbook and try again."}), 404

    cond_tag = conduit_row.get("Cond_Tag") or f"CONDUIT_{conduit_ident}"

    if project_number and seq_num is not None:
        # IC.EDC.S011 Interconnect-Diagram (IDP) drawing number: <project>-D.NN (category D,
        # no scope-of-work number, 2-digit index). seq_num is this conduit's project-sequential
        # position among the conduit sheets.
        output_filename = wdp_writer.conduit_drawing_no(project_number, int(seq_num)) + ".dwg"
    else:
        safe_tag = "".join(c for c in str(cond_tag) if c.isalnum() or c in "-_.")
        output_filename = f"{safe_tag}.dwg"

    output_path = os.path.join(output_folder, output_filename)

    fill_rows    = svc_parser.get_fill_rows(fill_index, cond_tag)
    loop_list    = svc_parser.build_loop_list(fill_rows)

    # NEC conduit-fill check: surface an over-fill conduit as a warning in the generation
    # log too (the frontend also warns before generating). Non-blocking -- notify only.
    from ..services import conduit_fill
    _fill_report = conduit_fill.evaluate(conduit_row, fill_rows)
    _fill_warning = _fill_report["message"] if _fill_report else None

    # Cross-conduit REF. DWG notes: for a 4-wire instrument whose power/TSP counterpart
    # lives in ANOTHER conduit, tag its loop with loop["ref_dwg"] = the counterpart's
    # drawing number. Purely additive — only 4W-split instruments get a key; every other
    # drawing is untouched. autocad_bridge places the NOTES flag + callout when present.
    svc_note_refs.annotate_instrument_refs(
        loop_list, fill_index, conduit_index, project_number, file_suffix, cond_tag)

    # Re-derive this conduit's Fill## slots from the RAW fill rows so the fill table
    # is always current, even if the frontend's cached conduit_index was parsed by an
    # older backend (stale Fill## values such as NONE/N/A). Raw fill rows are
    # backend-version-agnostic, so this self-heals without a re-parse. Strip any sent
    # Fill## keys first so _derive_fill_slots regenerates them.
    for _fk in [k for k in list(conduit_row.keys()) if str(k).startswith("Fill")]:
        del conduit_row[_fk]
    svc_mapper._derive_fill_slots(conduit_row, fill_rows)

    conduit_data = svc_parser.build_conduit_data(conduit_row)

    # Resolve ref doc rows for this conduit from ConduitIndex col L ("Ref Documents")
    ref_doc_names = [
        n.strip()
        for n in str(conduit_row.get("Ref_DocNames") or "").split(",")
        if n.strip()
    ]
    ref_doc_rows = [ref_docs[n] for n in ref_doc_names if n in ref_docs]

    # Resolve deviation notes for this conduit from ConduitIndex col K ("Deviations
    # Notes"): a comma-separated list of numbers. Each number maps to its note text
    # via the Ref Documents #→text lookup. Keep the ACTUAL number, in selection order,
    # no renumbering. dev_rows is a list of (number, text) pairs.
    def _norm_dev(n):
        try:
            return str(int(float(str(n).strip())))
        except (ValueError, TypeError):
            return str(n).strip()

    dev_nums = [
        n.strip()
        for n in str(conduit_row.get("Dev_Nums") or "").split(",")
        if n.strip()
    ]
    dev_rows = []
    for n in dev_nums:
        key = _norm_dev(n)
        if key in deviation_notes:
            dev_rows.append((key, deviation_notes[key]))
        elif n in deviation_notes:
            dev_rows.append((n, deviation_notes[n]))

    # Block heights (from BlockIndex) drive grid spacing; captured at parse time.
    block_heights = svc_parser.get_block_heights()

    # ── Paginate: a conduit whose stack would drop within 1" of the 1_BORDER bottom
    # is split across continuation sheets (start / middle / end). Pagination is pure
    # geometry (no AutoCAD), computed once over the whole conductor list.
    chunks = autocad_bridge.paginate_loops(loop_list, block_heights)

    base, ext = os.path.splitext(output_path)          # ext == ".dwg"
    n_sheets = len(chunks)
    if project_number and seq_num is not None:
        # Continuation sheets take the NEXT consecutive Interconnect-Diagram numbers (D.15 ->
        # D.16 -> D.17), matching IC.EDC.S011, instead of a "-N" suffix on the base name.
        # seq_num is this conduit's project-sequential start (already offset for earlier
        # conduits' continuations), so sheet k is D.(seq_num + k).
        sheet_paths = [
            os.path.join(output_folder,
                         wdp_writer.conduit_drawing_no(project_number, int(seq_num) + k) + ext)
            for k in range(n_sheets)
        ]
    else:
        # Unnumbered fallback (no project number / seq): keep the "-N" continuation suffix.
        sheet_paths = [output_path if k == 0 else f"{base}-{k + 1}{ext}"
                       for k in range(n_sheets)]

    # Sheet numbering leads with the GENERAL sheets (cover / index page(s) / legend) when
    # we're assembling a full project, so conduit sheet numbers continue after them; otherwise
    # conduit sheets number from 1. project_general_offset accounts for a multi-page drawing
    # index (a big project's index spills onto continuation sheets, pushing the conduits down),
    # and is computed identically here and in the finalize pass so the numbers agree.
    general_offset = wdp_writer.project_general_offset(conduit_index) if make_project else 0

    out_paths, all_warnings, errors, validations = [], [], [], []
    if _fill_warning:
        all_warnings.append(_fill_warning)   # NEC over-fill notice (non-blocking)
    if n_sheets > 1:
        # Continuation notice: this conduit didn't fit on one sheet, so it's split across
        # continuation sheets. Surface it in the generation log so the engineer knows.
        all_warnings.append(
            f"Conduit '{cond_tag}' has a continuation: it spans {n_sheets} sheets "
            f"({', '.join(os.path.basename(pp) for pp in sheet_paths)})."
        )
    for k, (a, b) in enumerate(chunks):
        loops_k = loop_list[a:b]

        if n_sheets == 1:
            cdata_k, state = conduit_data, None
        elif k == 0:
            # First sheet carries the full conduit schedule (all conductors).
            cdata_k, state = conduit_data, autocad_bridge.CONT_STATE_START
        else:
            # Continuation sheets carry NO conduit schedule -- the full fill table lives
            # only on the first sheet. Build conduit_data from a Fill-stripped copy and do
            # NOT re-derive: build_conduit_data stops at the first missing Fill01_Type, so
            # cdata_k gets the conduit identity (name/size/type, source/dest names) but zero
            # Fill## slots. Each sheet is a fresh template copy whose Conduit block Fill##
            # attributes all default to blank, so writing none leaves the schedule empty.
            # (Previously this re-derived a PARTIAL schedule from just this sheet's rows,
            # which is exactly the stray-context-on-continuation-pages bug.)
            cond_copy = {key: val for key, val in conduit_row.items()
                         if not str(key).startswith("Fill")}
            cdata_k = svc_parser.build_conduit_data(cond_copy)
            state = (autocad_bridge.CONT_STATE_END if k == n_sheets - 1
                     else autocad_bridge.CONT_STATE_MIDDLE)

        # Continuation cross-references show the adjacent sheet's DRAWING NUMBER (its
        # filename stem, e.g. 73.1159-15e -- no ".dwg"), so the "continued from / continued
        # on" comment reads as a drawing number, matching the title block's DRAWING_NO.
        prev_name = os.path.splitext(os.path.basename(sheet_paths[k - 1]))[0] if k > 0 else ""
        next_name = os.path.splitext(os.path.basename(sheet_paths[k + 1]))[0] if k < n_sheets - 1 else ""

        # DRAWING_NO on the title block = this sheet's own drawing number, i.e. its filename
        # stem (e.g. 73.1159-15e). SHEET = its running position in the whole deliverable:
        # the cover / index page(s) / legend take 1..general_offset, so conduit sheets continue
        # after them (no restart at 1). general_offset is 0 when not making a project.
        drawing_no  = os.path.splitext(os.path.basename(sheet_paths[k]))[0]
        conduit_seq = (int(seq_num) + k) if seq_num is not None else None   # 1-based across conduits + continuations
        sheet_number = (general_offset + conduit_seq) if conduit_seq is not None else None

        r = autocad_bridge.generate_dwg(
            cdata_k, loops_k, sheet_paths[k],
            ref_doc_rows=ref_doc_rows,          # supporting-docs table on every sheet
            dev_rows=dev_rows,                  # deviation notes (#→text) on every sheet
            block_heights=block_heights,
            cont_state=state, cont_prev=prev_name, cont_next=next_name,
            project_desc=project_desc, sheet_number=sheet_number, drawing_no=drawing_no)

        out_paths.append(sheet_paths[k])
        all_warnings += (r.get("warnings") or [])
        validations.append(r.get("validation"))
        if not r.get("success"):
            errors.append(f"sheet {k + 1}/{n_sheets}: {r.get('error')}")

    result = {
        "success":      not errors,
        "output_path":  out_paths[0] if out_paths else None,
        "output_paths": out_paths,
        "sheets":       n_sheets,
        "warnings":     all_warnings,
        "error":        "; ".join(errors) if errors else None,
        "validation":   validations if n_sheets > 1 else (validations[0] if validations else None),
    }

    # Remember this conduit's Drawing Properties Description 1/2/3 (Conduit name /
    # Source Name 1 / Destination Name 1), its Conduit tag, AND the project number it was
    # generated under, for every sheet just written -- so the .wdp writer, which rebuilds
    # its drawing list from scratch from whatever .dwgs are on disk, can still label this
    # conduit's drawing(s) on a LATER /generate call for a different conduit, and can tell
    # them apart both from an unrelated stray .dwg AND from an EARLIER project that reused
    # the same output folder and happened to share a conduit tag (project_number is what
    # tells those two apart -- cond_tag alone isn't enough).
    #
    # ONLY in make-project mode: this JSON sidecar (_idp_dwg_descriptions.json) is consumed
    # solely by /finalize-project's .wdp / drawing-index assembly. In plain (non-project) mode
    # there is no finalize, so we write NOTHING but the .dwg itself -- no sidecar, no .wdp.
    if not errors and out_paths and make_project:
        wdp_writer.record_dwg_descriptions(
            output_folder,
            [os.path.basename(p) for p in out_paths],
            desc1=conduit_data.get("Cdt_Name"),
            desc2=conduit_data.get("Src_Name01"),
            desc3=conduit_data.get("Dst_Name01"),
            cond_tag=cond_tag,
            project_number=project_number,
        )

    # Project files (.wdp/.aepx) are written ONLY when "Make it a project" is on -- and then
    # just ONCE, via /finalize-project after every conduit is generated (full-project assembly
    # re-opens G1-G3 over COM, so it must never run per conduit). In plain (non-project) mode we
    # deliberately write NO .wdp: the user just wants the drawing(s) they generated, nothing else.

    return jsonify(result)


# â”€â”€ Finalize a project (run ONCE after all conduits are generated) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/finalize-project", methods=["POST"])
def finalize_project():
    """Assemble the full AIC project ONCE, after every conduit has been generated.

    This is the work that used to ride on every per-conduit /generate call: copy the GENERAL
    template sheets (cover / drawing index / symbols legend) in, fill their title blocks over
    COM, populate the DRAWING INDEX table(s) with every drawing in the project (spilling onto
    continuation index sheets when the list is too big for one sheet), and write the sectioned
    .wdp/.aepx. The frontend calls this exactly once at the end of Generate All / Generate
    from list (and after a standalone single-conduit generate).

    Request body (JSON):
    {
        "output_folder":  str,
        "project_number": str,
        "project_desc":   {...},   # Project Description sheet -> title block project lines
        "conduit_index":  [ {...} ],  # for numbering + index rows (Seq_Start / Sheet_Count)
        "file_suffix":    "e"
    }
    Response: { ok, warnings, wdp_path, aepx_path, index_pages, error }
    """
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "JSON body required"}), 400

    output_folder = body.get("output_folder", "")
    project_number = str(body.get("project_number") or "").strip()
    project_desc = body.get("project_desc", {}) or {}
    conduit_index = body.get("conduit_index", []) or []
    file_suffix = body.get("file_suffix")
    file_suffix = "e" if file_suffix is None else str(file_suffix).strip()

    if not output_folder:
        return jsonify({"error": "No output folder is set."}), 400
    if not project_number:
        return jsonify({"error": "A project number is required to finalize a project."}), 400

    warnings = []
    try:
        # 1. Copy the GENERAL template sheets in, named per IC.EDC.S011 (<proj>-G.01 cover,
        # -G.02.. drawing-index page(s), legend last) + title-block map + drawing template.
        # Returns the authoritative sheet plan (correct numbers even with index continuation
        # pages). ensure_project_sheets creates every index page the plan calls for.
        plan = wdp_writer.ensure_project_sheets(output_folder, project_number, conduit_index)
        index_pages = sum(1 for g in plan if g["is_index"])

        # 3. Fill the NON-index general title blocks (cover + legend). The index pages' title
        # blocks are written by populate_drawing_index below (one open per page, table + TB).
        non_index = [
            (os.path.join(output_folder, g["name"]), g["drawing_no"], g["sheet_number"])
            for g in plan if not g["is_index"]
            and os.path.exists(os.path.join(output_folder, g["name"]))
        ]
        if non_index:
            warnings += autocad_bridge.fill_general_titleblocks(non_index, project_desc)

        # 4. Build the full drawing-index row list and split it across the index page(s), then
        # populate each index sheet's table (and its title block).
        rows, _pages = wdp_writer.build_index_rows(
            output_folder, project_number, conduit_index, file_suffix)
        cap = wdp_writer.INDEX_ROW_CAPACITY
        index_plan = [g for g in plan if g["is_index"]]
        index_pages_payload = []
        for i, g in enumerate(index_plan):
            slice_rows = rows[i * cap:(i + 1) * cap]
            index_pages_payload.append({
                "path": os.path.join(output_folder, g["name"]),
                "drawing_no": g["drawing_no"],
                "sheet_number": g["sheet_number"],
                "rows": slice_rows,
            })
        warnings += autocad_bridge.populate_drawing_index(index_pages_payload, project_desc)

        # 5. Write the sectioned .wdp + matching .aepx (GENERAL incl. index continuation pages,
        # then INTERCONNECTION DIAGRAMS = every current-workbook conduit drawing).
        valid_cond_tags = [
            str(r.get("Cond_Tag")).strip()
            for r in conduit_index
            if str(r.get("Cond_Tag") or "").strip()
        ]
        wdp_path = wdp_writer.write_full_project_wdp(
            output_folder, project_number, project_info=project_desc,
            valid_cond_tags=valid_cond_tags, conduit_index=conduit_index)
        aepx_path = wdp_writer.write_project_aepx(
            output_folder, project_number, valid_cond_tags=valid_cond_tags)

        return jsonify({
            "ok": True,
            "success": True,
            "warnings": warnings,
            "wdp_path": wdp_path,
            "aepx_path": aepx_path,
            "index_pages": index_pages,
            "error": None,
        })
    except Exception as e:
        return jsonify({"ok": False, "success": False, "error": str(e),
                        "warnings": warnings}), 500


# -- Regenerate the Drawing Index straight from a .wdp --------------------------

@idp_gen_bp.route("/reindex-drawing-index", methods=["POST"])
def reindex_drawing_index():
    """Rebuild the DRAWING INDEX table directly from an AutoCAD Electrical project (.wdp),
    independent of the loaded workbook -- so drawings added manually in ACADE are picked up.

    Pops a native file picker for the .wdp (unless a wdp_path is supplied), reads every drawing
    in the project (filename -> DRAWING NO., subsection -> SECTION, Drawing Properties
    Description 1/2/3 -> DRAWING DESCRIPTION, project order -> SHEET NO.), then writes those rows
    into the project's DRAWING INDEX sheet(s) -- the General sheet(s) whose Description 1 is
    'DRAWING INDEX'. It reads the .wdp as-is and does NOT rewrite it, so manual additions stay.

    Body (JSON, optional): { "wdp_path": "C:\\...\\<project>.wdp" }
    Returns { ok, warnings, drawings, index_pages, wdp_path } | { cancelled } | { error }.
    """
    body = request.get_json(silent=True) or {}
    wdp_path = (body.get("wdp_path") or "").strip()

    if not wdp_path:
        try:
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes("-topmost", True)
            wdp_path = filedialog.askopenfilename(
                title="Select the AutoCAD Electrical project (.wdp)",
                filetypes=[("AutoCAD Electrical project", "*.wdp"), ("All files", "*.*")],
            )
            root.destroy()
        except Exception as e:
            return jsonify({"ok": False, "error": f"Could not open the file dialog: {e}"})

    if not wdp_path:
        return jsonify({"ok": False, "cancelled": True})
    if not os.path.isfile(wdp_path):
        return jsonify({"ok": False, "error": f"File not found: {wdp_path}"}), 400

    try:
        drawings = wdp_writer.parse_wdp(wdp_path)
        if not drawings:
            return jsonify({"ok": False, "error": "No drawings found in that .wdp."})

        output_folder = os.path.dirname(wdp_path)
        # Every drawing becomes a row, in project order (SHEET NO. = running position).
        rows = [
            {"sheet_no": i, "drawing_no": d["drawing_no"], "section": d["section"],
             "description": d["description"]}
            for i, d in enumerate(drawings, start=1)
        ]

        # The DRAWING INDEX sheet(s) = drawings whose Description 1 marks them as the index.
        marker = wdp_writer.INDEX_DESC_MARKER.upper()
        index_dwgs = [d for d in drawings if marker in (d["description"] or "").upper()]
        if not index_dwgs:
            return jsonify({"ok": False, "error": (
                "No DRAWING INDEX sheet found in the project. The index sheet must be a General "
                "sheet whose Drawing Properties Description 1 is 'DRAWING INDEX'. Add/label one "
                "and try again.")})

        cap = wdp_writer.INDEX_ROW_CAPACITY
        sheet_no_by_dno = {r["drawing_no"]: r["sheet_no"] for r in rows}
        warnings, payload, n = [], [], len(index_dwgs)
        for i, d in enumerate(index_dwgs):
            slice_rows = rows[i * cap:] if i == n - 1 else rows[i * cap:(i + 1) * cap]
            if i == n - 1 and len(slice_rows) > cap:
                need = -(-len(rows) // cap)          # ceil: index pages the list actually needs
                warnings.append(
                    f"{len(rows)} drawings but only {n} DRAWING INDEX sheet(s) available -- the "
                    f"last index sheet holds {len(slice_rows)} rows and may overflow its border. "
                    f"Add {max(0, need - n)} more index sheet(s) (or regenerate the project) so "
                    f"the list fits.")
            payload.append({
                "path": os.path.join(output_folder, d["filename"]),
                "drawing_no": d["drawing_no"],
                "sheet_number": sheet_no_by_dno.get(d["drawing_no"]),
                "rows": slice_rows,
            })

        warnings += autocad_bridge.populate_drawing_index(payload, None)
        return jsonify({"ok": True, "success": True, "warnings": warnings,
                        "drawings": len(drawings), "index_pages": n, "wdp_path": wdp_path})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# â”€â”€ Wire Labels â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/wire-labels", methods=["POST"])
def wire_labels():
    """
    Generate a wire-labels Excel file from fill_index.

    Request body (JSON):
    {
        "fill_index": [ {...row dicts...} ],
        "filename":   str   (optional, used to derive the download name)
    }
    """
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "JSON body required"}), 400

    fill_index = body.get("fill_index", [])
    source_filename = body.get("filename", "IDP_Workbook.xlsx")
    stem = os.path.splitext(source_filename)[0]
    download_name = f"{stem}_WireLabels.xlsx"

    try:
        xl_bytes = svc_parser.build_wire_labels(fill_index)
        return send_file(
            io.BytesIO(xl_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Conduit fill % report (Excel) ────────────────────────────────────────────

@idp_gen_bp.route("/fill-report", methods=["POST"])
def fill_report():
    """Build an Excel report of every conduit's NEC Chapter 9 fill % and save it via a
    native Save-As dialog (a browser download doesn't work inside the LISA webview).

    Body: { "conduit_index": [...], "fill_index": [...], "filename": str, "default_dir": str }
    One row per conduit: raw fill %, % of the NEC-allowable limit, the conductor/cable
    breakdown, and whether it's over (over-fill rows shaded red).
    Returns { ok: true, path } | { ok: false, cancelled: true } | { ok: false, error }.
    """
    body = request.get_json(silent=True) or {}
    conduit_index = body.get("conduit_index", [])
    fill_index = body.get("fill_index", [])
    stem = os.path.splitext(body.get("filename", "IDP_Workbook.xlsx"))[0]
    download_name = f"{stem}_ConduitFill.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from ..services import conduit_fill

        by_tag = {}
        for r in fill_index:
            t = str(r.get("Cond_Tag") or "").strip()
            if t:
                by_tag.setdefault(t, []).append(r)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conduit Fill (NEC Ch.9)"
        headers = ["Conduit", "Type", "Size", "Type assumed?", "Conductors",
                   "Cables (by OD)", "Fill area (sq in)", "Conduit area (sq in)",
                   "Fill %", "NEC limit %", "% of limit", "Over limit?", "Notes"]
        ws.append(headers)
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        over_fill = PatternFill("solid", fgColor="F4CCCC")   # light red for over rows

        for c in conduit_index:
            tag = str(c.get("Cond_Tag") or "").strip()
            rep = conduit_fill.evaluate(c, by_tag.get(tag, [])) if tag else None
            if rep is None:
                ws.append([tag, c.get("Cond_Type"), c.get("Cond_Size"), "", "", "",
                           "", "", "n/a", "", "", "", "Not evaluable (no usable conduit type/size)"])
                continue
            notes = []
            if rep.get("assumed_type"):
                notes.append(f"type not specified - assumed {rep['conduit_type']}")
            if rep.get("skipped"):
                notes.append(f"{rep['skipped']} item(s) skipped (no gauge/OD) - actual fill higher")
            ws.append([
                tag, rep["conduit_type"], f"{rep['conduit_size']}\"",
                "Yes" if rep.get("assumed_type") else "",
                rep["conductors"], rep["cables"],
                rep["fill_area"], rep["conduit_area"],
                rep["fill_pct"] / 100.0, rep["allowed_pct"] / 100.0,
                (rep["of_limit_pct"] / 100.0) if rep.get("of_limit_pct") is not None else "",
                "OVER" if rep["over"] else "OK",
                "; ".join(notes),
            ])
            row = ws[ws.max_row]
            row[8].number_format = "0.0%"     # Fill %
            row[9].number_format = "0%"       # NEC limit %
            if isinstance(row[10].value, float):
                row[10].number_format = "0%"  # % of limit
            for cell in row:
                cell.border = border
            if rep["over"]:
                for cell in row:
                    cell.fill = over_fill

        widths = [16, 8, 8, 12, 11, 12, 15, 17, 9, 11, 10, 10, 46]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        xl_bytes = buf.getvalue()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    # Save via a native Save-As dialog (browser download doesn't work in the LISA webview).
    default_dir = body.get("default_dir") or ""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", True)
        path = filedialog.asksaveasfilename(
            title="Save Conduit Fill Report",
            initialfile=download_name,
            initialdir=default_dir or None,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        root.destroy()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    if not path:
        return jsonify({"ok": False, "cancelled": True})
    try:
        with open(path, "wb") as fh:
            fh.write(xl_bytes)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not write the file: {e}"}), 500
    return jsonify({"ok": True, "path": path})


# ── Wire-label print export (Excel, parity with IDPWireLabelPrintExcel.lsp) ──────

@idp_gen_bp.route("/wire-label-print", methods=["POST"])
def wire_label_print():
    """Build the wire-label PRINT workbook (grouped by size+label with Qty, Standard/Other
    worksheets, >14-char highlight, %%C/%C -> Ø) and save it via a native Save-As dialog
    (a browser download doesn't work inside the LISA webview).

    Body: { "fill_index": [...], "filename": str, "default_dir": str }
    Returns { ok: true, path } | { ok: false, cancelled: true } | { ok: false, error }.
    """
    body = request.get_json(silent=True) or {}
    fill_index = body.get("fill_index", [])
    stem = os.path.splitext(body.get("filename", "IDP_Workbook.xlsx"))[0]
    download_name = f"{stem}_WireLabelPrint.xlsx"
    try:
        xl_bytes = svc_parser.build_wire_label_print(fill_index)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    default_dir = body.get("default_dir") or ""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", True)
        path = filedialog.asksaveasfilename(
            title="Save Wire Label Print Report",
            initialfile=download_name,
            initialdir=default_dir or None,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        root.destroy()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    if not path:
        return jsonify({"ok": False, "cancelled": True})
    try:
        with open(path, "wb") as fh:
            fh.write(xl_bytes)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not write the file: {e}"}), 500
    return jsonify({"ok": True, "path": path})


# â”€â”€ Download (re-export Excel) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@idp_gen_bp.route("/download", methods=["POST"])
def download():
    """
    Re-export the current workbook state back to Excel.

    Request body (JSON):
    {
        "original_b64":  str,
        "conduit_index": [...],
        "fill_index":    [...],
        "filename":      str
    }
    """
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "JSON body required"}), 400

    original_b64  = body.get("original_b64")
    conduit_index = body.get("conduit_index", [])
    fill_index    = body.get("fill_index", [])
    filename      = body.get("filename", "IDP_Workbook.xlsx")

    if not original_b64:
        return jsonify({"error": "'original_b64' is required"}), 400

    try:
        original_bytes = base64.b64decode(original_b64)
        wb_bytes = svc_parser.write_workbook(original_bytes, filename, conduit_index, fill_index)
        return send_file(
            io.BytesIO(wb_bytes),
            mimetype=(
                "application/vnd.ms-excel.sheet.macroEnabled.12"
                if filename.lower().endswith(".xlsm")
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
